"""
검증 결과 엑셀 생성 모듈
원본 생산계획서 이미지를 테이블로 추출한 뒤,
교차검증 결과(일치/미입고/초과/부족)를 행별 색상으로 표시한 엑셀을 생성합니다.
"""

from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from modules.table_extractor import extract_table_from_image
from modules.erp_parser import normalize_color_code


# 스타일 상수
HEADER_FILL = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="맑은 고딕", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="A4B0BE"),
    right=Side(style="thin", color="A4B0BE"),
    top=Side(style="thin", color="A4B0BE"),
    bottom=Side(style="thin", color="A4B0BE"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

STATUS_FILLS = {
    "일치": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "초과": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
    "부족": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "미입고": PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
}
STATUS_FONTS = {
    "일치": Font(name="맑은 고딕", size=10, color="006100"),
    "초과": Font(name="맑은 고딕", size=10, color="9C5700"),
    "부족": Font(name="맑은 고딕", size=10, bold=True, color="9C0006"),
    "미입고": Font(name="맑은 고딕", size=10, bold=True, color="9C0006"),
}


def generate_verified_excel(
    image_bytes: bytes,
    file_name: str,
    result_df: pd.DataFrame,
    api_key: str,
) -> bytes:
    """
    원본 이미지를 테이블로 추출한 뒤 검증 결과를 색상으로 표시한 엑셀을 생성합니다.

    Returns:
        엑셀 파일 바이트
    """
    # 1. 이미지에서 테이블 추출
    table_data = extract_table_from_image(image_bytes, file_name, api_key)
    headers = table_data["headers"]
    rows = table_data["rows"]

    if not rows:
        raise ValueError("이미지에서 추출된 데이터가 없습니다.")

    # 2. 검증 결과 lookup 생성
    result_lookup = {}
    for _, row in result_df.iterrows():
        code = normalize_color_code(str(row["색상코드"]))
        result_lookup[code] = {
            "상태": row["상태"],
            "계획수량": int(row["계획수량"]),
            "입고수량": int(row["입고수량"]),
        }

    # 3. 각 행에서 품목코드를 찾아 상태 매칭
    row_statuses = []
    for row_data in rows:
        status = None
        for cell in row_data:
            if isinstance(cell, str) and len(cell) >= 5:
                code = normalize_color_code(cell)
                if code in result_lookup:
                    status = result_lookup[code]["상태"]
                    break
        row_statuses.append(status)

    # 4. 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "검증 결과"

    # 헤더 + 상태/입고수량 컬럼 추가
    all_headers = headers + ["상태", "입고수량"]
    for col_idx, header in enumerate(all_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 데이터 행
    for row_idx, (row_data, status) in enumerate(zip(rows, row_statuses), 2):
        # 원본 데이터
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = CENTER if isinstance(value, (int, float)) else LEFT

            # 상태에 따라 행 색상 적용
            if status and status in STATUS_FILLS:
                cell.fill = STATUS_FILLS[status]

        # 상태 컬럼
        status_col = len(headers) + 1
        status_cell = ws.cell(row=row_idx, column=status_col, value=status or "")
        status_cell.border = THIN_BORDER
        status_cell.alignment = CENTER
        if status and status in STATUS_FILLS:
            status_cell.fill = STATUS_FILLS[status]
            status_cell.font = STATUS_FONTS.get(status, DATA_FONT)

        # 입고수량 컬럼
        qty_col = len(headers) + 2
        qty_value = ""
        for cell_val in row_data:
            if isinstance(cell_val, str) and len(cell_val) >= 5:
                code = normalize_color_code(cell_val)
                if code in result_lookup:
                    qty_value = result_lookup[code]["입고수량"]
                    break
        qty_cell = ws.cell(row=row_idx, column=qty_col, value=qty_value)
        qty_cell.border = THIN_BORDER
        qty_cell.alignment = CENTER
        qty_cell.font = DATA_FONT
        if status and status in STATUS_FILLS:
            qty_cell.fill = STATUS_FILLS[status]

    # 열 너비 자동 조정
    for col_idx in range(1, len(all_headers) + 1):
        max_len = len(str(all_headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len * 1.5 + 4, 50)

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_incoming_plan_excel(plan_df: pd.DataFrame) -> bytes:
    """
    입고 예정 품목만 추려서 깔끔한 엑셀을 생성합니다.
    (신규 > 0인 항목만)

    Returns:
        엑셀 파일 바이트
    """
    incoming = plan_df[plan_df["신규"] > 0][["색상코드", "제조사", "신규"]].copy()
    incoming.columns = ["품목코드", "제조사", "입고예정수량"]
    incoming = incoming.reset_index(drop=True)

    if incoming.empty:
        raise ValueError("입고 예정 품목이 없습니다.")

    wb = Workbook()
    ws = wb.active
    ws.title = "입고 예정"

    headers = ["No.", "품목코드", "제조사", "입고예정수량"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(incoming.iterrows(), 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).font = DATA_FONT
        ws.cell(row=row_idx, column=1).border = THIN_BORDER
        ws.cell(row=row_idx, column=1).alignment = CENTER

        ws.cell(row=row_idx, column=2, value=row["품목코드"]).font = DATA_FONT
        ws.cell(row=row_idx, column=2).border = THIN_BORDER
        ws.cell(row=row_idx, column=2).alignment = CENTER

        ws.cell(row=row_idx, column=3, value=row["제조사"]).font = DATA_FONT
        ws.cell(row=row_idx, column=3).border = THIN_BORDER
        ws.cell(row=row_idx, column=3).alignment = CENTER

        qty_cell = ws.cell(row=row_idx, column=4, value=int(row["입고예정수량"]))
        qty_cell.font = Font(name="맑은 고딕", size=10, bold=True)
        qty_cell.border = THIN_BORDER
        qty_cell.alignment = CENTER
        qty_cell.number_format = "#,##0"

    # 합계 행
    sum_row = len(incoming) + 2
    sum_label = ws.cell(row=sum_row, column=1, value="합계")
    sum_label.font = Font(name="맑은 고딕", size=10, bold=True)
    sum_label.alignment = CENTER
    sum_label.border = THIN_BORDER
    sum_fill = PatternFill(start_color="DFE4EA", end_color="DFE4EA", fill_type="solid")
    for col in range(1, 5):
        ws.cell(row=sum_row, column=col).fill = sum_fill
        ws.cell(row=sum_row, column=col).border = THIN_BORDER
    ws.cell(row=sum_row, column=4).font = Font(name="맑은 고딕", size=10, bold=True)
    ws.cell(row=sum_row, column=4).alignment = CENTER
    col_letter = get_column_letter(4)
    ws.cell(row=sum_row, column=4, value=f"=SUM({col_letter}2:{col_letter}{sum_row - 1})")
    ws.cell(row=sum_row, column=4).number_format = "#,##0"

    # 열 너비
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
