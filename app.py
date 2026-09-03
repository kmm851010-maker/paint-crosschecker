# =====================================================================
# Project: paint-crosschecker
# Copyright (c) 2026 kmm851010-maker. All rights reserved.
# Unauthorized copying, modification, or distribution is strictly prohibited.
# =====================================================================
"""
스마트 공정/자재 관리 시스템 - Streamlit Dashboard
사이드바 메뉴 기반 다기능 웹 애플리케이션
"""

import os
import datetime
import time
import hmac
import hashlib
import base64

import pandas as pd
import streamlit as st
from dotenv import load_dotenv

load_dotenv()

st.set_page_config(
    page_title="KG스틸 업무도우미",
    page_icon="assets/favicon.png",
    layout="wide",
)

# ──────────────────────────────────────
# 세션 토큰 (HMAC 서명 - 서버 재시작 후에도 유효)
# ──────────────────────────────────────
_SESSION_TTL = 5 * 60 * 60  # 5시간

def _get_secret() -> bytes:
    return st.secrets.get("session_secret", "kg-steel-default-secret-2024").encode()

def _create_token(username: str) -> str:
    ts = str(int(time.time()))
    # rstrip("=") - URL에서 = 패딩 문제 방지
    payload = base64.urlsafe_b64encode(f"{username}:{ts}".encode()).decode().rstrip("=")
    sig = hmac.new(_get_secret(), payload.encode(), hashlib.sha256).hexdigest()[:24]
    return f"{payload}.{sig}"

def _validate_token(token: str):
    try:
        payload, sig = token.rsplit(".", 1)
        expected = hmac.new(_get_secret(), payload.encode(), hashlib.sha256).hexdigest()[:24]
        if not hmac.compare_digest(sig, expected):
            return None
        # 패딩 복원 후 디코드
        padded = payload + "=" * ((-len(payload)) % 4)
        decoded = base64.urlsafe_b64decode(padded.encode()).decode()
        username, ts = decoded.rsplit(":", 1)
        if time.time() - int(ts) > _SESSION_TTL:
            return None
        return username
    except Exception:
        return None

# ──────────────────────────────────────
# 로그인 처리
# ──────────────────────────────────────
def _check_login():
    auth_cfg = st.secrets.get("auth", {})
    users = auth_cfg.get("users", {})
    # users 없으면 단일 password 모드
    single_pw = auth_cfg.get("password", "")

    # URL 토큰으로 세션 복원 (새로고침 시 로그인 유지)
    token = st.query_params.get("t", "")
    if token:
        uname = _validate_token(token)
        if uname:
            st.session_state["authenticated"] = True
            st.session_state["username"] = uname
            st.session_state["_token"] = token
            return True
        else:
            # 만료된 토큰 제거
            st.query_params.pop("t", None)

    if st.session_state.get("authenticated"):
        return True

    import base64
    try:
        with open("assets/kg.jpg", "rb") as _f:
            _logo_b64 = base64.b64encode(_f.read()).decode()
        _logo_html = f'<img src="data:image/jpeg;base64,{_logo_b64}" style="width:90px;border-radius:12px;margin-bottom:12px;">'
    except Exception:
        _logo_html = '<div style="font-size:48px;margin-bottom:8px;">⚙️</div>'

    st.markdown(f"""
    <style>
        [data-testid="stSidebar"] {{ display: none; }}
        .block-container {{ max-width: 420px !important; margin: 60px auto !important; padding: 0 16px !important; }}
        .login-header {{ background: linear-gradient(135deg, #4B2D8E 0%, #6B3FA0 100%);
                        border-radius: 16px 16px 0 0; padding: 36px 24px 28px;
                        text-align: center; margin-bottom: 0; }}
        .login-title {{ color: #fff; font-size: 22px; font-weight: 700; margin-bottom: 4px; }}
        .login-sub {{ color: #D4C5F0; font-size: 13px; }}
        .login-body {{ background: #fff; border-radius: 0 0 16px 16px;
                      padding: 28px 24px 24px; box-shadow: 0 8px 32px rgba(75,45,142,0.18); }}
    </style>
    <div class="login-header">
        {_logo_html}
        <div class="login-title">KG스틸 업무도우미</div>
        <div class="login-sub">당진생산지원팀 전용 시스템</div>
    </div>
    <div class="login-body"></div>
    """, unsafe_allow_html=True)

    with st.form("login_form"):
        uid = st.text_input("아이디", placeholder="아이디를 입력하세요")
        pw = st.text_input("비밀번호", type="password", placeholder="비밀번호를 입력하세요")
        submitted = st.form_submit_button("🔐  로그인", use_container_width=True, type="primary")

    if submitted:
        # users dict 있으면 개인 ID/PW, 없으면 단일 PW
        if users:
            if users.get(uid.strip().lower()) == pw:
                uname = uid.strip().lower()
                token = _create_token(uname)
                st.session_state["authenticated"] = True
                st.session_state["username"] = uname
                st.session_state["_token"] = token
                st.query_params["t"] = token
                st.rerun()
            else:
                st.error("아이디 또는 비밀번호가 올바르지 않습니다.")
        else:
            if pw == single_pw:
                uname = uid or "user"
                token = _create_token(uname)
                st.session_state["authenticated"] = True
                st.session_state["username"] = uname
                st.session_state["_token"] = token
                st.query_params["t"] = token
                st.rerun()
            else:
                st.error("비밀번호가 올바르지 않습니다.")
    return False

if not _check_login():
    st.stop()

api_key = os.getenv("ANTHROPIC_API_KEY", "")

# KG스틸 브랜드 CSS
st.markdown("""
<style>
    /* 사이드바 */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #4B2D8E 0%, #3A2270 100%);
    }
    [data-testid="stSidebar"] * {
        color: #FFFFFF !important;
    }
    [data-testid="stSidebar"] .stCaption p {
        color: #D4C5F0 !important;
    }
    /* 사이드바 버튼 간격 제거 — 같은 섹션끼리 붙게 */
    [data-testid="stSidebar"] .stButton {
        margin-top: 1px !important;
        margin-bottom: 1px !important;
    }
    /* 사이드바 버튼: 항상 텍스트 보이게 */
    [data-testid="stSidebar"] .stButton > button {
        color: #FFFFFF !important;
        background: rgba(255, 255, 255, 0.08) !important;
        border: 1px solid rgba(255, 255, 255, 0.25) !important;
    }
    [data-testid="stSidebar"] .stButton > button:hover {
        background: rgba(255, 255, 255, 0.18) !important;
        border-color: rgba(255, 255, 255, 0.5) !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #F5A623 0%, #E8951A 100%) !important;
        border: none !important;
        color: #1A1A2E !important;
    }
    [data-testid="stSidebar"] .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #FFB84D 0%, #F5A623 100%) !important;
    }
    /* 메트릭 카드 */
    [data-testid="stMetric"] {
        background: #FFFFFF;
        border: 1px solid #E8E0F0;
        border-radius: 12px;
        padding: 12px;
        box-shadow: 0 2px 8px rgba(75,45,142,0.08);
    }
    /* 버튼 */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #4B2D8E 0%, #6B3FA0 100%);
        border: none;
        border-radius: 10px;
        font-weight: 700;
        color: white !important;
    }
    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #3A2270 0%, #5A3090 100%);
        color: white !important;
    }
    /* 다운로드 버튼 */
    .stDownloadButton > button {
        border: 2px solid #4B2D8E;
        border-radius: 10px;
        color: #4B2D8E !important;
        font-weight: 600;
    }
    .stDownloadButton > button[kind="primary"] {
        background: linear-gradient(135deg, #4B2D8E 0%, #6B3FA0 100%);
        color: white !important;
    }
    .stDownloadButton > button:hover {
        background: #4B2D8E;
        color: white !important;
    }
    /* 서브헤더 */
    h2, h3 {
        color: #4B2D8E !important;
        border-bottom: 2px solid #F5A623;
        padding-bottom: 6px;
    }
    /* 성공/경고 박스 */
    .stSuccess {
        border-left: 4px solid #2E7D32;
    }
    .stWarning {
        border-left: 4px solid #F5A623;
    }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────
# 사이드바
# ──────────────────────────────────────
st.sidebar.image("assets/kg.jpg", width=160)
_dept = st.secrets.get("company", {}).get("dept", "")
_team = st.secrets.get("company", {}).get("team", "")
st.sidebar.caption(f"{_dept}\n{_team} 업무도우미" if _dept else "KG스틸 업무도우미")
st.sidebar.markdown("---")

_VALID_PAGES = {"근태관리", "일일 작업 일지", "재고 현황", "입고 관리", "반품 관리"}
if st.session_state.get("page") not in _VALID_PAGES:
    st.session_state["page"] = "근태관리"

def _nav(label, key):
    is_active = st.session_state["page"] == key
    if st.sidebar.button(label, use_container_width=True, key=f"nav_{key}",
                         type="primary" if is_active else "secondary"):
        st.session_state["page"] = key
        st.rerun()

st.sidebar.markdown("**KG 근태관리**")
_nav("근태관리", "근태관리")
_nav("일일 작업 일지", "일일 작업 일지")
st.sidebar.markdown("**KG 재고관리**")
_nav("재고 현황", "재고 현황")
_nav("입고 관리", "입고 관리")
_nav("반품 관리", "반품 관리")

page = st.session_state["page"]

st.sidebar.markdown("---")
_uname = st.session_state.get("username", "")
if _uname:
    st.sidebar.caption(f"{_uname}")
if st.sidebar.button("🚪 로그아웃", use_container_width=True):
    st.query_params.clear()
    st.session_state.clear()
    st.rerun()
st.sidebar.markdown("---")
st.sidebar.markdown("**모바일 앱**")
st.sidebar.markdown(
    '<a href="https://expo.dev/artifacts/eas/w8QlmkgyJDLXqP-NOM7PfMjMOX4zkG1cc4CGWvT9tJo.apk" '
    'style="display:block;text-align:center;padding:10px;background:#F5A623;color:#1A1A2E;'
    'border-radius:8px;font-weight:700;text-decoration:none;">⬇️ KG OPS 설치</a>',
    unsafe_allow_html=True,
)



# ──────────────────────────────────────
# 근무 교대 스케줄 공통 헬퍼
# ──────────────────────────────────────
_CYCLE_20 = [
    ('B', 'C', 'D', 'A'), ('B', 'C', 'A', 'D'), ('B', 'C', 'A', 'D'),
    ('B', 'D', 'A', 'C'), ('B', 'D', 'A', 'C'), ('C', 'D', 'A', 'B'),
    ('C', 'D', 'B', 'A'), ('C', 'D', 'B', 'A'), ('C', 'A', 'B', 'D'),
    ('C', 'A', 'B', 'D'), ('D', 'A', 'B', 'C'), ('D', 'A', 'C', 'B'),
    ('D', 'A', 'C', 'B'), ('D', 'B', 'C', 'A'), ('D', 'B', 'C', 'A'),
    ('A', 'B', 'C', 'D'), ('A', 'B', 'D', 'C'), ('A', 'B', 'D', 'C'),
    ('A', 'C', 'D', 'B'), ('A', 'C', 'D', 'B'),
]
_BASE_DATE = datetime.date(2026, 3, 1)
_NIGHT_HOURS_BASE = {"1근": 0, "2근": 0.5, "3근": 7.5}

# ── 3조3교대 (3팀, 주중만, 3주 사이클) ──
_3S3_REF_MON = datetime.date(2026, 7, 27)
_3S3_SHIFTS = [
    {"A": "3근", "B": "1근", "C": "2근"},
    {"A": "2근", "B": "3근", "C": "1근"},
    {"A": "1근", "B": "2근", "C": "3근"},
]


def _shift_for_date_3s3(target_date, team):
    wd = target_date.weekday()
    if wd >= 5:
        return "휴무"
    mon = target_date - datetime.timedelta(days=wd)
    phase = ((mon - _3S3_REF_MON).days // 7) % 3
    return _3S3_SHIFTS[phase].get(team, "?")


# ── 2조2교대 (2팀, 주중만, 2주 사이클) ──
_2S2_REF_MON = datetime.date(2026, 7, 27)
_2S2_SHIFTS = [
    {"A": "2근", "B": "1근"},
    {"A": "1근", "B": "2근"},
]


def _shift_for_date_2s2(target_date, team):
    wd = target_date.weekday()
    if wd >= 5:
        return "휴무"
    mon = target_date - datetime.timedelta(days=wd)
    phase = ((mon - _2S2_REF_MON).days // 7) % 2
    return _2S2_SHIFTS[phase].get(team, "?")


# ── 4조2교대 (4팀, 연속 교대, 8일 사이클: 주주→휴휴→야야→휴휴) ──
_4S2_REF = datetime.date(2026, 7, 27)  # 기준일
_4S2_CYCLE = ["주간", "주간", "휴무", "휴무", "야간", "야간", "휴무", "휴무"]
_4S2_OFFSET = {"A": 1, "B": 5, "C": 3, "D": 7}  # A조 2026-08-31=야간, B조 2026-09-01=주간 검증완료


def _shift_for_date_4s2(target_date, team):
    days_since = (target_date - _4S2_REF).days
    offset = _4S2_OFFSET.get(team, 0)
    phase = (days_since + offset) % 8
    return _4S2_CYCLE[phase]


_KG_COMPANY_DAYS = {(9, 1): "창립기념일"}  # 매년 고정 회사 기념일 (월, 일)


def _get_holiday_name(date):
    """날짜의 공휴일/기념일 이름 반환 (없으면 '')"""
    c = _KG_COMPANY_DAYS.get((date.month, date.day))
    if c:
        return c
    try:
        import holidays as _hol
        kr = _hol.SouthKorea(years=date.year)
        if date in kr:
            return kr[date]
    except Exception:
        pass
    return ""


def _shift_for_date(target_date, members):
    idx = (target_date - _BASE_DATE).days % 20
    s1, s2, s3, off = _CYCLE_20[idx]
    prev_off = _CYCLE_20[(target_date - datetime.timedelta(days=1) - _BASE_DATE).days % 20][3]
    off_type = "주휴휴무" if prev_off == off else "교대휴무"
    return {
        "1근_조": s1, "1근_근무자": members.get(s1, s1),
        "2근_조": s2, "2근_근무자": members.get(s2, s2),
        "3근_조": s3, "3근_근무자": members.get(s3, s3),
        "휴무_조": off, "휴무_근무자": members.get(off, off),
        "휴무_구분": off_type, "is_2person": False,
        "leave_person": "", "leave_type": "",
    }


def _apply_leaves_stat(shift, target_date, leaves):
    result = shift.copy()
    for lv in leaves:
        try:
            start = datetime.date.fromisoformat(lv["start"][:10])
            end = datetime.date.fromisoformat(lv["end"][:10])
        except Exception:
            continue
        if start <= target_date <= end:
            absent, ltype = lv["name"], lv["type"]
            if result["1근_근무자"] == absent:
                result.update({"is_2person": True, "leave_person": absent, "leave_type": ltype,
                                "주간_근무자": result["2근_근무자"], "야간_근무자": result["3근_근무자"],
                                "주간_조": result["2근_조"], "야간_조": result["3근_조"]})
            elif result["2근_근무자"] == absent:
                result.update({"is_2person": True, "leave_person": absent, "leave_type": ltype,
                                "주간_근무자": result["1근_근무자"], "야간_근무자": result["3근_근무자"],
                                "주간_조": result["1근_조"], "야간_조": result["3근_조"]})
            elif result["3근_근무자"] == absent:
                result.update({"is_2person": True, "leave_person": absent, "leave_type": ltype,
                                "주간_근무자": result["1근_근무자"], "야간_근무자": result["2근_근무자"],
                                "주간_조": result["1근_조"], "야간_조": result["2근_조"]})
            break
    return result


# ══════════════════════════════════════
# 메뉴 1: 입고관리
# ══════════════════════════════════════
def page_cross_check():
    from modules.erp_parser import process_erp_file
    from modules.matcher import cross_check
    from modules.excel_generator import generate_report
    from utils.formatter import style_result_table, format_summary

    st.title("입고관리")

    # ── 2단 레이아웃 ──
    col_plan, col_erp = st.columns(2)

    # ── Dialog 1: 입고예정 품목 리스트 ──
    @st.dialog("입고예정 품목 리스트", width="large")
    def _incoming_items_dialog():
        _pln_df   = st.session_state.get("cc_plan_df")
        _pln_name = st.session_state.get("cc_plan_name", "")
        if _pln_df is None or "신규" not in _pln_df.columns:
            st.info("입고예정 품목이 없습니다.")
            return
        from modules.excel_converter import generate_incoming_plan_excel
        _inc_excel = None
        try: _inc_excel = generate_incoming_plan_excel(_pln_df)
        except Exception: pass

        _cols2 = ["색상코드", "제조사", "신규"]
        _has_rem = "비고" in _pln_df.columns
        if _has_rem: _cols2.append("비고")
        _inc_df = _pln_df[_pln_df["신규"] > 0][_cols2].copy()
        _cn = ["품목코드", "제조사", "입고예정수량"]
        if _has_rem: _cn.append("비고")
        _inc_df.columns = _cn
        _inc_df = _inc_df.reset_index(drop=True)
        _inc_df.insert(_inc_df.columns.get_loc("입고예정수량"), "기입고수량", 0)
        _inc_df["기입고수량"]   = _inc_df["기입고수량"].astype(int)
        _inc_df["입고예정수량"] = _inc_df["입고예정수량"].astype(int)
        _inc_df["비고"] = _inc_df.get("비고", pd.Series([""] * len(_inc_df))).fillna("").astype(str)

        if _inc_excel:
            st.download_button("입고예정 엑셀 다운로드", data=_inc_excel,
                file_name="incoming_plan.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        _ed_key = f"inc_dlg_ed_{_pln_name}_{len(_inc_df)}"
        _ed = st.data_editor(
            _inc_df[["품목코드", "제조사", "기입고수량", "입고예정수량", "비고"]],
            column_config={
                "품목코드":     st.column_config.TextColumn(disabled=True),
                "제조사":       st.column_config.TextColumn(disabled=True),
                "기입고수량":   st.column_config.NumberColumn("기입고수량", min_value=0, step=1),
                "입고예정수량": st.column_config.NumberColumn(disabled=True),
                "비고":         st.column_config.TextColumn("비고"),
            },
            use_container_width=True, hide_index=False, num_rows="fixed", key=_ed_key,
            height=min(len(_inc_df) * 35 + 40, 700),
        )
        _tp = int(_ed["입고예정수량"].sum())
        _ta = int(_ed["기입고수량"].fillna(0).sum())
        st.success(f"총 {len(_ed)}개 품목 | 입고예정: {_tp}개 | 기입고: {_ta}개 | 잔여: {_tp - _ta}개")

    # ── Dialog 2: 생산계획서 전체 변환 결과 ──
    @st.dialog("생산계획서 전체 변환 결과", width="large")
    def _plan_conversion_dialog():
        st.markdown(
            "<style>"
            "[data-testid='stDialog']>div{max-width:96vw!important;width:96vw!important;}"
            "[data-testid='stDialog'] [data-testid='stDataFrame']{overflow-x:visible!important;}"
            "</style>",
            unsafe_allow_html=True,
        )
        _pln_df    = st.session_state.get("cc_plan_df")
        _pln_bytes = st.session_state.get("cc_plan_bytes")
        _pln_name  = st.session_state.get("cc_plan_name", "")
        _tbl_data  = st.session_state.get("cc_table_data")

        if not (_tbl_data and _tbl_data.get("headers") and _tbl_data.get("rows")):
            st.info("변환 결과가 없습니다.")
            return

        _headers  = _tbl_data["headers"]
        _rows_raw = _tbl_data["rows"]
        _unique_h, _seen = [], {}
        for _h in _headers:
            _hs = str(_h) if _h else ""
            if _hs in _seen:
                _seen[_hs] += 1
                _unique_h.append(f"{_hs}_{_seen[_hs]}")
            else:
                _seen[_hs] = 0
                _unique_h.append(_hs)
        _padded = [list(r[:len(_unique_h)]) + [""] * max(0, len(_unique_h) - len(r)) for r in _rows_raw]
        _exp_h, _exp_r = [], [[] for _ in _padded]
        for _i, _h in enumerate(_unique_h):
            _exp_h.append(_h)
            for _j, _rw in enumerate(_padded):
                _exp_r[_j].append(_rw[_i] if _i < len(_rw) else "")
            if "신규" in str(_h):
                _in = str(_h).replace("신규", "입고")
                _b, _c = _in, 1
                while _in in _exp_h:
                    _in = f"{_b}_{_c}"; _c += 1
                _exp_h.append(_in)
                for _j in range(len(_padded)):
                    _exp_r[_j].append("")

        _full_df = pd.DataFrame(_exp_r, columns=_exp_h)
        def _ts(x):
            if x is None or str(x) in ("nan", "None", "NaN"): return ""
            if isinstance(x, float): return str(int(x)) if x == int(x) else str(x)
            return x if isinstance(x, str) else str(x)
        _full_df = _full_df.apply(lambda col: col.map(_ts))

        from modules.excel_converter import convert_to_excel
        _full_excel = convert_to_excel(list(_full_df.columns), _full_df.fillna("").values.tolist())

        # 상단 소형 버튼 행 (이모티콘 없음)
        _dc1, _dc2, _dc3 = st.columns([2, 2, 3])
        with _dc1:
            st.download_button("전체 엑셀 다운로드", data=_full_excel,
                file_name="plan_full_table.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with _dc2:
            if _pln_df is not None and "신규" in _pln_df.columns:
                if st.button("입고예정리스트 확인", use_container_width=True, key="btn_open_incoming_dlg"):
                    st.session_state["cc_open_incoming"] = True
                    st.rerun()

        st.caption(f"변환 결과 ({len(_rows_raw)}행 × {len(_exp_h)}열)")
        _tbl_key = f"dlg_tbl_{_pln_name}_{len(_rows_raw)}"
        _tbl_h   = len(_rows_raw) * 35 + 42  # 전체 행 높이 = 내부 수직 스크롤 없음
        _is_img  = _pln_name.lower().rsplit(".", 1)[-1] in ("jpg", "jpeg", "png", "webp")
        if _is_img and _pln_bytes:
            _ci, _ct = st.columns(2)
            with _ci:
                st.caption("원본 이미지")
                st.image(_pln_bytes, use_container_width=True)
            with _ct:
                _full_df = st.data_editor(_full_df, use_container_width=True, hide_index=True,
                    num_rows="fixed", key=_tbl_key, height=_tbl_h)
        else:
            _full_df = st.data_editor(_full_df, use_container_width=True, hide_index=True,
                num_rows="fixed", key=_tbl_key, height=_tbl_h)

    # 입고예정리스트 팝업: 변환결과 팝업 내 버튼 → rerun → 여기서 단독 트리거
    if st.session_state.get("cc_open_incoming"):
        st.session_state.pop("cc_open_incoming")
        _incoming_items_dialog()

    # ── 왼쪽: 생산계획서 ──
    with col_plan:
        st.subheader("① 생산계획서")
        _fu_col, _btn_col = st.columns([3, 1])
        with _fu_col:
            plan_file = st.file_uploader(
                "생산계획서 업로드 (이미지 · 엑셀 · PDF · Word)",
                type=["jpg", "jpeg", "png", "webp", "xlsx", "xls", "csv", "pdf", "docx"],
                key="plan_upload",
                help="인쇄물 사진, 엑셀, PDF, Word(.docx) 모두 지원",
            )
        with _btn_col:
            st.write("")
            st.write("")
            if plan_file and "cc_plan_df" not in st.session_state:
                if st.button("추출", type="primary", use_container_width=True, key="btn_extract_plan"):
                    if not api_key:
                        st.error("API Key가 설정되지 않았습니다.")
                        st.stop()
                    plan_bytes = plan_file.getvalue()
                    plan_fname = plan_file.name
                    with st.spinner("생산계획서 분석 중..."):
                        try:
                            from modules.vision_ocr import extract_production_plan
                            result = extract_production_plan(plan_bytes, plan_fname, api_key)
                            plan_df = pd.DataFrame(result["items"])
                            st.session_state["cc_table_data"] = result.get("table_data")
                        except Exception as e:
                            st.error(f"생산계획서 분석 실패: {e}")
                            st.stop()
                    st.session_state["cc_plan_df"] = plan_df
                    st.session_state["cc_plan_bytes"] = plan_bytes
                    st.session_state["cc_plan_name"] = plan_fname
                    st.rerun()
            elif "cc_plan_df" in st.session_state:
                table_data = st.session_state.get("cc_table_data")
                if table_data and table_data.get("headers") and table_data.get("rows"):
                    if st.button("변환결과", use_container_width=True, key="btn_show_plan_dlg"):
                        _plan_conversion_dialog()

        if "cc_plan_df" in st.session_state:
            plan_df   = st.session_state["cc_plan_df"]
            plan_bytes = st.session_state.get("cc_plan_bytes")
            plan_name  = st.session_state.get("cc_plan_name", "")
            table_data = st.session_state.get("cc_table_data")

            # ERP 입고 반영 결과 계산 (소스 데이터 변경 시에만 재계산, 사용자 편집 보존)
            if "cc_result_df" in st.session_state:
                from utils.helpers import auto_correct_code, is_valid_item_code
                _result_df = st.session_state["cc_result_df"]
                erp_qty_map = {
                    str(r.get("색상코드", "")).strip(): int(r.get("입고수량", 0) or 0)
                    for _, r in _result_df.iterrows()
                    if str(r.get("색상코드", "")).strip()
                }
                # 소스 해시: ERP 입고 데이터 + 생산계획 테이블이 바뀔 때만 재계산
                _src_hash = str(hash(str(sorted(erp_qty_map.items())) + str(table_data)))[-10:]
                _skip_recompute = (
                    "cc_filled_df" in st.session_state and
                    st.session_state.get("cc_filled_src_hash") == _src_hash
                )
                if table_data and table_data.get("headers") and table_data.get("rows") and not _skip_recompute:
                    _h2, _r2 = table_data["headers"], table_data["rows"]
                    _u2, _s2 = [], {}
                    for _h in _h2:
                        _hs = str(_h) if _h else ""
                        if _hs in _s2: _s2[_hs] += 1; _u2.append(f"{_hs}_{_s2[_hs]}")
                        else: _s2[_hs] = 0; _u2.append(_hs)
                    _pd2 = [list(r[:len(_u2)]) + [""] * max(0, len(_u2) - len(r)) for r in _r2]
                    _eh2, _er2 = [], [[] for _ in _pd2]
                    for _i2, _hh in enumerate(_u2):
                        _eh2.append(_hh)
                        for _j2, _rr in enumerate(_pd2):
                            _er2[_j2].append(_rr[_i2] if _i2 < len(_rr) else "")
                        if "신규" in str(_hh):
                            _inn = str(_hh).replace("신규", "입고")
                            _bb, _cc = _inn, 1
                            while _inn in _eh2: _inn = f"{_bb}_{_cc}"; _cc += 1
                            _eh2.append(_inn)
                            for _j2 in range(len(_pd2)): _er2[_j2].append("")
                    def _ts2(x):
                        if x is None or str(x) in ("nan", "None", "NaN"): return ""
                        if isinstance(x, float): return str(int(x)) if x == int(x) else str(x)
                        return x if isinstance(x, str) else str(x)
                    filled_df = pd.DataFrame(_er2, columns=_eh2).apply(lambda col: col.map(_ts2))
                    exp_headers = _eh2
                    신규_col_indices = [i for i, h in enumerate(exp_headers) if "신규" in str(h)]
                    for row_idx, row_vals in filled_df.iterrows():
                        for ni in 신규_col_indices:
                            if ni + 1 >= len(exp_headers): continue
                            inc_col = exp_headers[ni + 1]
                            if "입고" not in str(inc_col): continue
                            code_col_idx = ni - 3
                            if code_col_idx < 0: continue
                            raw_code = str(row_vals[exp_headers[code_col_idx]]).strip()
                            corrected = auto_correct_code(raw_code)
                            if is_valid_item_code(corrected) and corrected in erp_qty_map:
                                qty = erp_qty_map[corrected]
                                filled_df.at[row_idx, inc_col] = str(qty) if qty > 0 else ""
                    for _ni in 신규_col_indices:
                        if _ni + 1 >= len(exp_headers): continue
                        _inc_col = exp_headers[_ni + 1]
                        if "입고" not in str(_inc_col): continue
                        _신규_col = exp_headers[_ni]
                        _code_col_idx = _ni - 3
                        if _code_col_idx < 0: continue
                        _code_col = exp_headers[_code_col_idx]
                        _first_seen = {}
                        for _ridx, _rvals in filled_df.iterrows():
                            _raw = str(_rvals[_code_col]).strip()
                            _corr = auto_correct_code(_raw)
                            if not is_valid_item_code(_corr): continue
                            if _corr not in _first_seen:
                                _first_seen[_corr] = _ridx
                            else:
                                _fidx = _first_seen[_corr]
                                try: _fn = int(str(filled_df.at[_fidx, _신규_col]).strip() or "0")
                                except: _fn = 0
                                try: _dn = int(str(filled_df.at[_ridx, _신규_col]).strip() or "0")
                                except: _dn = 0
                                if _dn > 0:
                                    filled_df.at[_fidx, _신규_col] = str(_fn + _dn)
                                    filled_df.at[_ridx, _신규_col] = ""
                                filled_df.at[_ridx, _inc_col] = ""
                    st.session_state["cc_filled_df"] = filled_df
                    st.session_state["cc_filled_src_hash"] = _src_hash

    # ── 오른쪽: ERP 입고명세서 ──
    with col_erp:
        _erp_hdr, _erp_rst = st.columns([5, 1])
        with _erp_hdr:
            st.subheader("② ERP 입고명세서")
        with _erp_rst:
            if "cc_plan_df" in st.session_state:
                st.write("")
                if st.button("🔄", help="새로고침", use_container_width=True, key="btn_reset_cc"):
                    for _k in list(st.session_state.keys()):
                        if _k.startswith("cc_"):
                            del st.session_state[_k]
                    st.rerun()
        _fu2, _btn2 = st.columns([3, 1])
        with _fu2:
            erp_file = st.file_uploader(
                "ERP 입고명세서 업로드",
                type=["xlsx", "xls", "csv", "jpg", "jpeg", "png", "webp"],
                key="erp_upload",
                help="엑셀(.xlsx, .csv) 또는 화면 캡처 이미지",
            )
            if erp_file:
                ext_erp = erp_file.name.lower().rsplit(".", 1)[-1]
                if ext_erp in ("jpg", "jpeg", "png", "webp"):
                    st.image(erp_file, caption="ERP 명세서", use_container_width=True)
        with _btn2:
            st.write("")
            st.write("")
            run_button = st.button(
                "교차검증",
                type="primary",
                use_container_width=True,
                disabled=not (erp_file and "cc_plan_df" in st.session_state),
                key="btn_run_crosscheck",
            )

        if run_button:
            if not api_key:
                st.error("API Key가 설정되지 않았습니다.")
                st.stop()
            plan_df = st.session_state["cc_plan_df"]
            with st.spinner("ERP 입고명세서 분석 중..."):
                try:
                    erp_bytes = erp_file.getvalue()
                    erp_df = process_erp_file(erp_bytes, erp_file.name, api_key)
                except Exception as e:
                    st.error(f"ERP 명세서 분석 실패: {e}")
                    st.stop()
            result_df = cross_check(plan_df, erp_df)
            st.session_state["cc_erp_df"] = erp_df
            st.session_state["cc_result_df"] = result_df
            st.rerun()

    # ── 교차검증 결과 (①② 컬럼 아래 전체 너비) ──
    if "cc_result_df" in st.session_state:
        erp_df    = st.session_state["cc_erp_df"]
        result_df = st.session_state["cc_result_df"]
        plan_bytes = st.session_state["cc_plan_bytes"]
        plan_name  = st.session_state["cc_plan_name"]

        if not result_df.empty:
            summary = format_summary(result_df)

        if not result_df.empty:
            # ── ERP 입고 반영 결과 ──
            if "cc_filled_df" in st.session_state:
                _filled = st.session_state["cc_filled_df"]
                st.markdown("---")
                _res_hdr, _dl_col = st.columns([5, 1])
                with _res_hdr:
                    st.subheader("ERP 입고 반영 결과")
                    st.caption("신규 옆 입고 칸에 ERP 실입고 수량이 자동 기입된 양식입니다. 🟥 미입고 · 🟩 일치 · 🟡 초과 · 🟠 일부입고")
                with _dl_col:
                    from modules.excel_converter import convert_erp_filled_to_excel
                    _erp_excel = convert_erp_filled_to_excel(_filled)
                    st.write("")
                    st.write("")
                    st.download_button(
                        label="ERP 입고반영 엑셀",
                        data=_erp_excel,
                        file_name="plan_erp_result.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )

                # 상태 컬럼 동적 계산 (신규/입고 비교)
                _disp_filled = _filled.copy()
                _orig_cols = list(_filled.columns)
                _insert_offset = 0
                for _fi, _fh in enumerate(_orig_cols):
                    if "신규" in str(_fh):
                        _iq_fi = _fi + 1
                        if _iq_fi < len(_orig_cols) and "입고" in str(_orig_cols[_iq_fi]):
                            _h_sq2, _h_iq2 = _fh, _orig_cols[_iq_fi]
                            def _row_st(_row, _hs=_h_sq2, _hi=_h_iq2):
                                try: _sq = int(str(_row[_hs]).strip() or "0")
                                except: _sq = 0
                                try: _iq = int(str(_row[_hi]).strip() or "0")
                                except: _iq = 0
                                if _sq == 0: return ""
                                if _iq == 0: return "🟥 미입고"
                                if _iq == _sq: return "🟩 일치"
                                if _iq > _sq: return "🟡 초과"
                                return "🟠 일부"
                            _st_nm = f"__st_{_fh}"
                            _disp_filled.insert(_fi + _insert_offset + 2, _st_nm, _disp_filled.apply(_row_st, axis=1))
                            _insert_offset += 1
                _col_cfg_filled = {}
                for _fcol in list(_disp_filled.columns):
                    _fcol_s = str(_fcol)
                    if _fcol_s.startswith("__st_"):
                        _col_cfg_filled[_fcol_s] = st.column_config.TextColumn("상태", disabled=True, width="small")
                    elif any(k in _fcol_s for k in ["재고", "신규", "입고"]):
                        _col_cfg_filled[_fcol_s] = st.column_config.TextColumn(_fcol_s)
                    else:
                        _col_cfg_filled[_fcol_s] = st.column_config.TextColumn(_fcol_s, disabled=True)
                _filled_hash = str(hash(str(_filled.values.tolist())))[-6:]
                _edited_filled = st.data_editor(
                    _disp_filled,
                    column_config=_col_cfg_filled,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    key=f"cc_filled_editor_{_filled_hash}",
                    height=len(_disp_filled) * 35 + 38,
                )
                _status_cols2 = [c for c in _edited_filled.columns if str(c).startswith("__st_")]
                _clean_edited = _edited_filled.drop(columns=_status_cols2)
                _old_filled_ref = st.session_state.get("cc_filled_df")
                _data_changed = (_old_filled_ref is None or
                                 not _clean_edited.equals(_old_filled_ref))
                st.session_state["cc_filled_df"] = _clean_edited
                if _data_changed:
                    st.rerun()

            # ── 확인필요 목록 ──
            if summary['reverse_count'] > 0:
                st.markdown("---")
                _rev_l, _rev_c, _rev_r = st.columns([1, 2, 1])
                with _rev_c:
                    st.markdown(
                        f"<div style='text-align:center;font-size:14px'>⚠️ <b>확인필요 {summary['reverse_count']}건</b>: "
                        "생산계획서에 없지만 ERP에 입고 기록이 있는 품목입니다.</div>",
                        unsafe_allow_html=True,
                    )
                    _rev_df = result_df[result_df["상태"].str.contains("확인필요", na=False)][["색상코드", "입고수량", "상태"]].copy()
                    _rev_df.insert(0, "선택", False)
                    _rev_key = f"cc_reverse_{st.session_state.get('cc_reverse_ver', 0)}"
                    _edited_rev = st.data_editor(
                        _rev_df,
                        column_config={
                            "선택":     st.column_config.CheckboxColumn("선택", help="확인 완료 후 체크"),
                            "색상코드": st.column_config.TextColumn(disabled=True),
                            "입고수량": st.column_config.NumberColumn(disabled=True),
                            "상태":     st.column_config.TextColumn(disabled=True),
                        },
                        use_container_width=True,
                        hide_index=True,
                        num_rows="fixed",
                        key=_rev_key,
                        height=len(_rev_df) * 35 + 38,
                    )
                    if st.button("선택 항목 제거", use_container_width=True, key="cc_reverse_del"):
                        _to_del = _edited_rev[_edited_rev["선택"] == True]["색상코드"].tolist()
                        if _to_del:
                            st.session_state["cc_result_df"] = result_df[~result_df["색상코드"].isin(_to_del)].reset_index(drop=True)
                            st.session_state["cc_reverse_ver"] = st.session_state.get("cc_reverse_ver", 0) + 1
                            st.rerun()
                        else:
                            st.info("제거할 항목을 선택하세요.")

            is_plan_image = plan_name.lower().rsplit(".", 1)[-1] in ("jpg", "jpeg", "png", "webp")
            if is_plan_image:
                st.markdown("---")
                if st.button("검증 결과 엑셀 생성 (원본 이미지 기반)", use_container_width=True):
                    with st.spinner("원본 이미지를 엑셀로 변환 + 검증 결과 표시 중..."):
                        try:
                            from modules.image_annotator import generate_verified_excel
                            verified_excel = generate_verified_excel(plan_bytes, plan_name, result_df, api_key)
                            st.session_state["cc_verified_excel"] = verified_excel
                        except Exception as e:
                            st.error(f"검증 엑셀 생성 실패: {e}")
                if st.session_state.get("cc_verified_excel"):
                    st.download_button(
                        label="검증 결과 엑셀 다운로드 (입고/미입고 색상 표시)",
                        data=st.session_state["cc_verified_excel"],
                        file_name="verified_result.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
        else:
            st.info("신규 요청 수량이 있는 항목이 없습니다.")


# ══════════════════════════════════════
# 메뉴 2: 캡처 이미지 → 엑셀 변환기
# ══════════════════════════════════════
def page_image_to_excel():
    from modules.table_extractor import extract_table_from_image
    from modules.excel_converter import convert_to_excel

    st.title("캡처 이미지 → 엑셀 파일 변환기")
    st.caption("ERP 화면 캡처, 표 이미지를 서식 적용된 엑셀(.xlsx)로 즉시 변환합니다.")

    capture_file = st.file_uploader(
        "변환할 캡처 이미지를 업로드하세요",
        type=["jpg", "jpeg", "png", "webp"],
        key="capture_upload",
        help="ERP 화면 캡처, 엑셀 스크린샷 등 표가 포함된 이미지",
    )

    if capture_file:
        st.image(capture_file, caption="업로드된 캡처 이미지", use_container_width=True)

    st.markdown("---")

    convert_button = st.button(
        "📊 엑셀 파일로 변환하기",
        type="primary",
        use_container_width=True,
        disabled=not capture_file,
    )

    if convert_button:
        if not api_key:
            st.error("API Key가 설정되지 않았습니다. 관리자에게 문의하세요.")
            st.stop()

        capture_bytes = capture_file.getvalue()

        with st.spinner("이미지에서 표 데이터 추출 중..."):
            try:
                table_data = extract_table_from_image(capture_bytes, capture_file.name, api_key)
            except Exception as e:
                st.error(f"테이블 추출 실패: {e}")
                st.stop()

        headers = table_data["headers"]
        rows = table_data["rows"]

        if not rows:
            st.warning("추출된 데이터가 없습니다.")
            st.stop()

        st.markdown("---")
        st.subheader("변환 결과")

        col_img, col_table = st.columns(2)
        with col_img:
            st.caption("원본 캡처 이미지")
            st.image(capture_bytes, use_container_width=True)
        with col_table:
            st.caption(f"추출된 데이터 ({len(rows)}행 × {len(headers)}열)")
            # 중복 컬럼명 처리
            unique_headers = []
            seen = {}
            for h in headers:
                if h in seen:
                    seen[h] += 1
                    unique_headers.append(f"{h}_{seen[h]}")
                else:
                    seen[h] = 0
                    unique_headers.append(h)
            # 행 길이를 헤더에 맞춤
            padded_rows = []
            for r in rows:
                if len(r) < len(unique_headers):
                    padded_rows.append(list(r) + [""] * (len(unique_headers) - len(r)))
                else:
                    padded_rows.append(list(r[:len(unique_headers)]))
            result_df = pd.DataFrame(padded_rows, columns=unique_headers)
            st.dataframe(result_df, use_container_width=True, hide_index=True)

        st.info(f"총 {len(rows)}개 행, {len(headers)}개 컬럼 추출 완료")

        st.markdown("---")
        with st.spinner("엑셀 파일 생성 중..."):
            excel_bytes = convert_to_excel(headers, rows)

        st.download_button(
            label="📥 변환된 엑셀(.xlsx) 다운로드",
            data=excel_bytes,
            file_name=f"converted_{capture_file.name.rsplit('.', 1)[0]}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
            use_container_width=True,
        )


# ══════════════════════════════════════
# 메뉴 3: 일일 작업일지 작성
# ══════════════════════════════════════
def page_work_log():
    import datetime
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 4조 3교대 로테이션 (KG스틸 20일 주기)
    MEMBERS = dict(st.secrets.get("members", {'A': '직원A', 'B': '직원B', 'C': '직원C', 'D': '직원D'}))
    CYCLE_20 = [
        ('B', 'C', 'D', 'A'), ('B', 'C', 'A', 'D'), ('B', 'C', 'A', 'D'),
        ('B', 'D', 'A', 'C'), ('B', 'D', 'A', 'C'), ('C', 'D', 'A', 'B'),
        ('C', 'D', 'B', 'A'), ('C', 'D', 'B', 'A'), ('C', 'A', 'B', 'D'),
        ('C', 'A', 'B', 'D'), ('D', 'A', 'B', 'C'), ('D', 'A', 'C', 'B'),
        ('D', 'A', 'C', 'B'), ('D', 'B', 'C', 'A'), ('D', 'B', 'C', 'A'),
        ('A', 'B', 'C', 'D'), ('A', 'B', 'D', 'C'), ('A', 'B', 'D', 'C'),
        ('A', 'C', 'D', 'B'), ('A', 'C', 'D', 'B'),
    ]
    BASE_DATE = datetime.date(2026, 3, 1)

    def get_shift_info(target_date):
        idx = (target_date - BASE_DATE).days % 20
        s1, s2, s3, off = CYCLE_20[idx]
        prev_idx = (target_date - datetime.timedelta(days=1) - BASE_DATE).days % 20
        prev_off = CYCLE_20[prev_idx][3]
        off_type = "주휴휴무" if prev_off == off else "교대휴무"
        return {
            "1근_조": s1, "1근_근무자": MEMBERS[s1],
            "2근_조": s2, "2근_근무자": MEMBERS[s2],
            "3근_조": s3, "3근_근무자": MEMBERS[s3],
            "휴무_조": off, "휴무_근무자": MEMBERS[off],
            "휴무_구분": off_type
        }

    def _fill_work_log_sheet(ws, selected_date, shift_data, work_items, safety_items, note_text):
        """워크시트에 일일 작업일지를 채운다 (10컬럼 A-J)."""
        font_title = Font(name="맑은 고딕", size=18, bold=True, underline="single")
        font_date  = Font(name="맑은 고딕", size=11, bold=True)
        font_sec   = Font(name="맑은 고딕", size=11, bold=True)
        font_hdr   = Font(name="맑은 고딕", size=9, bold=True)
        font_body  = Font(name="맑은 고딕", size=9)
        font_bold  = Font(name="맑은 고딕", size=9, bold=True)
        fill_gray  = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        _t = Side(style='thin', color='000000')
        thin_border = Border(left=_t, right=_t, top=_t, bottom=_t)
        align_c  = Alignment(horizontal="center", vertical="center")
        align_cw = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 열 너비 A~J
        for col, w in zip("ABCDEFGHIJ", [9, 5, 13, 9, 9, 9, 9, 9, 9, 10]):
            ws.column_dimensions[col].width = w

        # 제목
        ws.row_dimensions[2].height = 32
        ws.merge_cells('A2:J2')
        _doc_team = st.secrets.get("company", {}).get("team", "")
        ws['A2'] = f"{_doc_team} 일일 업무 보고" if _doc_team else "일일 업무 보고"
        ws['A2'].font, ws['A2'].alignment = font_title, align_c

        # 날짜
        ws.merge_cells('H4:J4')
        ws['H4'] = selected_date.strftime("%Y년 %m월 %d일")
        ws['H4'].font = font_date
        ws['H4'].alignment = Alignment(horizontal="right", vertical="center")

        # 1. 인원 현황
        ws['A5'] = "1. 인원 현황"
        ws['A5'].font = font_sec
        for pos, txt in [("A6","구분"),("B6","조"),("C6","근무시간"),("D6","근무자"),("E6","휴무자"),("F6","연장근무")]:
            ws[pos] = txt
            ws[pos].font, ws[pos].fill, ws[pos].alignment, ws[pos].border = font_hdr, fill_gray, align_c, thin_border
        ws.merge_cells('G6:J6')
        ws['G6'] = "비고"
        ws['G6'].font, ws['G6'].fill, ws['G6'].alignment, ws['G6'].border = font_hdr, fill_gray, align_c, thin_border
        for c in "HIJ":
            ws[f"{c}6"].fill, ws[f"{c}6"].border = fill_gray, thin_border

        if shift_data.get("is_2person"):
            rows_1 = [
                ("주간", shift_data.get("주간_조", shift_data.get("1근_조","")), "06:30 – 18:30",
                 shift_data.get("주간_근무자", shift_data.get("1근_근무자","")), "", shift_data.get("1근_연장",""), shift_data.get("1근_비고","")),
                ("야간", shift_data.get("야간_조", shift_data.get("2근_조","")), "18:30 – 06:30",
                 shift_data.get("야간_근무자", shift_data.get("2근_근무자","")), "", shift_data.get("2근_연장",""), shift_data.get("2근_비고","")),
                ("휴무", "", "", "", shift_data.get("3근_근무자",""), "", shift_data.get("3근_비고","")),
                ("휴무", shift_data.get("휴무_조",""), "", "", shift_data.get("휴무_근무자",""), "", shift_data.get("휴무_구분",""))
            ]
        else:
            rows_1 = [
                ("1근", shift_data.get("1근_조",""), "06:30 – 14:30", shift_data.get("1근_근무자",""), "", shift_data.get("1근_연장",""), shift_data.get("1근_비고","")),
                ("2근", shift_data.get("2근_조",""), "14:30 – 22:30", shift_data.get("2근_근무자",""), "", shift_data.get("2근_연장",""), shift_data.get("2근_비고","")),
                ("3근", shift_data.get("3근_조",""), "22:30 – 06:30", shift_data.get("3근_근무자",""), "", shift_data.get("3근_연장",""), shift_data.get("3근_비고","")),
                ("휴무", shift_data.get("휴무_조",""), "", "", shift_data.get("휴무_근무자",""), "", shift_data.get("휴무_구분",""))
            ]
        for idx, r in enumerate(rows_1, start=7):
            ws.row_dimensions[idx].height = 28
            for ci, val in zip("ABCDEF", r[:6]):
                ws[f"{ci}{idx}"] = val
                ws[f"{ci}{idx}"].font  = font_body
                ws[f"{ci}{idx}"].border = thin_border
                ws[f"{ci}{idx}"].alignment = align_cw if ci == "C" else align_c
            ws.merge_cells(f"G{idx}:J{idx}")
            ws[f"G{idx}"] = r[6]
            for c in "GHIJ":
                ws[f"{c}{idx}"].font, ws[f"{c}{idx}"].border = font_body, thin_border
            ws[f"G{idx}"].alignment = align_c

        # 2. 업무 현황
        ws['A12'] = "2. 업무 현황"
        ws['A12'].font = font_sec
        ws.merge_cells('A13:C13')
        ws['A13'] = "작업 내용"
        ws['A13'].font, ws['A13'].fill, ws['A13'].alignment, ws['A13'].border = font_hdr, fill_gray, align_c, thin_border
        for c in "BC":
            ws[f"{c}13"].fill, ws[f"{c}13"].border = fill_gray, thin_border
        for col_l, hdr in zip("DEFGHIJ", ["1근","2근","3근","주간","야간","합계","월합계"]):
            ws[f"{col_l}13"] = hdr
            ws[f"{col_l}13"].font, ws[f"{col_l}13"].fill, ws[f"{col_l}13"].alignment, ws[f"{col_l}13"].border = font_hdr, fill_gray, align_c, thin_border

        for idx, item in enumerate(work_items, start=14):
            ws.row_dimensions[idx].height = 20
            ws.merge_cells(f"A{idx}:C{idx}")
            ws[f"A{idx}"] = item["name"]
            ws[f"A{idx}"].font, ws[f"A{idx}"].alignment, ws[f"A{idx}"].border = font_bold, align_c, thin_border
            for c in "BC":
                ws[f"{c}{idx}"].border = thin_border
            ws[f"D{idx}"] = item.get("s1") or ""
            ws[f"E{idx}"] = item.get("s2") or ""
            ws[f"F{idx}"] = item.get("s3") or ""
            ws[f"G{idx}"] = item.get("day") or ""
            ws[f"H{idx}"] = item.get("night") or ""
            for c in "DEFGH":
                ws[f"{c}{idx}"].font, ws[f"{c}{idx}"].alignment, ws[f"{c}{idx}"].border = font_body, align_c, thin_border
            ws[f"I{idx}"] = f"=SUM(D{idx}:H{idx})"
            ws[f"I{idx}"].font, ws[f"I{idx}"].alignment, ws[f"I{idx}"].border = font_bold, align_c, thin_border
            ws[f"J{idx}"] = item.get("month_total", 0)
            ws[f"J{idx}"].font, ws[f"J{idx}"].alignment, ws[f"J{idx}"].border = font_body, align_c, thin_border

        # 3. 안전 관리 사항
        ws['A27'] = "3. 안전 관리 사항"
        ws['A27'].font = font_sec
        ws.merge_cells('A28:E28')
        for c in "ABCDE":
            ws[f"{c}28"].fill, ws[f"{c}28"].border = fill_gray, thin_border
        for pos, txt in [("F28","1근"),("G28","2근"),("H28","3근"),("I28","주간"),("J28","야간")]:
            ws[pos] = txt
            ws[pos].font, ws[pos].fill, ws[pos].alignment, ws[pos].border = font_hdr, fill_gray, align_c, thin_border

        align_safety = Alignment(horizontal="left", vertical="center", wrap_text=True)
        for idx, s_row in enumerate(safety_items, start=29):
            ws.row_dimensions[idx].height = 34
            ws.merge_cells(f"A{idx}:E{idx}")
            ws[f"A{idx}"] = s_row["text"]
            ws[f"A{idx}"].font, ws[f"A{idx}"].alignment, ws[f"A{idx}"].border = font_bold, align_safety, thin_border
            for c in "BCDE":
                ws[f"{c}{idx}"].border = thin_border
            for pc, key in [("F","s1"),("G","s2"),("H","s3"),("I","day"),("J","night")]:
                cell = ws[f"{pc}{idx}"]
                cell.value = "☑" if s_row.get(key) else "□"
                cell.font, cell.alignment, cell.border = Font(name="맑은 고딕", size=10), align_c, thin_border

        # 4. 특이 사항
        ws['A36'] = "4. 특이 사항"
        ws['A36'].font = font_sec
        ws.merge_cells('A37:J41')
        ws['A37'] = note_text
        ws['A37'].font = font_body
        ws['A37'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        for r in range(37, 42):
            for c_idx in range(1, 11):
                cl = get_column_letter(c_idx)
                ws[f"{cl}{r}"].border = Border(
                    top=_t if r == 37 else Side(),
                    bottom=_t if r == 41 else Side(),
                    left=_t if c_idx == 1 else Side(),
                    right=_t if c_idx == 10 else Side()
                )

    def generate_monthly_work_log_excel(up_to_date):
        """당월 1일~up_to_date까지 저장된 작업일지를 날짜별 탭으로 묶어 반환.
        Returns: (bytes, int) - Excel 바이트, 포함된 날짜 수
        """
        import calendar as _cal
        from utils.supabase_db import load_monthly_data
        year, month = up_to_date.year, up_to_date.month
        work_by_date, detail_by_date = load_monthly_data(year, month)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        last_day = min(up_to_date.day, _cal.monthrange(year, month)[1])
        added = 0
        for day in range(1, last_day + 1):
            date_obj = datetime.date(year, month, day)
            date_str = date_obj.strftime("%Y-%m-%d")
            if date_str not in work_by_date and date_str not in detail_by_date:
                continue
            detail = detail_by_date.get(date_str, {})
            ws_new = wb.create_sheet(title=f"{month}월{day}일")
            _fill_work_log_sheet(
                ws_new, date_obj,
                detail.get("shift", {}),
                work_by_date.get(date_str, []),
                detail.get("safety", []),
                detail.get("note", "")
            )
            added += 1

        if added == 0:
            ws_empty = wb.create_sheet(title="데이터없음")
            ws_empty['A1'] = "저장된 작업일지가 없습니다."

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue(), added

    # 휴가 목록 (sheets 연동 — 근무표 메뉴에서 등록)
    if "leave_list" not in st.session_state:
        try:
            from utils.supabase_db import load_leaves as _ll
            st.session_state["leave_list"] = _ll()
        except Exception:
            st.session_state["leave_list"] = []

    ALL_MEMBERS = list(MEMBERS.values())

    def apply_leaves(shift_data, target_date):
        """등록된 휴가를 근무 데이터에 반영. 2인 근무 시 주간/야간 체계로 전환."""
        result = shift_data.copy()
        result["is_2person"] = False
        result["leave_person"] = ""
        result["leave_type"] = ""

        for leave in st.session_state["leave_list"]:
            try:
                start = datetime.date.fromisoformat(leave["start"])
                end = datetime.date.fromisoformat(leave["end"])
            except Exception:
                continue  # 날짜 형식 오류 항목 무시
            if start <= target_date <= end:
                absent = leave["name"]
                ltype = leave["type"]

                # 누가 빠지는지 확인
                if result["1근_근무자"] == absent:
                    # 1근 휴가 → 2근→주간, 3근→야간
                    result["is_2person"] = True
                    result["leave_person"] = absent
                    result["leave_type"] = ltype
                    result["주간_근무자"] = result["2근_근무자"]
                    result["야간_근무자"] = result["3근_근무자"]
                    result["주간_조"] = result["2근_조"]
                    result["야간_조"] = result["3근_조"]
                elif result["2근_근무자"] == absent:
                    # 2근 휴가 → 1근→주간, 3근→야간
                    result["is_2person"] = True
                    result["leave_person"] = absent
                    result["leave_type"] = ltype
                    result["주간_근무자"] = result["1근_근무자"]
                    result["야간_근무자"] = result["3근_근무자"]
                    result["주간_조"] = result["1근_조"]
                    result["야간_조"] = result["3근_조"]
                elif result["3근_근무자"] == absent:
                    # 3근 휴가 → 1근→주간, 2근→야간
                    result["is_2person"] = True
                    result["leave_person"] = absent
                    result["leave_type"] = ltype
                    result["주간_근무자"] = result["1근_근무자"]
                    result["야간_근무자"] = result["2근_근무자"]
                    result["주간_조"] = result["1근_조"]
                    result["야간_조"] = result["2근_조"]
                break  # 한 날짜에 한 명만 휴가 가정

        return result

    # ── UI ──
    _title_team = st.secrets.get("company", {}).get("team", "")
    st.title(f"{_title_team} 일일 업무 보고 작성" if _title_team else "일일 업무 보고 작성")

    # 달력 크게 표시
    st.markdown("""
    <style>
        [data-testid="stDateInput"] > div { transform: scale(1.15); transform-origin: left top; }
        [data-testid="stDateInput"] input { font-size: 20px !important; font-weight: 700 !important; padding: 12px !important; }
    </style>
    """, unsafe_allow_html=True)

    _dcol1, _dcol2 = st.columns([5, 1])
    with _dcol1:
        st.subheader("작업 일자 선택")
    with _dcol2:
        st.markdown("<div style='padding-top:28px'>", unsafe_allow_html=True)
        if st.button("🔄 새로고침", key="wl_refresh", use_container_width=True, help="시트에서 데이터를 다시 불러옵니다"):
            _now_kst2 = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
            _today2 = (_now_kst2 - datetime.timedelta(hours=6, minutes=30)).date()
            for _k in list(st.session_state.keys()):
                if _k.startswith("wl_autoloaded_") or _k.startswith("wl_loaded_") or _k.startswith("wl_monthly_totals_"):
                    del st.session_state[_k]
            st.session_state.pop("wl_current_date", None)
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

    _now_kst = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
    # 06:30 이전이면 전날 일지로 취급 (하루 기준 시각 06:30)
    today_kst = (_now_kst - datetime.timedelta(hours=6, minutes=30)).date()
    selected_date = st.date_input("날짜를 클릭하세요", today_kst,
                                   min_value=datetime.date(2026, 1, 1),
                                   max_value=datetime.date(2100, 12, 31),
                                   label_visibility="collapsed")

    shift_auto = get_shift_info(selected_date)
    shift_auto = apply_leaves(shift_auto, selected_date)

    is_2person = shift_auto.get("is_2person", False)

    if is_2person:
        st.warning(
            f"⚠️ **{shift_auto['leave_person']}** {shift_auto['leave_type']} — "
            f"2인 근무 체계 (주간/야간 12시간) 자동 전환"
        )
        st.success(
            f"**{selected_date.strftime('%Y년 %m월 %d일')}** 2인 근무\n\n"
            f"주간(06:30-18:30): **{shift_auto['주간_조']}조 {shift_auto['주간_근무자']}** | "
            f"야간(18:30-06:30): **{shift_auto['야간_조']}조 {shift_auto['야간_근무자']}** | "
            f"휴가: **{shift_auto['leave_person']}** | "
            f"휴무: **{shift_auto['휴무_조']}조 {shift_auto['휴무_근무자']}**"
        )
    else:
        st.success(
            f"**{selected_date.strftime('%Y년 %m월 %d일')}** 근무 매칭 완료\n\n"
            f"1근: **{shift_auto['1근_조']}조 {shift_auto['1근_근무자']}** | "
            f"2근: **{shift_auto['2근_조']}조 {shift_auto['2근_근무자']}** | "
            f"3근: **{shift_auto['3근_조']}조 {shift_auto['3근_근무자']}** | "
            f"휴무: **{shift_auto['휴무_조']}조 {shift_auto['휴무_근무자']}**"
        )

    st.markdown("---")
    st.subheader("1. 인원 현황")

    # 0.5시간 단위 시간 옵션 (00:00 ~ 23:30)
    _TIME_OPTS = ["없음"] + [f"{h:02d}:{m:02d}" for h in range(0, 24) for m in (0, 30)]

    def _calc_ot(start_str, end_str):
        """시작~종료 시간 → (주간연장H, 야간연장H). 야간 기준: 22:00~06:00"""
        if start_str == "없음" or end_str == "없음":
            return 0.0, 0.0
        sh, sm = int(start_str[:2]), int(start_str[3:])
        eh, em = int(end_str[:2]), int(end_str[3:])
        start_min = sh * 60 + sm
        end_min = eh * 60 + em
        if end_min <= start_min:
            end_min += 24 * 60  # 자정 넘어가는 경우
        total_min = end_min - start_min
        if total_min <= 0:
            return 0.0, 0.0
        # 야간: 22:00(1320분) ~ 06:00 다음날(1800분)
        night_start, night_end = 22 * 60, 30 * 60
        overlap_start = max(start_min, night_start)
        overlap_end = min(end_min, night_end)
        night_min = max(0, overlap_end - overlap_start)
        day_min = total_min - night_min
        return round(day_min / 60 * 2) / 2, round(night_min / 60 * 2) / 2

    if is_2person:
        _leave_type_2p  = shift_auto.get("leave_type", "")
        _leave_person_2p = shift_auto.get("leave_person", "")
        _is_gonghu = _leave_type_2p == "공휴"
        if _is_gonghu:
            _default_note = "공휴일 휴일연장대근"
        elif _leave_person_2p and _leave_type_2p:
            _default_note = f"{_leave_person_2p} {_leave_type_2p}로 대근"
        else:
            _default_note = "대근"

        if _is_gonghu:
            st.info(
                f"🗓️ **공휴일 근무 체제** — {shift_auto.get('leave_person', '')} 공휴 처리 "
                f"→ 나머지 2인이 **휴일연장대근**으로 커버합니다."
            )

        c1, c2, c3 = st.columns(3)
        with c1:
            st.markdown("**주간** (06:30~18:30)")
            day_name = st.text_input("주간 근무자", value=shift_auto["주간_근무자"])
            day_day_ot, day_night_ot = 0.0, 0.0
            day_note = st.text_input("주간 비고", _default_note)
        with c2:
            st.markdown("**야간** (18:30~06:30)")
            night_name = st.text_input("야간 근무자", value=shift_auto["야간_근무자"])
            night_day_ot, night_night_ot = 0.0, 0.0
            night_note = st.text_input("야간 비고", _default_note)
        with c3:
            st.markdown("**휴무**")
            off_name = st.text_input("휴무자", value=shift_auto["휴무_근무자"])
            _off_opts_2p = ["교대휴무","주휴휴무","정기휴가","연차","특별휴가","명휴","생일휴가","공가","공상휴업","산재","휴직","대휴","교육","결근","조퇴","외출","청원휴가","공휴"]
            _off_idx_2p = _off_opts_2p.index(shift_auto["휴무_구분"]) if shift_auto["휴무_구분"] in _off_opts_2p else 0
            off_type = st.selectbox("휴무 구분", _off_opts_2p, index=_off_idx_2p)

        shift_data_final = {
            "1근_조": shift_auto["주간_조"], "1근_근무자": day_name,
            "1근_연장": 4 + day_day_ot + day_night_ot,
            "1근_주간연장": 4 + day_day_ot, "1근_야간연장": day_night_ot, "1근_비고": day_note,
            "2근_조": shift_auto["야간_조"], "2근_근무자": night_name,
            "2근_연장": 4 + night_day_ot + night_night_ot,
            "2근_주간연장": night_day_ot, "2근_야간연장": 4 + night_night_ot, "2근_비고": night_note,
            "3근_조": "", "3근_근무자": shift_auto["leave_person"], "3근_연장": 0, "3근_비고": shift_auto["leave_type"],
            "휴무_조": shift_auto["휴무_조"], "휴무_근무자": off_name, "휴무_구분": off_type,
            "is_2person": True,
        }
    else:
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.markdown("**1근** (06:30~14:30)")
            s1_name = st.text_input("1근 근무자", value=shift_auto["1근_근무자"])
            s1_day_ot, s1_night_ot = 0.0, 0.0
            s1_note = st.text_input("1근 비고", "")
        with c2:
            st.markdown("**2근** (14:30~22:30)")
            s2_name = st.text_input("2근 근무자", value=shift_auto["2근_근무자"])
            s2_day_ot, s2_night_ot = 0.0, 0.0
            s2_note = st.text_input("2근 비고", "")
        with c3:
            st.markdown("**3근** (22:30~06:30)")
            s3_name = st.text_input("3근 근무자", value=shift_auto["3근_근무자"])
            s3_day_ot, s3_night_ot = 0.0, 0.0
            s3_note = st.text_input("3근 비고", "")
        with c4:
            st.markdown("**휴무**")
            off_name = st.text_input("휴무자", value=shift_auto["휴무_근무자"])
            _off_opts_3p = ["교대휴무","주휴휴무","정기휴가","연차","특별휴가","명휴","생일휴가","공가","공상휴업","산재","휴직","대휴","교육","결근","조퇴","외출","청원휴가","공휴"]
            _off_idx_3p = _off_opts_3p.index(shift_auto["휴무_구분"]) if shift_auto["휴무_구분"] in _off_opts_3p else 0
            off_type = st.selectbox("휴무 구분", _off_opts_3p, index=_off_idx_3p)

        shift_data_final = {
            "1근_조": shift_auto["1근_조"], "1근_근무자": s1_name,
            "1근_연장": s1_day_ot + s1_night_ot, "1근_주간연장": s1_day_ot, "1근_야간연장": s1_night_ot, "1근_비고": s1_note,
            "2근_조": shift_auto["2근_조"], "2근_근무자": s2_name,
            "2근_연장": s2_day_ot + s2_night_ot, "2근_주간연장": s2_day_ot, "2근_야간연장": s2_night_ot, "2근_비고": s2_note,
            "3근_조": shift_auto["3근_조"], "3근_근무자": s3_name,
            "3근_연장": s3_day_ot + s3_night_ot, "3근_주간연장": s3_day_ot, "3근_야간연장": s3_night_ot, "3근_비고": s3_note,
            "휴무_조": shift_auto["휴무_조"], "휴무_근무자": off_name, "휴무_구분": off_type,
            "is_2person": False,
        }

    st.markdown("---")

    # 기존 데이터 불러오기
    # 휴가 데이터 로드 (최초 1회)
    if "leave_loaded" not in st.session_state:
        st.session_state["leave_loaded"] = True
        try:
            from utils.supabase_db import load_leaves
            saved_leaves = load_leaves()
            if saved_leaves:
                st.session_state["leave_list"] = saved_leaves
                st.rerun()
        except Exception:
            pass

    # ─── 작업일지 자동 로드 (날짜별 캐시) ───────────────────────────────
    _grid_key = f"wl_grid_{selected_date}"
    _fetched_key = f"wl_fetched_{selected_date}"

    if _grid_key not in st.session_state:
        if _fetched_key not in st.session_state:
            try:
                from utils.supabase_db import load_all
                _all_data = load_all(selected_date)
                _loaded_wi = _all_data.get("work_items") or {}
                st.session_state[f"wl_detail_{selected_date}"] = _all_data.get("detail") or {}
                if _loaded_wi:
                    st.toast(f"📂 {selected_date.strftime('%Y-%m-%d')} 저장 데이터 자동 불러옴")
            except Exception:
                _loaded_wi = {}
            st.session_state[_fetched_key] = True
        else:
            _loaded_wi = {}  # 새로 작성 후 재진입
    else:
        _loaded_wi = {}

    loaded_detail = st.session_state.get(f"wl_detail_{selected_date}") or {}

    # 새로 작성 버튼
    if _grid_key in st.session_state:
        if st.button('새로 작성 (저장 데이터 무시)', key='new_write'):
            st.session_state.pop(_grid_key, None)
            st.session_state.pop(f'wl_safety_{selected_date}', None)
            st.session_state.pop(f'wl_note_{selected_date}', None)
            st.rerun()

    import pandas as _pd

    item_names = [
        '페인트 하차 수량', '페인트 공급 수량', '재고 페인트 창고 입고',
        '신나 하차 수량', '신나 공급 수량', '크롬 공급 수량',
        '공드럼 운반 수량', '페보루 운반 수량', '페신너 운반 및 상차',
        '반품 , 불량 페인트 수량', '코터롤 운반 횟수', '필름 하차, 장소 이동 횟수',
        'AGV 입/출고 작업 수량',
    ]
    _all_shifts = ['1근', '2근', '3근', '주간', '야간']
    _shift_keys = ['s1', 's2', 's3', 'day', 'night']
    _dis_2p  = {'1근', '2근', '3근'}
    _dis_3sh = {'주간', '야간'}

    _mt_key = f'wl_monthly_totals_{selected_date}'
    if _mt_key not in st.session_state:
        try:
            from utils.supabase_db import get_monthly_totals
            st.session_state[_mt_key] = get_monthly_totals(selected_date)
        except Exception:
            st.session_state[_mt_key] = {}
    monthly_totals = st.session_state[_mt_key]

    # data_editor 공통 테두리 CSS
    _DE_CSS = (
        '<style>'
        'div[data-testid="stDataEditor"] > div:first-child {'
        '  border:2px solid #595959 !important;'
        '  border-top:none !important;'
        '  border-radius:0 0 4px 4px !important;'
        '}'
        '</style>'
    )
    st.markdown(_DE_CSS, unsafe_allow_html=True)

    # 2. 업무 현황
    st.markdown(
        '<div style="background:#4B2D8E;color:white;padding:6px 14px;'
        'font-size:13px;font-weight:bold;margin-top:8px;'
        'border-radius:4px 4px 0 0;border:2px solid #4B2D8E;">'
        '2. 업무 현황</div>',
        unsafe_allow_html=True,
    )

    # 연산식 평가 시 data_editor 키 버전을 올려 강제 재초기화
    _de_ver_key = f'wl_de_ver_{selected_date}'
    _de_ver = st.session_state.get(_de_ver_key, 0)

    if _grid_key in st.session_state:
        _df_grid = st.session_state[_grid_key]
    else:
        _rows = []
        for _nm in item_names:
            _itm = _loaded_wi.get(_nm, {})
            _row = {'작업 내용': _nm}
            for _sl, _lk in zip(_all_shifts, _shift_keys):
                _row[_sl] = str(int(_itm.get(_lk, 0) or 0))
            _ds = sum(int(_row[_sl]) for _sl in _all_shifts)
            _row['합계']   = str(_ds)
            _row['월합계'] = str(monthly_totals.get(_nm, 0) + _ds)
            _rows.append(_row)
        _df_grid = _pd.DataFrame(_rows)

    # Excel 양식 비율: 작업내용 185px, 근무열 66px, 합계 63px, 월합계 70px
    _W = {'item': 185, 'shift': 66, 'sum': 63, 'mon': 70}

    def _eval_cell(v):
        import re as _re
        s = str(v).strip() if v is not None else ''
        if not s or s in ('0', 'None', 'nan'):
            return 0
        if _re.match(r'^[\d\s\+\-\*\/\(\)\.]+$', s):
            try:
                return max(0, int(eval(s)))  # noqa: S307 – 숫자+연산자만 허용
            except Exception:
                pass
        try:
            return max(0, int(float(s)))
        except Exception:
            return 0

    _col_cfg_w = {
        '작업 내용': st.column_config.TextColumn('작업 내용', width=_W['item'], disabled=True),
        '합계':   st.column_config.TextColumn('합계',   width=_W['sum'], disabled=True),
        '월합계': st.column_config.TextColumn('월합계', width=_W['mon'], disabled=True),
    }
    for _sl in _all_shifts:
        _col_cfg_w[_sl] = st.column_config.TextColumn(
            _sl, width=_W['shift'],
            disabled=(is_2person and _sl in _dis_2p) or (not is_2person and _sl in _dis_3sh),
        )

    _edited_df = st.data_editor(
        _df_grid,
        key=f'de_{_grid_key}_v{_de_ver}',
        column_config=_col_cfg_w,
        column_order=['작업 내용'] + _all_shifts + ['합계', '월합계'],
        hide_index=True,
        use_container_width=True,
        num_rows='fixed',
        height=495,
    )

    _expr_evaluated = False
    _vals_changed = False
    for _i, _nm in enumerate(item_names):
        _row_vals = {}
        for _sl in _all_shifts:
            _raw = _edited_df.at[_i, _sl]
            _ev = _eval_cell(_raw)
            _raw_s = str(_raw).strip()
            if _raw_s not in ('', '0', 'None', 'nan', str(_ev)):
                _expr_evaluated = True
            if _raw_s != str(_df_grid.at[_i, _sl]).strip():
                _vals_changed = True
            _edited_df.at[_i, _sl] = str(_ev)
            _row_vals[_sl] = _ev
        _ds = sum(_row_vals.values())
        _edited_df.at[_i, '합계']   = str(_ds)
        _edited_df.at[_i, '월합계'] = str(monthly_totals.get(_nm, 0) + _ds)
    st.session_state[_grid_key] = _edited_df
    if _expr_evaluated:
        st.session_state[_de_ver_key] = _de_ver + 1
        st.rerun()
    elif _vals_changed:
        st.rerun()

    work_items_data = []
    for _i, _nm in enumerate(item_names):
        _wd = {'name': _nm, 'month_total': int(_edited_df.at[_i, '월합계'] or 0)}
        for _sl, _lk in zip(_all_shifts, _shift_keys):
            _v = int(_edited_df.at[_i, _sl] or 0)
            _wd[_lk] = _v if _v > 0 else None
        work_items_data.append(_wd)

    # 3. 안전 관리 사항
    st.markdown(
        '<div style="background:#4B2D8E;color:white;padding:6px 14px;'
        'font-size:13px;font-weight:bold;margin-top:18px;'
        'border-radius:4px 4px 0 0;border:2px solid #4B2D8E;">'
        '3. 안전 관리 사항</div>',
        unsafe_allow_html=True,
    )

    safety_questions = [
        '작업 계획에 따라 작업 절차를 준수 하였는가?',
        '안전장치(후방 경보장치 , 안전밸트 등) 기능의 이상 유무를 점검 하였는가?',
        '주행시 급출발 , 급정거 , 급선회를 하지 않았는가?',
        '화물 적재시 허용 하중을 초과하지 않았는가?',
        '작업장소에 적합한 제한 속도를 준수 하였는가?',
        '지게차 작업 안전 수칙에 위배 되는 작업을 하지 않았는가?',
    ]

    _safety_key = f'wl_safety_{selected_date}'
    if _safety_key not in st.session_state:
        _loaded_safety = (loaded_detail.get('safety') or []) if loaded_detail else []
        _srows = []
        for _qi, _q in enumerate(safety_questions):
            if _qi < len(_loaded_safety):
                _si = _loaded_safety[_qi]
                _srows.append({
                    '안전 관리 사항': _q,
                    '1근':  bool(_si.get('s1',    False)),
                    '2근':  bool(_si.get('s2',    False)),
                    '3근':  bool(_si.get('s3',    False)),
                    '주간': bool(_si.get('day',   False)),
                    '야간': bool(_si.get('night', False)),
                })
            else:
                _srows.append({
                    '안전 관리 사항': _q,
                    '1근': not is_2person, '2근': not is_2person, '3근': not is_2person,
                    '주간': is_2person,    '야간': is_2person,
                })
        st.session_state[_safety_key] = _pd.DataFrame(_srows)

    _safety_col_cfg = {
        '안전 관리 사항': st.column_config.TextColumn('안전 관리 사항', width=450, disabled=True),
        '1근':  st.column_config.CheckboxColumn('1근',  width=66),
        '2근':  st.column_config.CheckboxColumn('2근',  width=66),
        '3근':  st.column_config.CheckboxColumn('3근',  width=66),
        '주간': st.column_config.CheckboxColumn('주간', width=66),
        '야간': st.column_config.CheckboxColumn('야간', width=66),
    }

    _safety_edit = st.data_editor(
        st.session_state[_safety_key],
        key=f'de_safety_{_grid_key}',
        column_config=_safety_col_cfg,
        column_order=['안전 관리 사항', '1근', '2근', '3근', '주간', '야간'],
        hide_index=True,
        use_container_width=True,
        num_rows='fixed',
    )
    st.session_state[_safety_key] = _safety_edit

    safety_items_data = []
    for _qi in range(len(safety_questions)):
        safety_items_data.append({
            'text':  safety_questions[_qi],
            's1':    bool(_safety_edit.at[_qi, '1근']),
            's2':    bool(_safety_edit.at[_qi, '2근']),
            's3':    bool(_safety_edit.at[_qi, '3근']),
            'day':   bool(_safety_edit.at[_qi, '주간']),
            'night': bool(_safety_edit.at[_qi, '야간']),
        })

    # 4. 특이 사항
    st.markdown(
        '<div style="background:#4B2D8E;color:white;padding:6px 14px;'
        'font-size:13px;font-weight:bold;margin-top:18px;'
        'border-radius:4px;border:2px solid #4B2D8E;margin-bottom:4px;">'
        '4. 특이 사항</div>',
        unsafe_allow_html=True,
    )
    _note_key = f'wl_note_{selected_date}'
    if _note_key not in st.session_state:
        st.session_state[_note_key] = loaded_detail.get('note', '') if loaded_detail else ''
    note_text = st.text_area('특이사항', key=_note_key, height=100, label_visibility='collapsed')

    # 저장 / 다운로드 / 메일
    st.markdown('---')
    import time as _time
    _fn_team   = st.secrets.get('company', {}).get('team', '일일업무')
    _last_save = st.session_state.get('_last_save_ts', 0)
    _cooldown  = 20
    _elapsed   = _time.time() - _last_save
    _can_save  = _elapsed >= _cooldown

    if st.button('저장', use_container_width=True, type='primary', disabled=not _can_save):
        try:
            with st.spinner('저장 중...'):
                from utils.supabase_db import save_all
                save_all(
                    selected_date, work_items_data, shift_data_final,
                    safety_items_data, note_text,
                    st.session_state.get('leave_list', [])
                )
            st.session_state['_last_save_ts'] = _time.time()
            st.session_state.pop(f'wl_monthly_totals_{selected_date}', None)
            with st.spinner(f'{selected_date.month}월 통합 Excel 생성 중...'):
                _mbytes, _mcount = generate_monthly_work_log_excel(selected_date)
            st.session_state['_monthly_bytes'] = _mbytes
            st.session_state['_monthly_count'] = _mcount
            st.session_state['_monthly_ym']    = (selected_date.year, selected_date.month)
            st.success(f'저장 완료! {selected_date.month}월 Excel 준비됨 ({_mcount}일)')
        except Exception as e:
            st.error(f'저장 실패: {type(e).__name__}: {e}')
    if not _can_save:
        st.caption(f'재저장까지 {int(_cooldown - _elapsed)}초 대기')

    _my, _mm = st.session_state.get('_monthly_ym', (selected_date.year, selected_date.month))
    _monthly_fname = f'{_fn_team}_{_my}년{_mm}월_작업일지.xlsx'

    if st.session_state.get('_monthly_bytes'):
        _dl_label = f'{_my}년 {_mm}월 통합 다운로드 ({st.session_state["_monthly_count"]}일)'
        st.download_button(
            _dl_label,
            data=st.session_state['_monthly_bytes'],
            file_name=_monthly_fname,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            use_container_width=True,
        )

    _email_cfg = st.secrets.get('email', {})
    if _email_cfg.get('gmail_user'):
        st.markdown('---')
        if st.button('통합일지 메일 전송', use_container_width=True):
            st.session_state['_show_mail_form'] = True
        if st.session_state.get('_show_mail_form'):
            with st.expander('메일 작성 및 미리보기', expanded=True):
                _def_to      = _email_cfg.get('to', '')
                _def_subject = _email_cfg.get('subject', '{year}년 {month}월 일일업무보고 통합').format(year=_my, month=_mm)
                _def_body    = _email_cfg.get('body', '{month}월 업무보고 통합 파일을 첨부합니다.').format(year=_my, month=_mm)
                _to      = st.text_input('받는 사람', value=_def_to, help='쉼표로 여러 명 입력 가능')
                _subject = st.text_input('제목', value=_def_subject)
                _body    = st.text_area('본문', value=_def_body, height=150)
                st.markdown('**첨부파일**')
                _attach_default = st.checkbox(f'{_monthly_fname} (통합 Excel)', value=True)
                _extra_files = st.file_uploader('추가 첨부파일 (선택)', accept_multiple_files=True, key='mail_extra_attach')
                col_send, col_cancel = st.columns(2)
                with col_cancel:
                    if st.button('취소', use_container_width=True):
                        st.session_state['_show_mail_form'] = False
                        st.rerun()
                with col_send:
                    if st.button('전송', use_container_width=True, type='primary'):
                        if not _to.strip():
                            st.error('받는 사람 이메일을 입력하세요.')
                        else:
                            try:
                                import base64 as _b64
                                from email.mime.multipart import MIMEMultipart
                                from email.mime.text import MIMEText
                                from email.mime.base import MIMEBase
                                from email import encoders
                                from email.header import Header
                                import google.oauth2.credentials as _goauth
                                import googleapiclient.discovery as _gdisco
                                from utils.sheets import _get_or_create_sheet as _gos
                                _recipients = [r.strip() for r in _to.split(',') if r.strip()]
                                msg = MIMEMultipart('mixed')
                                msg['To']      = ', '.join(_recipients)
                                msg['Subject'] = Header(_subject, 'utf-8').encode()
                                msg.attach(MIMEText(_body, 'plain', 'utf-8'))
                                def _attach_file(data: bytes, fname: str):
                                    part = MIMEBase('application', 'octet-stream')
                                    part.set_payload(data)
                                    encoders.encode_base64(part)
                                    part.add_header('Content-Disposition', 'attachment', filename=('utf-8', '', fname))
                                    msg.attach(part)
                                if _attach_default:
                                    _attach_file(st.session_state['_monthly_bytes'], _monthly_fname)
                                for _ef in (_extra_files or []):
                                    _attach_file(_ef.getvalue(), _ef.name)
                                with st.spinner('메일 전송 중...'):
                                    _gcfg_ws = _gos('gmail_config')
                                    _gcfg = {r[0]: r[1] for r in _gcfg_ws.get_all_values() if len(r) >= 2}
                                    _gcreds = _goauth.Credentials(
                                        token=None,
                                        refresh_token=_gcfg['refresh_token'],
                                        token_uri='https://oauth2.googleapis.com/token',
                                        client_id=_gcfg['client_id'],
                                        client_secret=_gcfg['client_secret'],
                                    )
                                    _svc = _gdisco.build('gmail', 'v1', credentials=_gcreds)
                                    _raw = _b64.urlsafe_b64encode(msg.as_bytes()).decode()
                                    _svc.users().messages().send(userId='me', body={'raw': _raw}).execute()
                                st.success(f'메일 전송 완료! -> {_to}')
                                st.session_state['_show_mail_form'] = False
                            except Exception as _e:
                                st.error(f'전송 실패: {_e}')



def page_statistics():
    import datetime
    import calendar

    st.title("월별 근무 통계")

    MEMBERS = dict(st.secrets.get("members", {'A': '직원A', 'B': '직원B', 'C': '직원C', 'D': '직원D'}))
    ALL_MEMBERS = list(MEMBERS.values())

    # 야간근로 기준시간 (근무 유형별)
    NIGHT_HOURS = {"1근": 0, "2근": 0.5, "3근": 7.5, "주간": 0, "야간": 7.5}

    today_kst = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()
    col1, col2 = st.columns(2)
    with col1:
        years = list(range(2026, 2101))
        year_idx = max(0, min(today_kst.year - 2026, len(years) - 1))
        year = st.selectbox("연도", years, index=year_idx)
    with col2:
        month = st.selectbox("월", range(1, 13), index=max(0, today_kst.month - 1))

    target_month = datetime.date(year, month, 1)
    days_in_month = calendar.monthrange(year, month)[1]

    # 해당 월 데이터 로드
    try:
        from utils.supabase_db import load_daily_detail_month
        daily_details = load_daily_detail_month(target_month.year, target_month.month)
    except Exception:
        daily_details = {}

    # 휴가 등록 데이터 로드 (교대 스케줄 기반 계산용)
    try:
        from utils.supabase_db import load_leaves
        base_leaves = load_leaves()
    except Exception:
        base_leaves = []

    def _sf(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            import re
            nums = re.findall(r'[\d.]+', str(val))
            return float(nums[0]) if nums else 0

    # 직원별 통계 계산 (전체 날짜: 저장 일지 우선, 없으면 교대 스케줄 기반)
    stats = {name: {
        "근무일수": 0, "기본근로": 0, "연장근로_대근": 0,
        "연장근로_주간": 0, "연장근로_야간": 0,
        "야간근로": 0, "휴가일수": 0, "대근횟수": 0, "휴가내역": [], "대근내역": []
    } for name in ALL_MEMBERS}

    saved_days = len(daily_details)

    for day in range(1, days_in_month + 1):
        date = datetime.date(year, month, day)
        date_str = date.strftime("%Y-%m-%d")

        if date_str in daily_details:
            # ── 저장된 일지 데이터 사용 ──
            shift = daily_details[date_str].get("shift", {})
            is_2p = shift.get("is_2person", False)

            if is_2p:
                day_worker = shift.get("1근_근무자", "")
                night_worker = shift.get("2근_근무자", "")
                leave_worker = shift.get("3근_근무자", "")
                leave_type = shift.get("3근_비고", "")
                if day_worker in stats:
                    day_total = _sf(shift.get("1근_연장", 4))
                    stats[day_worker]["근무일수"] += 1
                    stats[day_worker]["기본근로"] += 8
                    stats[day_worker]["연장근로_대근"] += 4
                    stats[day_worker]["연장근로_주간"] += max(0, _sf(shift.get("1근_주간연장", 4)) - 4)
                    stats[day_worker]["연장근로_야간"] += _sf(shift.get("1근_야간연장", 0))
                    stats[day_worker]["야간근로"] += NIGHT_HOURS.get("주간", 0)
                    stats[day_worker]["대근횟수"] += 1
                    stats[day_worker]["대근내역"].append({"날짜": date_str, "휴가자": leave_worker, "휴가구분": leave_type, "시간": day_total, "구분": "주간"})
                if night_worker in stats:
                    night_total = _sf(shift.get("2근_연장", 4))
                    stats[night_worker]["근무일수"] += 1
                    stats[night_worker]["기본근로"] += 8
                    stats[night_worker]["연장근로_대근"] += 4
                    stats[night_worker]["연장근로_주간"] += _sf(shift.get("2근_주간연장", 0))
                    stats[night_worker]["연장근로_야간"] += max(0, _sf(shift.get("2근_야간연장", 4)) - 4)
                    stats[night_worker]["야간근로"] += 8
                    stats[night_worker]["대근횟수"] += 1
                    stats[night_worker]["대근내역"].append({"날짜": date_str, "휴가자": leave_worker, "휴가구분": leave_type, "시간": night_total, "구분": "야간"})
                if leave_worker in stats and leave_type:
                    stats[leave_worker]["휴가일수"] += 1
                    stats[leave_worker]["휴가내역"].append(f"{date_str}: {leave_type}")
            else:
                for shift_key, night_key in [("1근_근무자", "1근"), ("2근_근무자", "2근"), ("3근_근무자", "3근")]:
                    worker = shift.get(shift_key, "")
                    ot = _sf(shift.get(f"{night_key}_연장", 0))
                    if worker in stats:
                        stats[worker]["근무일수"] += 1
                        stats[worker]["기본근로"] += 8
                        base_night = NIGHT_HOURS.get(night_key, 0)
                        explicit_day = shift.get(f"{night_key}_주간연장")
                        explicit_night = shift.get(f"{night_key}_야간연장")
                        if explicit_day is not None or explicit_night is not None:
                            day_ot = _sf(explicit_day or 0)
                            night_ot = _sf(explicit_night or 0)
                        elif shift.get(f"{night_key}_연장유형") == "야간연장":
                            day_ot, night_ot = 0, ot
                        else:
                            day_ot, night_ot = ot, 0
                        stats[worker]["연장근로_주간"] += day_ot
                        stats[worker]["연장근로_야간"] += night_ot
                        stats[worker]["야간근로"] += base_night + night_ot
            off_worker = shift.get("휴무_근무자", "")
            off_type = shift.get("휴무_구분", "")
            if off_worker in stats and off_type not in ("교대휴무", "주휴휴무", ""):
                stats[off_worker]["휴가일수"] += 1
                stats[off_worker]["휴가내역"].append(f"{date_str}: {off_type}")

        else:
            # ── 교대 스케줄 기반 기본값 계산 ──
            shift = _shift_for_date(date, MEMBERS)
            shift = _apply_leaves_stat(shift, date, base_leaves)
            is_2p = shift.get("is_2person", False)

            if is_2p:
                day_w = shift.get("주간_근무자", "")
                night_w = shift.get("야간_근무자", "")
                leave_w = shift.get("leave_person", "")
                leave_t = shift.get("leave_type", "")
                if day_w in stats:
                    stats[day_w]["근무일수"] += 1
                    stats[day_w]["기본근로"] += 8
                    stats[day_w]["연장근로_대근"] += 4
                    stats[day_w]["대근횟수"] += 1
                    stats[day_w]["대근내역"].append({"날짜": date_str, "휴가자": leave_w, "휴가구분": leave_t, "시간": 4, "구분": "주간"})
                if night_w in stats:
                    stats[night_w]["근무일수"] += 1
                    stats[night_w]["기본근로"] += 8
                    stats[night_w]["연장근로_대근"] += 4
                    stats[night_w]["야간근로"] += 8
                    stats[night_w]["대근횟수"] += 1
                    stats[night_w]["대근내역"].append({"날짜": date_str, "휴가자": leave_w, "휴가구분": leave_t, "시간": 4, "구분": "야간"})
                if leave_w in stats and leave_t:
                    stats[leave_w]["휴가일수"] += 1
                    stats[leave_w]["휴가내역"].append(f"{date_str}: {leave_t} (예정)")
            else:
                for worker_key, sk in [("1근_근무자", "1근"), ("2근_근무자", "2근"), ("3근_근무자", "3근")]:
                    worker = shift.get(worker_key, "")
                    if worker in stats:
                        stats[worker]["근무일수"] += 1
                        stats[worker]["기본근로"] += 8
                        stats[worker]["야간근로"] += _NIGHT_HOURS_BASE.get(sk, 0)
                off_worker = shift.get("휴무_근무자", "")
                off_type = shift.get("휴무_구분", "")
                if off_worker in stats and off_type not in ("교대휴무", "주휴휴무", ""):
                    stats[off_worker]["휴가일수"] += 1
                    stats[off_worker]["휴가내역"].append(f"{date_str}: {off_type} (예정)")

    # 올해 전체 휴가 데이터 로드
    year_leaves = {name: [] for name in ALL_MEMBERS}
    saved_dates_year = set()
    try:
        year_prefix = f"{year}-"
        for row in all_data[1:]:
            if row and row[0].startswith(year_prefix) and len(row) > 1:
                saved_dates_year.add(row[0])
                try:
                    import json as _json
                    detail = _json.loads(row[1])
                    shift_y = detail.get("shift", {})
                    is_2p_y = shift_y.get("is_2person", False)

                    if is_2p_y:
                        lw = shift_y.get("3근_근무자", "")
                        lt = shift_y.get("3근_비고", "")
                        if lw in year_leaves and lt:
                            year_leaves[lw].append({"날짜": row[0], "구분": lt})

                    off_w = shift_y.get("휴무_근무자", "")
                    off_t = shift_y.get("휴무_구분", "")
                    if off_w in year_leaves and off_t not in ("교대휴무", "주휴휴무", ""):
                        year_leaves[off_w].append({"날짜": row[0], "구분": off_t})
                except Exception:
                    pass
    except Exception:
        pass

    # 휴가 레지스트리에서 저장 일지 없는 날짜 보완
    today_kst = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()
    for lv in base_leaves:
        try:
            lv_start = datetime.date.fromisoformat(lv["start"])
            lv_end = datetime.date.fromisoformat(lv["end"])
        except Exception:
            continue
        cur = lv_start
        while cur <= lv_end and cur <= today_kst:
            if cur.year == year:
                ds = cur.strftime("%Y-%m-%d")
                if ds not in saved_dates_year and lv["name"] in year_leaves:
                    year_leaves[lv["name"]].append({"날짜": ds, "구분": lv["type"]})
            cur += datetime.timedelta(days=1)

    # 통계 표시
    st.markdown("---")
    st.subheader(f"{year}년 {month}월 직원별 근무 통계")

    def _fh(v):
        """float/int 시간값을 깔끔하게 표시 (4.0 → 4, 5.5 → 5.5)"""
        try:
            v = float(v)
            return int(v) if v == int(v) else v
        except Exception:
            return v

    # ── 급여시간표 헬퍼 ──
    _SALARY_COLS = [
        "정상근로", "유휴근로", "휴일근로", "연장근로", "휴일연장", "야간근로",
        "휴일비근로", "휴가비근로", "스틸아카데미", "항군교육",
        "사내교육(1)", "사내교육(1.5)", "사외교육(1)", "사외교육(1.5)", "공가",
    ]

    def _build_salary_rows(name):
        def _sf(v):
            try:
                return float(v)
            except Exception:
                return 0.0

        # 한국 법정공휴일 + 대체공휴일 세트 (해당 월 포함 연도)
        try:
            import holidays as _hol
            _kr_holidays = _hol.SouthKorea(years=[year, year - 1, year + 1])
        except Exception:
            _kr_holidays = set()

        def _is_holiday(d):
            return d in _kr_holidays

        def _fill_2p(row, day_worker, night_worker, leave_person, leave_type_val,
                     is_hol, day_ot=4.0, night_ot=4.0, night_base=7.5):
            """2인 근무 시 해당 인원의 행 채우기"""
            if name == leave_person:
                if leave_type_val == "공가":
                    row["공가"] = 8.0
                elif leave_type_val == "공휴":
                    pass  # 공휴일 → 급여시간 없음
                else:
                    row["휴가비근로"] = 8.0
            elif name == day_worker:
                if is_hol or leave_type_val == "공휴":
                    row["유휴근로"] = 8.0
                    row["휴일연장"] = day_ot
                    row["휴일비근로"] = 8.0
                else:
                    row["정상근로"] = 8.0
                    row["연장근로"] = day_ot
            elif name == night_worker:
                if is_hol or leave_type_val == "공휴":
                    row["유휴근로"] = 8.0
                    row["야간근로"] = night_base
                    row["휴일연장"] = night_ot
                    row["휴일비근로"] = 8.0
                else:
                    row["정상근로"] = 8.0
                    row["야간근로"] = night_base
                    row["연장근로"] = night_ot

        def _fill_3p_worker(row, 근, is_hol, day_ot=0.0, night_ot_h=0.0):
            """3인 근무 해당 근의 행 채우기"""
            night_base = NIGHT_HOURS.get(근, 0)
            if is_hol:
                row["유휴근로"] = 8.0
                row["휴일비근로"] = 8.0
                row["휴일연장"] = day_ot + night_ot_h
                if night_base > 0:
                    row["야간근로"] = night_base
            else:
                row["정상근로"] = 8.0
                row["연장근로"] = day_ot
                if night_base > 0:
                    row["야간근로"] = night_base + night_ot_h

        rows = []
        for day in range(1, days_in_month + 1):
            date = datetime.date(year, month, day)
            date_str = date.strftime("%Y-%m-%d")
            is_hol = _is_holiday(date)

            row = {c: 0.0 for c in _SALARY_COLS}
            row["날짜"] = f"{month:02d}/{day:02d}"

            if date_str in daily_details:
                # ── 저장된 일지 우선 사용 ──
                shift = daily_details[date_str].get("shift", {})
                is_2p = shift.get("is_2person", False)

                if is_2p:
                    lp = shift.get("leave_person", "") or shift.get("3근_근무자", "")
                    dw = shift.get("주간_근무자", "") or shift.get("1근_근무자", "")
                    nw = shift.get("야간_근무자", "") or shift.get("2근_근무자", "")
                    lv = shift.get("leave_type", "") or shift.get("3근_비고", "")
                    d_ot = _sf(shift.get("1근_연장", 4))
                    n_ot = _sf(shift.get("2근_연장", 4))
                    nb = NIGHT_HOURS.get("야간", 7.5)
                    _fill_2p(row, dw, nw, lp, lv, is_hol, d_ot, n_ot, nb)
                else:
                    for sk, 근 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        if shift.get(sk) == name:
                            exp_d = shift.get(f"{근}_주간연장")
                            exp_n = shift.get(f"{근}_야간연장")
                            base = _sf(shift.get(f"{근}_연장", 0))
                            d_ot = _sf(exp_d) if exp_d is not None else base
                            n_ot = _sf(exp_n) if exp_n is not None else 0.0
                            _fill_3p_worker(row, 근, is_hol, d_ot, n_ot)
                            break
                    else:
                        if shift.get("휴무_근무자") == name:
                            ot = shift.get("휴무_구분", "")
                            if ot in ("정기휴가","연차","특별휴가","명휴","생일휴가","청원휴가"):
                                row["휴가비근로"] = 8.0
                            elif ot == "공가":
                                row["공가"] = 8.0

            else:
                # ── 저장 없음 → 교대 스케줄 기반 기본값 ──
                sched = _shift_for_date(date, MEMBERS)
                sched = _apply_leaves_stat(sched, date, base_leaves)
                is_2p = sched.get("is_2person", False)

                if is_2p:
                    lp = sched.get("leave_person", "")
                    dw = sched.get("주간_근무자", "")
                    nw = sched.get("야간_근무자", "")
                    lv = sched.get("leave_type", "")
                    nb = NIGHT_HOURS.get("야간", 7.5)
                    _fill_2p(row, dw, nw, lp, lv, is_hol, 4.0, 4.0, nb)
                else:
                    for sk, 근 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        if sched.get(sk) == name:
                            _fill_3p_worker(row, 근, is_hol)
                            break
                    else:
                        if sched.get("휴무_근무자") == name:
                            ot = sched.get("휴무_구분", "")
                            if ot in ("정기휴가","연차","특별휴가","명휴","생일휴가","청원휴가"):
                                row["휴가비근로"] = 8.0
                            elif ot == "공가":
                                row["공가"] = 8.0
                            # 교대휴무/주휴휴무 → 시간 없음 (휴무일)

            total = sum(row[c] for c in _SALARY_COLS)
            row["일별합계"] = total
            if total > 0:
                rows.append(row)
        return rows

    def _render_salary_table(rows):
        totals = {c: sum(r[c] for r in rows) for c in _SALARY_COLS}
        totals["일별합계"] = sum(r["일별합계"] for r in rows)

        def _fmt(v):
            return f"{v:.2f}" if v else ""

        header_sub = "".join(f'<th style="background:#4472C4;color:#fff;text-align:center;padding:5px 3px;border:1px solid #2F5496;font-size:10px;white-space:nowrap;">{c}</th>' for c in _SALARY_COLS)
        html = f"""<div style="overflow-x:auto;margin-top:8px;">
<table style="border-collapse:collapse;font-size:11px;width:100%;min-width:900px;">
<thead>
<tr>
  <th rowspan="2" style="background:#4472C4;color:#fff;text-align:center;padding:6px 4px;border:1px solid #2F5496;white-space:nowrap;">날짜</th>
  <th colspan="{len(_SALARY_COLS)}" style="background:#4472C4;color:#fff;text-align:center;padding:6px 4px;border:1px solid #2F5496;">일일 급여시간</th>
  <th rowspan="2" style="background:#2F5496;color:#fff;text-align:center;padding:6px 4px;border:1px solid #2F5496;white-space:nowrap;">일별합계</th>
</tr>
<tr>{header_sub}</tr>
</thead><tbody>"""

        for i, row in enumerate(rows):
            bg = "#f0f4fb" if i % 2 == 0 else "#ffffff"
            cells = "".join(
                f'<td style="text-align:right;padding:3px 5px;border:1px solid #D0D7E4;background:{bg};">{_fmt(row[c])}</td>'
                for c in _SALARY_COLS
            )
            total_val = _fmt(row["일별합계"])
            html += (
                f'<tr><td style="text-align:center;padding:3px 5px;border:1px solid #D0D7E4;background:#EEF2FA;font-weight:600;">{row["날짜"]}</td>'
                f'{cells}'
                f'<td style="text-align:right;padding:3px 5px;border:1px solid #2F5496;background:#D9E1F2;font-weight:700;color:#1F3864;">{total_val}</td></tr>'
            )

        total_cells = "".join(
            f'<td style="text-align:right;padding:4px 5px;border:1px solid #9DC3E6;background:#BDD7EE;font-weight:700;color:#1F3864;">{_fmt(totals[c])}</td>'
            for c in _SALARY_COLS
        )
        html += (
            f'<tr><td style="text-align:center;padding:4px 5px;border:1px solid #9DC3E6;background:#9DC3E6;font-weight:700;color:#1F3864;">근로별 월합계</td>'
            f'{total_cells}'
            f'<td style="text-align:right;padding:4px 5px;border:1px solid #9DC3E6;background:#9DC3E6;font-weight:700;color:#1F3864;">{_fmt(totals["일별합계"])}</td></tr>'
        )
        html += "</tbody></table></div>"
        return html

    for name in ALL_MEMBERS:
        s = stats[name]
        total_ot = s["연장근로_대근"] + s["연장근로_주간"] + s["연장근로_야간"]
        yr_leaves = year_leaves.get(name, [])
        with st.container(border=True):
            st.markdown(f"### {name}")
            mc1, mc2, mc3, mc4, mc5, mc6, mc7 = st.columns(7)
            mc1.metric("근무일수", f"{s['근무일수']}일")
            mc2.metric("기본근로", f"{_fh(s['기본근로'])}H")
            mc3.metric("연장(대근)", f"{_fh(s['연장근로_대근'])}H")
            mc4.metric("주간연장", f"{_fh(s['연장근로_주간'])}H")
            mc5.metric("야간연장", f"{_fh(s['연장근로_야간'])}H")
            mc6.metric("야간근로", f"{_fh(s['야간근로'])}H")
            mc7.metric("휴가(월)", f"{s['휴가일수']}일")

            total_daegeun_h = sum(d["시간"] for d in s["대근내역"])

            # 대근 상세 내역
            if s["대근내역"]:
                with st.expander(f"대근 내역 ({s['대근횟수']}회 · 계 {_fh(total_daegeun_h)}H)"):
                    for d in s["대근내역"]:
                        st.write(f"{d['날짜']} | {d['구분']} | {d['휴가자']} {d['휴가구분']}으로 대근 | {_fh(d['시간'])}H")

            # 이번 달 휴가 내역
            if s["휴가내역"]:
                with st.expander(f"{month}월 휴가 내역 ({s['휴가일수']}일)"):
                    for h in s["휴가내역"]:
                        st.write(h)

            # 올해 전체 휴가 내역
            if yr_leaves:
                with st.expander(f"{year}년 전체 휴가 내역 ({len(yr_leaves)}일)"):
                    # 유형별 집계
                    type_count = {}
                    for lv in yr_leaves:
                        type_count[lv["구분"]] = type_count.get(lv["구분"], 0) + 1
                    summary = " | ".join(f"{k}: {v}일" for k, v in type_count.items())
                    st.info(summary)
                    for lv in yr_leaves:
                        st.write(f"{lv['날짜']} — {lv['구분']}")

            # 급여시간표
            with st.expander(f"{month}월 급여시간표 자세히보기"):
                try:
                    _salary_rows = _build_salary_rows(name)
                except Exception as _serr:
                    st.error(f"급여시간표 계산 오류: {_serr}")
                    _salary_rows = []
                # _build_salary_rows가 비어 있으면 직접 인라인 계산 (클로저 우회)
                if not _salary_rows:
                    try:
                        import holidays as _hol2
                        _kr2 = _hol2.SouthKorea(years=[year, year-1, year+1])
                    except Exception:
                        _kr2 = set()
                    _salary_rows = []
                    for _fd in range(1, days_in_month + 1):
                        _fdate = datetime.date(year, month, _fd)
                        _fds = _fdate.strftime("%Y-%m-%d")
                        _fhol = _fdate in _kr2
                        _frow = {c: 0.0 for c in _SALARY_COLS}
                        _frow["날짜"] = f"{month:02d}/{_fd:02d}"
                        if _fds in daily_details:
                            _sh = daily_details[_fds].get("shift", {})
                            if _sh.get("is_2person"):
                                _lp = _sh.get("leave_person","") or _sh.get("3근_근무자","")
                                _dw = _sh.get("주간_근무자","") or _sh.get("1근_근무자","")
                                _nw = _sh.get("야간_근무자","") or _sh.get("2근_근무자","")
                                _lv = _sh.get("leave_type","") or _sh.get("3근_비고","")
                                _nb2 = NIGHT_HOURS.get("야간", 7.5)
                                _dot = float(_sh.get("1근_연장", 4) or 4)
                                _not = float(_sh.get("2근_연장", 4) or 4)
                                if name == _lp:
                                    if _lv == "공가": _frow["공가"] = 8.0
                                    elif _lv != "공휴": _frow["휴가비근로"] = 8.0
                                elif name == _dw:
                                    if _fhol or _lv == "공휴":
                                        _frow["유휴근로"] = 8.0; _frow["휴일연장"] = _dot; _frow["휴일비근로"] = 8.0
                                    else:
                                        _frow["정상근로"] = 8.0; _frow["연장근로"] = _dot
                                elif name == _nw:
                                    if _fhol or _lv == "공휴":
                                        _frow["유휴근로"] = 8.0; _frow["야간근로"] = _nb2; _frow["휴일연장"] = _not; _frow["휴일비근로"] = 8.0
                                    else:
                                        _frow["정상근로"] = 8.0; _frow["야간근로"] = _nb2; _frow["연장근로"] = _not
                            else:
                                for _sk2, _근2 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                                    if _sh.get(_sk2) == name:
                                        _nb2 = NIGHT_HOURS.get(_근2, 0)
                                        _ot2 = float(_sh.get(f"{_근2}_연장", 0) or 0)
                                        _doy2 = float(_sh.get(f"{_근2}_주간연장") or _ot2)
                                        _noy2 = float(_sh.get(f"{_근2}_야간연장") or 0)
                                        if _fhol:
                                            _frow["유휴근로"] = 8.0; _frow["휴일비근로"] = 8.0
                                            _frow["휴일연장"] = _doy2 + _noy2
                                            if _nb2 > 0: _frow["야간근로"] = _nb2
                                        else:
                                            _frow["정상근로"] = 8.0; _frow["연장근로"] = _doy2
                                            if _nb2 > 0: _frow["야간근로"] = _nb2 + _noy2
                                        break
                        else:
                            _fsched = _shift_for_date(_fdate, MEMBERS)
                            _fsched = _apply_leaves_stat(_fsched, _fdate, base_leaves)
                            if _fsched.get("is_2person"):
                                _nb2 = NIGHT_HOURS.get("야간", 7.5)
                                _lp2 = _fsched.get("leave_person","")
                                _dw2 = _fsched.get("주간_근무자","")
                                _nw2 = _fsched.get("야간_근무자","")
                                _lv2 = _fsched.get("leave_type","")
                                if name == _lp2:
                                    if _lv2 == "공가": _frow["공가"] = 8.0
                                    elif _lv2 != "공휴": _frow["휴가비근로"] = 8.0
                                elif name == _dw2:
                                    if _fhol or _lv2 == "공휴":
                                        _frow["유휴근로"] = 8.0; _frow["휴일연장"] = 4.0; _frow["휴일비근로"] = 8.0
                                    else:
                                        _frow["정상근로"] = 8.0; _frow["연장근로"] = 4.0
                                elif name == _nw2:
                                    if _fhol or _lv2 == "공휴":
                                        _frow["유휴근로"] = 8.0; _frow["야간근로"] = _nb2; _frow["휴일연장"] = 4.0; _frow["휴일비근로"] = 8.0
                                    else:
                                        _frow["정상근로"] = 8.0; _frow["야간근로"] = _nb2; _frow["연장근로"] = 4.0
                            else:
                                for _sk2, _근2 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                                    if _fsched.get(_sk2) == name:
                                        _nb2 = NIGHT_HOURS.get(_근2, 0)
                                        if _fhol:
                                            _frow["유휴근로"] = 8.0; _frow["휴일비근로"] = 8.0
                                            if _nb2 > 0: _frow["야간근로"] = _nb2
                                        else:
                                            _frow["정상근로"] = 8.0
                                            if _nb2 > 0: _frow["야간근로"] = _nb2
                                        break
                        _ftotal = sum(_frow[c] for c in _SALARY_COLS)
                        _frow["일별합계"] = _ftotal
                        if _ftotal > 0:
                            _salary_rows.append(_frow)
                if _salary_rows:
                    if not daily_details:
                        st.caption("※ 저장된 일지 없음 — 교대 스케줄 기준 기본 급여시간 (예상치)")
                    st.markdown(_render_salary_table(_salary_rows), unsafe_allow_html=True)
                else:
                    st.info("저장된 근무 데이터가 없습니다.")

            # 교대주기별 연장 시간
            with st.expander("교대주기별 연장 시간", expanded=False):
                st.caption("조회월 기준 해당 근무조 교대 주기(연속 근무 5일)별 연장 현황. 주기당 최대 12H — 초과 시 빨간색 경고.")
                _OT_LIMIT = 12
                _team_key = next((k for k, v in MEMBERS.items() if v == name), None)
                if _team_key is None:
                    st.info("조 코드 미매핑")
                else:
                    _m_start = datetime.date(year, month, 1)
                    _next_m = datetime.date(year, month + 1, 1) if month < 12 else datetime.date(year + 1, 1, 1)
                    _m_end = _next_m - datetime.timedelta(days=1)
                    _scan_start = _m_start - datetime.timedelta(days=6)
                    _scan_end = _m_end + datetime.timedelta(days=6)

                    # 날짜별 연장시간 맵: 대근내역(이미 정확히 계산됨) + 일지 저장 주간/야간 연장
                    _ot_by_date = {}
                    for _dk in s.get("대근내역", []):
                        _ot_by_date[_dk["날짜"]] = _ot_by_date.get(_dk["날짜"], 0) + _dk["시간"]
                    for _ds2, _dd2 in daily_details.items():
                        if _ds2 in _ot_by_date:
                            continue
                        _sh2 = _dd2.get("shift", {})
                        if not _sh2.get("is_2person"):
                            for _sk2, _ok2 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                                if _sh2.get(_sk2) == name:
                                    _ot2 = _sf(_sh2.get(f"{_ok2}_연장", 0))
                                    if _ot2 > 0:
                                        _ot_by_date[_ds2] = _ot2
                                    break

                    _work_seq = []
                    _d = _scan_start
                    while _d <= _scan_end:
                        _idx = (_d - _BASE_DATE).days % 20
                        _s1, _s2, _s3, _off = _CYCLE_20[_idx]
                        if _team_key in (_s1, _s2, _s3):
                            _ds = _d.strftime("%Y-%m-%d")
                            _ot = _ot_by_date.get(_ds, 0)
                            _work_seq.append((_d, _ot))
                        _d += datetime.timedelta(days=1)

                    _blocks = []
                    if _work_seq:
                        _cur = [_work_seq[0]]
                        for _i in range(1, len(_work_seq)):
                            if (_work_seq[_i][0] - _work_seq[_i - 1][0]).days == 1:
                                _cur.append(_work_seq[_i])
                            else:
                                _blocks.append(_cur)
                                _cur = [_work_seq[_i]]
                        _blocks.append(_cur)

                    _month_blocks = [b for b in _blocks if b[-1][0] >= _m_start and b[0][0] <= _m_end]

                    if not _month_blocks:
                        st.info("해당 월 교대 주기 없음")
                    else:
                        _n = len(_month_blocks)
                        _TH = "background:#374151;color:#D1D5DB;padding:5px 4px;border:1px solid #4B5563;text-align:center;font-size:12px;"
                        _TD_BASE = "padding:6px 4px;border:1px solid #4B5563;text-align:center;font-size:13px;"
                        _TD_GRAY = _TD_BASE + "background:#1F2937;color:#9CA3AF;"
                        _TD_OK   = _TD_BASE + "background:#1F2937;color:#E5E7EB;font-weight:600;"
                        _TD_RED  = _TD_BASE + "background:#DC2626;color:#fff;font-weight:700;"
                        _TD_WARN = _TD_BASE + "background:#78350F;color:#FDE68A;font-weight:600;"

                        _date_hdr = "".join(
                            '<th colspan="3" style="' + _TH + '">' +
                            b[0][0].strftime("%m/%d") + "~" + b[-1][0].strftime("%m/%d") + "</th>"
                            for b in _month_blocks
                        )
                        _col_sub = (
                            '<th style="' + _TH + '">연장(발생)</th>' +
                            '<th style="' + _TH + '">연장(잔여)</th>' +
                            '<th style="' + _TH + '">탄력근로 사용</th>'
                        ) * _n

                        _data_cells = ""
                        _any_warn = False
                        for _b in _month_blocks:
                            _b_ot = int(sum(_x[1] for _x in _b))
                            _remain = max(0, _OT_LIMIT - _b_ot)
                            _is_over = _b_ot >= _OT_LIMIT
                            _is_near = not _is_over and _b_ot >= _OT_LIMIT - 2
                            _any_warn = _any_warn or _is_over
                            _cell_style = _TD_RED if _is_over else (_TD_WARN if _is_near else _TD_OK)
                            _data_cells += (
                                '<td style="' + _cell_style + '">' + str(_b_ot) + '</td>' +
                                '<td style="' + _TD_GRAY + '">' + str(_remain) + '</td>' +
                                '<td style="' + _TD_GRAY + '">0</td>'
                            )

                        _html_ot = (
                            '<div style="overflow-x:auto;margin-top:8px;">'
                            '<table style="border-collapse:collapse;font-size:12px;min-width:100%;">'
                            '<thead>'
                            '<tr><th style="' + _TH + 'min-width:90px;">구분</th>' + _date_hdr + '</tr>'
                            '<tr><th style="' + _TH + '"></th>' + _col_sub + '</tr>'
                            '</thead>'
                            '<tbody>'
                            '<tr><td style="' + _TH + 'text-align:left;white-space:nowrap;">연장/잔여(H)</td>' + _data_cells + '</tr>'
                            '</tbody></table></div>'
                        )
                        st.markdown(_html_ot, unsafe_allow_html=True)

                        if _any_warn:
                            _warn_list = [
                                b[0][0].strftime("%m/%d") + "~" + b[-1][0].strftime("%m/%d") +
                                " (" + str(int(sum(_x[1] for _x in b))) + "H)"
                                for b in _month_blocks if int(sum(_x[1] for _x in b)) >= _OT_LIMIT
                            ]
                            st.error("⚠️ 주 52시간 위배 주기: " + ", ".join(_warn_list))


# ══════════════════════════════════════
# 메뉴 5: 근무 일정표
# ══════════════════════════════════════
def page_my_schedule():
    import calendar as _cal

    st.markdown("""
<style>
.sched-header {
    display: flex; align-items: center; justify-content: space-between;
    margin-bottom: 6px;
}
.cal-table { width: 100%; border-collapse: collapse; margin-top: 4px; }
.cal-table th {
    background: #1E1E2E; color: #CDD6F4;
    text-align: center; padding: 8px 0; font-size: 13px;
    border: 1px solid #313244;
}
.cal-table td {
    border: 1px solid #313244; vertical-align: top;
    padding: 6px 5px; min-height: 64px; min-width: 48px;
    background: #1E1E2E; cursor: default;
}
.cal-table td.today { background: #2A2A3E; border: 2px solid #89B4FA !important; }
.cal-table td.othermonth { background: #181825; }
.day-num {
    font-size: 12px; font-weight: 600; color: #BAC2DE; margin-bottom: 4px;
}
.day-num.sun { color: #F38BA8; }
.day-num.sat { color: #89B4FA; }
.day-num.today-num {
    background: #89B4FA; color: #1E1E2E; border-radius: 50%;
    width: 22px; height: 22px; display: inline-flex;
    align-items: center; justify-content: center;
}
.badge {
    display: inline-block; padding: 2px 7px; border-radius: 10px;
    font-size: 11px; font-weight: 700; margin-top: 2px; width: 100%;
    text-align: center; box-sizing: border-box;
}
.badge-1 { background: #1D4ED8; color: #EFF6FF; }
.badge-2 { background: #15803D; color: #F0FDF4; }
.badge-3 { background: #B91C1C; color: #FEF2F2; }
.badge-off { background: #374151; color: #D1D5DB; }
.badge-leave { background: #92400E; color: #FEF3C7; }
.badge-sub { background: #6D28D9; color: #EDE9FE; font-size: 10px; }
.detail-card {
    background: #1E1E2E; border: 1px solid #313244; border-radius: 12px;
    padding: 20px 24px; margin-top: 12px;
}
.detail-row {
    display: flex; align-items: center; gap: 12px;
    padding: 10px 0; border-bottom: 1px solid #313244;
}
.detail-row:last-child { border-bottom: none; }
.shift-pill {
    padding: 4px 14px; border-radius: 20px; font-weight: 700; font-size: 14px;
}
.legend-box {
    display: flex; gap: 10px; flex-wrap: wrap; margin-bottom: 14px;
}
.legend-item {
    display: flex; align-items: center; gap: 5px; font-size: 12px; color: #BAC2DE;
}
.legend-dot {
    width: 12px; height: 12px; border-radius: 3px;
}
.summary-grid {
    display: grid; grid-template-columns: repeat(5, 1fr); gap: 8px;
    margin-top: 12px; margin-bottom: 4px;
}
.summary-card {
    background: #1E1E2E; border: 1px solid #313244; border-radius: 8px;
    padding: 10px; text-align: center;
}
.summary-card .sc-num { font-size: 22px; font-weight: 700; }
.summary-card .sc-label { font-size: 11px; color: #BAC2DE; margin-top: 2px; }
</style>
""", unsafe_allow_html=True)

    st.markdown("## 근무표")

    MEMBERS = dict(st.secrets.get("members", {'A': '직원A', 'B': '직원B', 'C': '직원C', 'D': '직원D'}))
    ALL_MEMBERS = list(MEMBERS.values())

    # ── 휴가신청서 작성 ──
    if "leave_list" not in st.session_state:
        try:
            from utils.supabase_db import load_leaves as _ll
            st.session_state["leave_list"] = _ll()
        except Exception:
            st.session_state["leave_list"] = []

    with st.expander("📝 휴가신청서 작성", expanded=False):
        _rc1, _rc2, _rc3, _rc4, _rc5 = st.columns(5)
        with _rc1:
            _lv_name = st.selectbox("대상자", ALL_MEMBERS, key="leave_name")
        with _rc2:
            _lv_type = st.selectbox("구분", ["정기휴가", "연차", "특별휴가", "명휴", "생일휴가", "공가", "공상휴업", "산재", "휴직", "대휴", "교육", "결근", "조퇴", "외출", "청원휴가", "공휴"], key="leave_type")
        with _rc3:
            _lv_start = st.date_input("시작일", datetime.date.today(), key="leave_start")
        with _rc4:
            _lv_end = st.date_input("종료일", datetime.date.today(), key="leave_end")
        with _rc5:
            _lv_sub = st.selectbox("대근자", [""] + ALL_MEMBERS, key="leave_sub")

        _gonghu_nonholiday = []
        if _lv_type == "공휴":
            try:
                import holidays as _hol
                _cur = _lv_start
                _wday = ["월", "화", "수", "목", "금", "토", "일"]
                while _cur <= _lv_end:
                    _kr = _hol.SouthKorea(years=_cur.year)
                    if _cur not in _kr:
                        _gonghu_nonholiday.append(f"{_cur.strftime('%m/%d')}({_wday[_cur.weekday()]})")
                    _cur += datetime.timedelta(days=1)
            except Exception:
                pass

        if _gonghu_nonholiday:
            st.warning(
                "⚠️ **공휴는 국경일·명절·대체공휴일에만 적용 가능합니다.**  \n"
                f"선택 기간 중 법정공휴일이 아닌 날: **{', '.join(_gonghu_nonholiday)}**  \n"
                "그래도 등록하려면 아래를 체크하세요."
            )
            _gonghu_confirmed = st.checkbox("⚠️ 비공휴일 포함을 확인하고 등록합니다.", key="gonghu_confirm")
        else:
            _gonghu_confirmed = True

        if st.button("✅ 등록", use_container_width=True, key="add_leave"):
            if _lv_start > _lv_end:
                st.error("시작일이 종료일보다 늦습니다.")
            elif _gonghu_nonholiday and not _gonghu_confirmed:
                st.error("⛔ 비공휴일 포함 시 확인 체크가 필요합니다.")
            else:
                st.session_state["leave_list"].append({
                    "name": _lv_name, "type": _lv_type,
                    "start": _lv_start.isoformat(), "end": _lv_end.isoformat(),
                    "sub": _lv_sub,
                })
                try:
                    from utils.supabase_db import save_leaves
                    save_leaves(st.session_state["leave_list"])
                except Exception:
                    pass
                st.rerun()

        if st.session_state["leave_list"]:
            st.markdown("**등록된 일정 조회**")
            _today = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()
            _lv_years = sorted({datetime.date.fromisoformat(lv["start"]).year for lv in st.session_state["leave_list"]}, reverse=True)
            _lv_fc1, _lv_fc2 = st.columns(2)
            with _lv_fc1:
                _filter_yr = st.selectbox("연도", _lv_years, index=0 if _today.year not in _lv_years else _lv_years.index(_today.year), key="lv_filter_yr")
            with _lv_fc2:
                _filter_mo = st.selectbox("월", list(range(1, 13)), index=_today.month - 1, key="lv_filter_mo", format_func=lambda m: f"{m}월")

            # 선택 연/월에 걸치는 일정만 표시
            _sel_start = datetime.date(_filter_yr, _filter_mo, 1)
            _next_mo = (_filter_mo % 12) + 1
            _next_yr = _filter_yr + (1 if _filter_mo == 12 else 0)
            _sel_end = datetime.date(_next_yr, _next_mo, 1) - datetime.timedelta(days=1)

            _shown = [(i, lv) for i, lv in enumerate(st.session_state["leave_list"])
                      if datetime.date.fromisoformat(lv["end"]) >= _sel_start
                      and datetime.date.fromisoformat(lv["start"]) <= _sel_end]

            if not _shown:
                st.caption(f"{_filter_yr}년 {_filter_mo}월 등록 일정 없음")
            else:
                st.caption(f"{_filter_yr}년 {_filter_mo}월 — {len(_shown)}건")
                for _i, _lv in _shown:
                    _ct, _cd = st.columns([4, 1])
                    with _ct:
                        _sub = _lv.get('sub', '')
                        _sub_txt = f" | 대근: {_sub}" if _sub else ""
                        st.write(f"{_lv['name']} | {_lv['type']} | {_lv['start']} ~ {_lv['end']}{_sub_txt}")
                    with _cd:
                        if st.button("삭제", key=f"del_leave_{_i}"):
                            _deleted = st.session_state["leave_list"].pop(_i)
                            try:
                                from utils.supabase_db import save_leaves, delete_daily_details_for_leave
                                save_leaves(st.session_state["leave_list"])
                                _rcnt = delete_daily_details_for_leave(_deleted)
                                if _rcnt > 0:
                                    st.toast(f"휴가 삭제 — 관련 저장 일지 {_rcnt}일 초기화", icon="♻️")
                            except Exception:
                                pass
                            st.rerun()

    st.markdown("---")

    today = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()

    st.markdown("""<style>
/* 근무표 셀렉트박스 글자 크기·색상 강화 — Streamlit DOM 직접 타겟 */
div[data-testid="stSelectbox"] label p {
    font-size: 15px !important;
    font-weight: 700 !important;
    color: #CDD6F4 !important;
}
div[data-testid="stSelectbox"] div[data-baseweb="select"] > div:first-child {
    min-height: 48px !important;
    border-radius: 10px !important;
    border: 1.5px solid #45475A !important;
    background: #1E1E2E !important;
}
div[data-testid="stSelectbox"] [data-testid="stSelectboxVirtualDropdown"] span,
div[data-testid="stSelectbox"] div[data-baseweb="select"] span {
    font-size: 16px !important;
    font-weight: 700 !important;
    color: #CDD6F4 !important;
}
</style>""", unsafe_allow_html=True)

    _sc1, _sc2, _sc3, _sc4 = st.columns([2, 2, 1, 1])
    with _sc1:
        shift_type = st.selectbox(
            "근무 형태", ["4조3교대", "3조3교대", "2조2교대", "4조2교대"],
            key="sched_shift_type",
        )

    _SHIFT_TEAMS = {
        "4조3교대": ALL_MEMBERS,
        "3조3교대": ["A조", "B조", "C조"],
        "2조2교대": ["A조", "B조"],
        "4조2교대": ["A조", "B조", "C조", "D조"],
    }
    _team_label = "이름" if shift_type == "4조3교대" else "조"
    _team_code = ""  # non-4조3교대용 팀 코드 (A/B/C/D)

    with _sc2:
        selected_name = st.selectbox(_team_label, _SHIFT_TEAMS[shift_type], key="sched_name")
    with _sc3:
        selected_year = st.selectbox("년도", list(range(2020, today.year + 6)), index=list(range(2020, today.year + 6)).index(today.year), key="sched_yr")
    with _sc4:
        selected_month = st.selectbox("월", list(range(1, 13)), index=today.month - 1, key="sched_mo")

    st.markdown(
        f'<div style="background:#1E1E2E;border:1.5px solid #45475A;border-radius:10px;'
        f'padding:10px 18px;margin:6px 0 10px;display:flex;align-items:center;gap:16px;">'
        f'<span style="font-size:14px;font-weight:600;color:#6C7086;">근무형태</span>'
        f'<span style="font-size:18px;font-weight:800;color:#CDD6F4;">{shift_type}</span>'
        f'<span style="color:#45475A;font-size:20px;">|</span>'
        f'<span style="font-size:14px;font-weight:600;color:#6C7086;">{"이름" if shift_type == "4조3교대" else "조"}</span>'
        f'<span style="font-size:22px;font-weight:900;color:#89B4FA;">{selected_name}</span>'
        f'<span style="color:#45475A;font-size:20px;">|</span>'
        f'<span style="font-size:20px;font-weight:800;color:#CDD6F4;">{selected_year}년 {selected_month}월</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if shift_type != "4조3교대":
        _team_code = selected_name[0]  # "A조" → "A"

    # 이름 → 조 매핑 (4조3교대 전용)
    name_to_team = {v: k for k, v in MEMBERS.items()}
    my_team = name_to_team.get(selected_name, "")

    # 휴가 목록 로드 (4조3교대 전용)
    from utils.supabase_db import load_leaves as _load_leaves
    leave_list = []
    if shift_type == "4조3교대":
        try:
            leave_list = _load_leaves()
        except Exception:
            pass

    def _get_day_info(d):
        """특정 날짜의 근무 정보 반환"""
        # ── 4조3교대 외 단순 패턴 ──
        if shift_type == "3조3교대":
            s = _shift_for_date_3s3(d, _team_code)
            return {"shift": s, "is_leave": False, "leave_type": "", "sub_for": "", "sub_for_shift": "", "sub_role": "", "raw": None}
        if shift_type == "2조2교대":
            s = _shift_for_date_2s2(d, _team_code)
            return {"shift": s, "is_leave": False, "leave_type": "", "sub_for": "", "sub_for_shift": "", "sub_role": "", "raw": None}
        if shift_type == "4조2교대":
            s = _shift_for_date_4s2(d, _team_code)
            return {"shift": s, "is_leave": False, "leave_type": "", "sub_for": "", "sub_for_shift": "", "sub_role": "", "raw": None}

        # ── 4조3교대 ──
        shift = _shift_for_date(d, MEMBERS)
        s1, s2, s3, off = shift["1근_근무자"], shift["2근_근무자"], shift["3근_근무자"], shift["휴무_근무자"]

        # 본인 기본 근무
        if s1 == selected_name:
            my_shift = "1근"
        elif s2 == selected_name:
            my_shift = "2근"
        elif s3 == selected_name:
            my_shift = "3근"
        elif off == selected_name:
            my_shift = "휴무"
        else:
            my_shift = "?"

        # 휴가 적용 (본인)
        is_my_leave = False
        leave_type_my = ""
        for lv in leave_list:
            try:
                ls = datetime.date.fromisoformat(lv["start"])
                le = datetime.date.fromisoformat(lv["end"])
            except Exception:
                continue
            if lv["name"] == selected_name and ls <= d <= le:
                is_my_leave = True
                leave_type_my = lv["type"]
                break

        # 대근 필요 체크 (다른 사람 휴가인데 내가 근무 중이면 대근)
        sub_for = ""
        sub_for_shift = ""  # 결근자의 원래 근무
        if not is_my_leave and my_shift != "휴무":
            for lv in leave_list:
                if lv["name"] == selected_name:
                    continue
                try:
                    ls = datetime.date.fromisoformat(lv["start"])
                    le = datetime.date.fromisoformat(lv["end"])
                except Exception:
                    continue
                if ls <= d <= le:
                    sub_for = lv["name"]
                    # 결근자가 원래 어떤 근무였는지 파악
                    absent_name = lv["name"]
                    if shift["1근_근무자"] == absent_name:
                        sub_for_shift = "1근"
                    elif shift["2근_근무자"] == absent_name:
                        sub_for_shift = "2근"
                    elif shift["3근_근무자"] == absent_name:
                        sub_for_shift = "3근"
                    break

        # 주간/야간 대근 판별
        # 3근 결근 시: 1근→주간, 2근→야간
        # 1근 또는 2근 결근 시: 3근→야간, 나머지→주간
        sub_role = ""
        if sub_for:
            if sub_for_shift == "3근":
                sub_role = "주간 대근" if my_shift == "1근" else "야간 대근"
            else:
                sub_role = "야간 대근" if my_shift == "3근" else "주간 대근"

        return {
            "shift": my_shift,
            "is_leave": is_my_leave,
            "leave_type": leave_type_my,
            "sub_for": sub_for,
            "sub_for_shift": sub_for_shift,
            "sub_role": sub_role,
            "raw": shift,
        }

    SHIFT_TIME = {"1근": "06:30~14:30", "2근": "14:30~22:30", "3근": "22:30~06:30", "휴무": "",
                  "주간": "07:00~19:00", "야간": "19:00~07:00"}
    SHIFT_COLOR = {"1근": "#1D4ED8", "2근": "#15803D", "3근": "#B91C1C", "휴무": "#374151",
                   "주간": "#1D4ED8", "야간": "#B91C1C"}

    # ── 달력 설정 ──
    first_day = datetime.date(selected_year, selected_month, 1)
    last_day = datetime.date(selected_year, selected_month, _cal.monthrange(selected_year, selected_month)[1])
    start_col = (first_day.weekday() + 1) % 7  # 일요일=0

    # 선택 날짜 상태관리
    _skey = f"sched_sel_{selected_year}_{selected_month}"
    if _skey not in st.session_state:
        st.session_state[_skey] = today.day if (today.year == selected_year and today.month == selected_month) else 1
    sel_day_num = st.session_state[_skey]

    # 범례 (Samsung Calendar 스타일)
    if shift_type == "4조2교대":
        _legend_html = """
<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;">
  <span style="display:inline-flex;align-items:center;gap:5px;background:#FFC107;border-radius:20px;padding:3px 12px;font-size:13px;color:#3D2A00;font-weight:700;">주 주간</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#212121;border-radius:20px;padding:3px 12px;font-size:13px;color:#fff;font-weight:700;">야 야간</span>
  <span style="display:inline-flex;align-items:center;gap:5px;border:1px solid #E5E7EB;border-radius:20px;padding:3px 12px;font-size:13px;color:#EF4444;font-weight:700;">휴 휴무</span>
</div>"""
    elif shift_type in ("3조3교대", "2조2교대"):
        _legend_html = """
<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;">
  <span style="display:inline-flex;align-items:center;gap:5px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:20px;padding:3px 12px;font-size:13px;color:#1D4ED8;font-weight:700;">1근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:20px;padding:3px 12px;font-size:13px;color:#15803D;font-weight:700;">2근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#FEF2F2;border:1px solid #FECACA;border-radius:20px;padding:3px 12px;font-size:13px;color:#B91C1C;font-weight:700;">3근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;border:1px solid #E5E7EB;border-radius:20px;padding:3px 12px;font-size:13px;color:#EF4444;font-weight:700;">휴 휴무(주말)</span>
</div>"""
    else:
        _legend_html = """
<div style="display:flex;flex-wrap:wrap;gap:8px;margin-bottom:10px;">
  <span style="display:inline-flex;align-items:center;gap:5px;background:#EFF6FF;border:1px solid #BFDBFE;border-radius:20px;padding:3px 12px;font-size:13px;color:#1D4ED8;font-weight:700;">1근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#F0FDF4;border:1px solid #BBF7D0;border-radius:20px;padding:3px 12px;font-size:13px;color:#15803D;font-weight:700;">2근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#FEF2F2;border:1px solid #FECACA;border-radius:20px;padding:3px 12px;font-size:13px;color:#B91C1C;font-weight:700;">3근</span>
  <span style="display:inline-flex;align-items:center;gap:5px;border:1px solid #E5E7EB;border-radius:20px;padding:3px 12px;font-size:13px;color:#9CA3AF;font-weight:700;">휴무</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#FFFBEB;border:1px solid #FDE68A;border-radius:20px;padding:3px 12px;font-size:13px;color:#92400E;font-weight:700;">🏖 휴가</span>
  <span style="display:inline-flex;align-items:center;gap:5px;background:#F5F3FF;border:1px solid #DDD6FE;border-radius:20px;padding:3px 12px;font-size:13px;color:#6D28D9;font-weight:700;">🔄 대근</span>
</div>"""
    st.markdown(_legend_html, unsafe_allow_html=True)

    # ── 달력 CSS (Samsung Calendar 라이트 스타일) ──
    # column 단위 색상 적용: 각 날짜는 st.columns(7)의 독립 column 안에 있음
    st.markdown("""<style>
div[data-testid="column"]:has(.cmark) button {
    min-height: 44px !important; max-height: 44px !important;
    font-size: 17px !important; font-weight: 800 !important;
    padding: 2px 3px !important; white-space: nowrap !important;
    border-radius: 0 0 10px 10px !important;
    border-top: none !important;
    width: 100% !important;
    transition: filter 0.1s !important;
    margin-top: -6px !important;
}
div[data-testid="column"]:has(.cmark) button:hover {
    filter: brightness(0.92) !important;
}
div[data-testid="column"]:has(.cs-day) button {
    background: #FFC107 !important; color: #3D2A00 !important;
    border: 1.5px solid #FFB300 !important;
}
div[data-testid="column"]:has(.cs-night) button {
    background: #1a1a1a !important; color: #ffffff !important;
    border: 1.5px solid #000000 !important;
}
div[data-testid="column"]:has(.cs-off) button {
    background: #ffffff !important; color: #EF4444 !important;
    border: 1.5px solid #E5E7EB !important;
}
div[data-testid="column"]:has(.cs1) button {
    background: #EFF6FF !important; color: #1D4ED8 !important;
    border: 1.5px solid #BFDBFE !important;
}
div[data-testid="column"]:has(.cs2) button {
    background: #F0FDF4 !important; color: #15803D !important;
    border: 1.5px solid #BBF7D0 !important;
}
div[data-testid="column"]:has(.cs3) button {
    background: #FEF2F2 !important; color: #B91C1C !important;
    border: 1.5px solid #FECACA !important;
}
div[data-testid="column"]:has(.cs-leave) button {
    background: #FFFBEB !important; color: #92400E !important;
    border: 1.5px solid #FDE68A !important;
}
div[data-testid="column"]:has(.cs-sub) button {
    background: #F5F3FF !important; color: #6D28D9 !important;
    border: 1.5px solid #DDD6FE !important;
}
div[data-testid="column"]:has(.csel) button {
    box-shadow: inset 0 0 0 2.5px #3B82F6 !important;
}
div[data-testid="column"]:has(.ctoday) button {
    box-shadow: inset 0 0 0 2.5px #3B82F6 !important;
}
</style>""", unsafe_allow_html=True)

    # ── 월별 메모 일괄 로드 (달력 셀 표시용, 4조3교대만) ──
    _month_notes = {}
    if shift_type == "4조3교대":
        _mn_key = f"sched_month_notes_{selected_name}_{selected_year}_{selected_month}"
        if _mn_key not in st.session_state:
            try:
                from utils.supabase_db import load_schedule_notes_month as _lsnm
                st.session_state[_mn_key] = _lsnm(selected_name, selected_year, selected_month)
            except Exception:
                pass  # 실패 시 캐시하지 않음 (다음 렌더에서 재시도)
        _month_notes = st.session_state.get(_mn_key, {})

    # ── 근태관리 특이사항 메모 연동 (전체 달력) ──
    import json as _json_fs
    _att_memo_key = f"att_cal_memos_{selected_year}_{selected_month}"
    if _att_memo_key not in st.session_state:
        try:
            from utils.supabase_db import load_schedule_note as _att_ln
            _raw = _att_ln("전체", datetime.date(selected_year, selected_month, 1))
            st.session_state[_att_memo_key] = _json_fs.loads(_raw) if _raw else []
        except Exception:
            st.session_state[_att_memo_key] = []
    _cal_memo_dates_fs = {}
    for _m in st.session_state[_att_memo_key]:
        try:
            _ms = datetime.date.fromisoformat(str(_m.get("start", "")))
            _me = datetime.date.fromisoformat(str(_m.get("end", _m.get("start", ""))))
            _md = _ms
            while _md <= _me:
                if _md.year == selected_year and _md.month == selected_month:
                    _cal_memo_dates_fs.setdefault(_md.strftime("%Y-%m-%d"), []).append(_m.get("text", ""))
                _md += datetime.timedelta(days=1)
        except Exception:
            pass

    # ── 요일 헤더 ──
    WD_LABELS = ["일", "월", "화", "수", "목", "금", "토"]
    WD_COLORS = ["#EF4444", "#374151", "#374151", "#374151", "#374151", "#374151", "#3B82F6"]
    hdr_html = '<div style="display:grid;grid-template-columns:repeat(7,1fr);gap:4px;margin-bottom:4px;">'
    for wd, clr in zip(WD_LABELS, WD_COLORS):
        hdr_html += (
            f'<div style="text-align:center;padding:6px 0;font-size:13px;font-weight:700;color:{clr};">{wd}</div>'
        )
    hdr_html += '</div>'
    st.markdown(hdr_html, unsafe_allow_html=True)

    # ── 주(week) 목록 구성 ──
    weeks = []
    week = [None] * start_col
    for dn in range(1, last_day.day + 1):
        week.append(dn)
        if len(week) == 7:
            weeks.append(week)
            week = []
    if week:
        weeks.append(week + [None] * (7 - len(week)))

    # ── 근무 → CSS 클래스 & 버튼 텍스트 ──
    SHIFT_MCLS  = {"1근": "cs1", "2근": "cs2", "3근": "cs3", "휴무": "cs-off",
                   "주간": "cs-day", "야간": "cs-night"}
    SHIFT_LABEL = {"1근": "1근", "2근": "2근", "3근": "3근", "휴무": "휴",
                   "주간": "주", "야간": "야"}

    for wk in weeks:
        wcols = st.columns(7, gap="small")
        for ci, dn in enumerate(wk):
            with wcols[ci]:
                if dn is None:
                    st.markdown('<div style="height:90px;"></div>', unsafe_allow_html=True)
                else:
                    d = datetime.date(selected_year, selected_month, dn)
                    info = _get_day_info(d)
                    is_today = (d == today)
                    is_sel = (sel_day_num == dn)
                    wi = (d.weekday() + 1) % 7  # 0=일,6=토
                    holiday = _get_holiday_name(d)

                    # 근무 분류 → CSS 클래스 & 버튼 텍스트
                    if info["is_leave"]:
                        mcls = "cs-leave"
                        btn_txt = (info['leave_type'] or '휴가')[:3]
                    elif info["sub_for"]:
                        mcls = "cs-sub"
                        btn_txt = "대근"
                    else:
                        s = info["shift"]
                        mcls = SHIFT_MCLS.get(s, "cs-off")
                        btn_txt = SHIFT_LABEL.get(s, s)

                    marker_cls = f"cmark {mcls}"
                    if is_sel:   marker_cls += " csel"
                    if is_today: marker_cls += " ctoday"

                    # 날짜 숫자 색 (공휴일/일요일=빨강, 토요일=파랑)
                    day_clr = "#EF4444" if (wi == 0 or holiday) else ("#3B82F6" if wi == 6 else "#111827")
                    if is_today:
                        num_html = (f'<span style="background:#3B82F6;color:#fff;border-radius:50%;'
                                    f'width:24px;height:24px;display:inline-flex;align-items:center;'
                                    f'justify-content:center;font-size:13px;font-weight:800;">{dn}</span>')
                    else:
                        num_html = f'<span style="font-size:15px;font-weight:700;color:{day_clr};">{dn}</span>'

                    # 공휴일/기념일 라벨
                    hol_html = ""
                    if holiday:
                        hol_html = (f'<div style="font-size:13px;font-weight:600;color:#EF4444;'
                                    f'white-space:nowrap;overflow:hidden;text-overflow:ellipsis;'
                                    f'line-height:1.2;margin-top:1px;">{holiday}</div>')

                    # 비고 메모 표시
                    _cell_note = _month_notes.get(d.strftime("%Y-%m-%d"), "")
                    note_html = ""
                    if _cell_note:
                        note_html = (
                            f'<div style="font-size:12px;color:#374151;margin-top:2px;'
                            f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'
                            f'line-height:1.3;font-weight:600;background:#FEF9C3;'
                            f'border-radius:3px;padding:1px 3px;" title="{_cell_note}">'
                            f'📝{_cell_note}</div>'
                        )

                    # 특이사항 메모 (근태관리 전체 달력 연동)
                    att_memo_html = ""
                    _d_str_fs = d.strftime("%Y-%m-%d")
                    if _d_str_fs in _cal_memo_dates_fs:
                        _atxt = " / ".join(_cal_memo_dates_fs[_d_str_fs])
                        _ashort = _atxt[:8] + ("…" if len(_atxt) > 8 else "")
                        att_memo_html = (
                            f'<div style="background:#FDE68A;color:#92400E;border-radius:3px;'
                            f'font-size:12px;padding:1px 3px;margin-top:1px;overflow:hidden;'
                            f'white-space:nowrap;text-overflow:ellipsis;font-weight:600;" title="{_atxt}">'
                            f'{_ashort}</div>'
                        )

                    # 대근 상세 표시 (날짜 아래 소형 텍스트)
                    sub_html = ""
                    if info["sub_for"]:
                        sub_html = (
                            f'<div style="font-size:9px;color:#7C3AED;margin-top:1px;'
                            f'overflow:hidden;text-overflow:ellipsis;white-space:nowrap;'
                            f'line-height:1.2;font-weight:600;">'
                            f'↺{info["sub_for"]} 휴가</div>'
                        )

                    ring = "2px solid #3B82F6" if (is_sel or is_today) else "1.5px solid #E5E7EB"
                    _has_extra = _cell_note or info["sub_for"] or (_d_str_fs in _cal_memo_dates_fs)
                    _cell_mh = "58px" if _has_extra else "44px"
                    st.markdown(
                        f'<div class="{marker_cls}" style="background:#ffffff;border:{ring};'
                        f'border-bottom:none;border-radius:10px 10px 0 0;padding:4px 5px 2px;min-height:{_cell_mh};">'
                        f'{num_html}{hol_html}{sub_html}{note_html}{att_memo_html}</div>',
                        unsafe_allow_html=True
                    )

                    if st.button(btn_txt, key=f"cd_{selected_year}_{selected_month}_{dn}", use_container_width=True):
                        st.session_state[_skey] = dn
                        st.rerun()
    # ── 월간 요약 ──
    if shift_type == "4조2교대":
        counts = {"주간": 0, "야간": 0, "휴무": 0}
        color_map = {"주간": "#1D4ED8", "야간": "#B91C1C", "휴무": "#6B7280"}
    elif shift_type in ("3조3교대", "2조2교대"):
        counts = {"1근": 0, "2근": 0, "3근": 0, "휴무": 0}
        color_map = {"1근": "#1D4ED8", "2근": "#15803D", "3근": "#B91C1C", "휴무": "#6B7280"}
    else:
        counts = {"1근": 0, "2근": 0, "3근": 0, "휴무": 0, "휴가": 0, "대근": 0}
        color_map = {"1근": "#1D4ED8", "2근": "#15803D", "3근": "#B91C1C", "휴무": "#6B7280", "휴가": "#D97706", "대근": "#7C3AED"}

    for dn in range(1, last_day.day + 1):
        d = datetime.date(selected_year, selected_month, dn)
        info = _get_day_info(d)
        if info["is_leave"]:
            if "휴가" in counts:
                counts["휴가"] += 1
        else:
            s = info["shift"]
            if s in counts:
                counts[s] += 1
        if info["sub_for"] and "대근" in counts:
            counts["대근"] += 1

    _grid_cols = len(counts)
    summary_html = f'<div style="display:grid;grid-template-columns:repeat({_grid_cols},1fr);gap:8px;margin-top:12px;margin-bottom:4px;">'
    for label, cnt in counts.items():
        summary_html += f'<div class="summary-card"><div class="sc-num" style="color:{color_map[label]}">{cnt}</div><div class="sc-label">{label}</div></div>'
    summary_html += "</div>"
    st.markdown(summary_html, unsafe_allow_html=True)

    # ── 선택 날짜 세부내역 ──
    st.markdown("---")
    sel_date = datetime.date(selected_year, selected_month, sel_day_num)
    info = _get_day_info(sel_date)
    raw = info["raw"]
    wd_names = ["월", "화", "수", "목", "금", "토", "일"]
    wd_str = wd_names[sel_date.weekday()]

    if info["is_leave"]:
        my_label = f"🌴 {info['leave_type'] or '휴가'}"
        my_color = "#D97706"
    elif info["sub_for"]:
        my_label = info["sub_role"]
        my_color = "#6D28D9"
    else:
        my_label = info["shift"]
        my_color = SHIFT_COLOR.get(info["shift"], "#374151")

    time_str = SHIFT_TIME.get(info["shift"], "") if not info["is_leave"] else ""

    detail_html = f'''
<div class="detail-card">
  <div style="font-size:18px;font-weight:700;color:#CDD6F4;margin-bottom:14px;">
    {selected_year}년 {selected_month}월 {sel_day_num}일 ({wd_str})
  </div>
  <div class="detail-row">
    <span style="color:#BAC2DE;width:80px;font-size:13px;">내 근무</span>
    <span class="shift-pill" style="background:{my_color};color:#fff;">{my_label}</span>
    <span style="color:#9399B2;font-size:13px;">{time_str}</span>
    {"<span style='background:#6D28D9;color:#EDE9FE;padding:3px 10px;border-radius:12px;font-size:12px;font-weight:700;'>🔄 대근</span>" if info["sub_for"] else ""}
  </div>'''

    if info["sub_for"]:
        detail_html += f'''
  <div class="detail-row">
    <span style="color:#BAC2DE;width:80px;font-size:13px;">대근 대상</span>
    <span style="color:#C4B5FD;font-weight:600;">{info["sub_for"]} 휴가로 인한 대근</span>
  </div>'''

    if shift_type == "4조3교대":
        # 전체 근무 배치 — 휴가자 있으면 대근 체제로 표시
        _vacationers_today = []
        for lv in leave_list:
            try:
                _ls = datetime.date.fromisoformat(lv["start"])
                _le = datetime.date.fromisoformat(lv["end"])
            except Exception:
                continue
            if _ls <= sel_date <= _le:
                _vacationers_today.append(lv)

        detail_html += '<div style="margin-top:14px;font-size:12px;color:#6C7086;font-weight:600;letter-spacing:1px;">전체 근무 배치</div>'

        if _vacationers_today:
            absent_names = {lv["name"] for lv in _vacationers_today}
            workers = []
            for shift_key, team_key, label in [
                ("1근_근무자", "1근_조", "1근"),
                ("2근_근무자", "2근_조", "2근"),
                ("3근_근무자", "3근_조", "3근"),
            ]:
                name = raw[shift_key]
                if name not in absent_names:
                    workers.append((name, raw[team_key], label))

            absent_shifts = set()
            for lv in _vacationers_today:
                nm = lv["name"]
                if raw["1근_근무자"] == nm: absent_shifts.add("1근")
                elif raw["2근_근무자"] == nm: absent_shifts.add("2근")
                elif raw["3근_근무자"] == nm: absent_shifts.add("3근")

            def _role(orig_shift):
                if "3근" in absent_shifts:
                    return "주간" if orig_shift == "1근" else "야간"
                return "야간" if orig_shift == "3근" else "주간"

            ROLE_COLOR = {"주간": "#1D6FA4", "야간": "#7C3AED"}
            ROLE_TIME  = {"주간": "06:30~22:30", "야간": "22:30~06:30"}

            for name, team, orig in workers:
                role = _role(orig)
                clr  = ROLE_COLOR[role]
                t    = ROLE_TIME[role]
                detail_html += f'''
  <div class="detail-row">
    <span style="background:{clr};color:#fff;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">{role}</span>
    <span style="color:#CDD6F4;">{name} ({team}조)</span>
    <span style="color:#6C7086;font-size:12px;">{t}</span>
  </div>'''

            for lv in _vacationers_today:
                detail_html += f'''
  <div class="detail-row">
    <span style="background:#92400E;color:#FEF3C7;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">🌴 {lv["type"]}</span>
    <span style="color:#9399B2;">{lv["name"]} — 휴가</span>
  </div>'''

            detail_html += f'''
  <div class="detail-row">
    <span style="background:#374151;color:#D1D5DB;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">휴무</span>
    <span style="color:#CDD6F4;">{raw["휴무_근무자"]} ({raw["휴무_조"]}조)</span>
    <span style="color:#6C7086;font-size:12px;">{raw["휴무_구분"]}</span>
  </div>'''
        else:
            detail_html += f'''
  <div class="detail-row">
    <span style="background:#1D4ED8;color:#fff;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">1근</span>
    <span style="color:#CDD6F4;">{raw["1근_근무자"]} ({raw["1근_조"]}조)</span>
    <span style="color:#6C7086;font-size:12px;">06:30~14:30</span>
  </div>
  <div class="detail-row">
    <span style="background:#15803D;color:#fff;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">2근</span>
    <span style="color:#CDD6F4;">{raw["2근_근무자"]} ({raw["2근_조"]}조)</span>
    <span style="color:#6C7086;font-size:12px;">14:30~22:30</span>
  </div>
  <div class="detail-row">
    <span style="background:#B91C1C;color:#fff;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">3근</span>
    <span style="color:#CDD6F4;">{raw["3근_근무자"]} ({raw["3근_조"]}조)</span>
    <span style="color:#6C7086;font-size:12px;">22:30~06:30</span>
  </div>
  <div class="detail-row">
    <span style="background:#374151;color:#D1D5DB;padding:2px 10px;border-radius:8px;font-size:12px;font-weight:700;">휴무</span>
    <span style="color:#CDD6F4;">{raw["휴무_근무자"]} ({raw["휴무_조"]}조)</span>
    <span style="color:#6C7086;font-size:12px;">{raw["휴무_구분"]}</span>
  </div>'''

    detail_html += "</div>"
    st.markdown(detail_html, unsafe_allow_html=True)

    # ── 특이사항 입력 (4조3교대 전용) ──
    if shift_type != "4조3교대":
        return
    from utils.supabase_db import save_schedule_note as _save_note, load_schedule_note as _load_note
    _note_key = f"sched_note_{selected_name}_{sel_date}"
    if _note_key not in st.session_state:
        try:
            st.session_state[_note_key] = _load_note(selected_name, sel_date)
        except Exception:
            st.session_state[_note_key] = ""

    def _on_note_submit():
        val = st.session_state.get(f"note_input_{selected_name}_{sel_date}", "")
        try:
            _save_note(selected_name, sel_date, val)
            st.session_state[_note_key] = val
            # 달력 셀 표시용 월별 캐시 무효화
            _mn_key2 = f"sched_month_notes_{selected_name}_{selected_year}_{selected_month}"
            st.session_state.pop(_mn_key2, None)
        except Exception as e:
            st.error(f"저장 실패: {e}")

    st.markdown("""<div style="background:#1E1E2E;border:1px solid #313244;border-radius:12px;padding:16px 20px;margin-top:10px;">
<div style="font-size:13px;color:#6C7086;font-weight:600;margin-bottom:8px;">📝 특이사항 메모</div>""", unsafe_allow_html=True)
    st.text_input(
        "특이사항 (엔터로 저장)",
        value=st.session_state[_note_key],
        key=f"note_input_{selected_name}_{sel_date}",
        placeholder="이 날짜에 대한 메모를 입력하세요...",
        on_change=_on_note_submit,
        label_visibility="collapsed",
    )
    if st.session_state[_note_key]:
        st.caption(f"✅ 저장됨: {st.session_state[_note_key]}")
    st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════
# ══════════════════════════════════════
# 근태관리 다이얼로그 (모듈 수준)
# ══════════════════════════════════════

@st.dialog("특이사항", width="large")
def _att_cal_note_dialog():
    info = st.session_state.get("_att_cal_note_info", {})
    d    = info.get("date")
    if not d:
        return

    # 기간 선택
    _dc1, _dc2 = st.columns(2)
    with _dc1:
        _ds = st.date_input("시작일", value=d, key="att_cal_note_ds")
    with _dc2:
        _de = st.date_input("종료일", value=d, key="att_cal_note_de")

    if _ds > _de:
        st.warning("종료일이 시작일보다 이릅니다.")
        _de = _ds

    _days_cnt = (_de - _ds).days + 1
    if _days_cnt > 1:
        st.caption(f"📅 {_days_cnt}일간 동일한 특이사항이 저장됩니다.")

    st.markdown("---")
    _new_note = st.text_area(
        "내용", value=info.get("note", ""), height=180,
        key="att_cal_note_ta",
        placeholder="특이사항을 입력하세요…",
        label_visibility="collapsed",
    )
    _nb1, _nb2 = st.columns(2)
    with _nb1:
        if st.button("저장", use_container_width=True, key="att_cal_note_sv", type="primary"):
            from utils.supabase_db import save_daily_detail as _sdd2, load_daily_detail as _ldd2
            import datetime as _dt2
            _cur = _ds
            while _cur <= _de:
                _fresh = _ldd2(_cur) or {}
                _sdd2(_cur, _fresh.get("shift", {}), _fresh.get("safety", []), _new_note)
                _cur += _dt2.timedelta(days=1)
            st.rerun()
    with _nb2:
        if st.button("닫기", use_container_width=True, key="att_cal_note_cl"):
            st.rerun()


@st.dialog("년월 선택", width="small")
def _att_date_picker_dialog():
    _init = st.session_state.get("_att_dlg_pick_init", datetime.date.today())
    _cur_y = _init.year
    _cur_m = _init.month
    _py, _pm = st.columns(2)
    with _py:
        _sel_y = st.selectbox("년", list(range(_cur_y - 5, _cur_y + 6)), index=5, key="att_pick_year")
    with _pm:
        _sel_m = st.selectbox("월", list(range(1, 13)), index=_cur_m - 1, key="att_pick_month")
    if st.button("이동", use_container_width=True, key="att_date_pick_go"):
        st.session_state["att_cal_year"]  = _sel_y
        st.session_state["att_cal_month"] = _sel_m
        st.rerun()


@st.dialog("휴가/연장 신청서", width="large")
def _att_lv_dialog():
    d = st.session_state.get("_att_lv_dlg_data", {})
    ALL_MEMBERS = d.get("ALL_MEMBERS", [])
    today_d = d.get("today", datetime.date.today())

    _TIME_30M = [f"{h:02d}:{m:02d}" for h in range(0, 24) for m in (0, 30)]

    _ra, _rb = st.columns(2)
    with _ra: _lvn = st.selectbox("대상자", ALL_MEMBERS, key="att_lv_name")
    with _rb: _lvt = st.selectbox(
        "구분",
        ["연장근무", "정기휴가", "연차", "특별휴가", "명휴", "생일휴가", "공가",
         "공상휴업", "산재", "휴직", "대휴", "교육", "결근", "조퇴", "외출", "청원휴가", "공휴"],
        key="att_lv_type",
    )

    if _lvt == "연장근무":
        _oc1, _oc2, _oc3, _oc4 = st.columns(4)
        with _oc1: _ot_ds = st.date_input("시작 날짜", today_d, key="att_ot_date_s")
        with _oc2:
            _ts_def = "06:30" if "06:30" in _TIME_30M else _TIME_30M[0]
            _ot_ts = st.selectbox("시작 시간", _TIME_30M, index=_TIME_30M.index(_ts_def), key="att_ot_time_s")
        with _oc3: _ot_de = st.date_input("종료 날짜", today_d, key="att_ot_date_e")
        with _oc4:
            _te_def = "18:30" if "18:30" in _TIME_30M else _TIME_30M[0]
            _ot_te = st.selectbox("종료 시간", _TIME_30M, index=_TIME_30M.index(_te_def), key="att_ot_time_e")

        try:
            _s_dt = datetime.datetime.fromisoformat(f"{_ot_ds}T{_ot_ts}")
            _e_dt = datetime.datetime.fromisoformat(f"{_ot_de}T{_ot_te}")
            _mins = int((_e_dt - _s_dt).total_seconds() / 60)
            if _mins > 0:
                st.info(f"연장 시간: **{_mins // 60}시간 {_mins % 60:02d}분** ({_ot_ts} ~ {_ot_te})")
            else:
                st.warning("종료가 시작보다 이릅니다.")
        except Exception:
            pass

        if st.button("✅ 등록", use_container_width=True, key="att_add_lv"):
            if "leave_list" not in st.session_state:
                st.session_state["leave_list"] = []
            st.session_state["leave_list"].append({
                "name": _lvn, "type": "연장근무",
                "start": f"{_ot_ds}T{_ot_ts}", "end": f"{_ot_de}T{_ot_te}", "sub": "",
            })
            try:
                from utils.supabase_db import save_leaves
                save_leaves(st.session_state["leave_list"])
            except Exception: pass
            st.success("등록되었습니다.")
            st.rerun()

    else:
        _rc, _rd = st.columns(2)
        with _rc: _lvs2 = st.date_input("시작일", today_d, key="att_lv_start")
        with _rd: _lve2 = st.date_input("종료일", today_d, key="att_lv_end")

        _gnh = []
        if _lvt == "공휴":
            try:
                import holidays as _hh4; _cd3 = _lvs2; _wd3 = ["월","화","수","목","금","토","일"]
                while _cd3 <= _lve2:
                    _krr = _hh4.SouthKorea(years=_cd3.year)
                    if _cd3 not in _krr: _gnh.append(f"{_cd3.strftime('%m/%d')}({_wd3[_cd3.weekday()]})")
                    _cd3 += datetime.timedelta(days=1)
            except Exception: pass
        if _gnh:
            st.warning(f"⚠️ 비공휴일 포함: **{', '.join(_gnh)}**")
            _gc = st.checkbox("⚠️ 확인 후 등록", key="att_gc")
        else:
            _gc = True

        if st.button("✅ 등록", use_container_width=True, key="att_add_lv"):
            if _lvs2 > _lve2: st.error("시작일이 종료일보다 늦습니다.")
            elif _gnh and not _gc: st.error("⛔ 확인 체크 필요")
            else:
                if "leave_list" not in st.session_state:
                    st.session_state["leave_list"] = []
                st.session_state["leave_list"].append({
                    "name": _lvn, "type": _lvt,
                    "start": _lvs2.isoformat(), "end": _lve2.isoformat(), "sub": "",
                })
                try:
                    from utils.supabase_db import save_leaves
                    save_leaves(st.session_state["leave_list"])
                except Exception: pass
                st.success("등록되었습니다.")
                st.rerun()

    if st.session_state.get("leave_list"):
        st.markdown("---")
        st.markdown("**등록된 일정 조회**")
        _lyrs = sorted({datetime.date.fromisoformat(lv["start"][:10]).year for lv in st.session_state["leave_list"]}, reverse=True)
        _lf1, _lf2 = st.columns(2)
        with _lf1: _fyr = st.selectbox("연도", _lyrs, index=0, key="att_lf_yr")
        with _lf2: _fmo = st.selectbox("월", list(range(1, 13)), index=today_d.month - 1, key="att_lf_mo", format_func=lambda m: f"{m}월")
        _fss = datetime.date(_fyr, _fmo, 1)
        _fnm = (_fmo % 12) + 1; _fny = _fyr + (1 if _fmo == 12 else 0)
        _fse = datetime.date(_fny, _fnm, 1) - datetime.timedelta(days=1)
        _shwn = [(i, lv) for i, lv in enumerate(st.session_state["leave_list"])
                 if datetime.date.fromisoformat(lv["end"][:10]) >= _fss and datetime.date.fromisoformat(lv["start"][:10]) <= _fse]
        if not _shwn: st.caption(f"{_fyr}년 {_fmo}월 등록 일정 없음")
        else:
            st.caption(f"{_fyr}년 {_fmo}월 — {len(_shwn)}건")
            for _ii2, _lv5 in _shwn:
                _ct2, _cd4 = st.columns([4, 1])
                with _ct2:
                    _disp_s = _lv5["start"].replace("T", " ")
                    _disp_e = _lv5["end"].replace("T", " ")
                    st.write(f"{_lv5['name']} | {_lv5['type']} | {_disp_s} ~ {_disp_e}")
                with _cd4:
                    if st.button("삭제", key=f"att_del_{_ii2}"):
                        _deld = st.session_state["leave_list"].pop(_ii2)
                        try:
                            from utils.supabase_db import save_leaves, delete_daily_details_for_leave
                            save_leaves(st.session_state["leave_list"])
                            _rc2 = delete_daily_details_for_leave(_deld)
                            if _rc2 > 0: st.toast(f"휴가 삭제 — 관련 저장 일지 {_rc2}일 초기화", icon="♻️")
                        except Exception: pass
                        st.rerun()


@st.dialog("근무 통계 상세", width="large")
def _att_stats_dialog():
    d = st.session_state.get("_att_dlg_data", {})
    if not d:
        st.error("데이터 없음"); return
    nm            = d["nm"]
    s             = d["s"]
    yr_lv         = d["yr_lv"]
    selected_year = d["year"]
    selected_month= d["month"]
    days_in_month = d["days_in_month"]
    daily_details = d["daily_details"]
    NIGHT_HOURS   = d["NIGHT_HOURS"]
    MEMBERS       = d["MEMBERS"]
    _kr2          = d["_kr2"]
    _SCOLS        = d["_SCOLS"]
    base_leaves   = d["base_leaves"]
    _mblks        = d["_mblks"]

    _조_map = {v: k + "조" for k, v in MEMBERS.items()}
    st.markdown(f"### {nm} ({_조_map.get(nm,'')}) — {selected_year}년 {selected_month}월")
    _tdh = sum(dd["시간"] for dd in s["대근내역"])

    def _fhh(v):
        try: v = float(v); return int(v) if v == int(v) else v
        except Exception: return v

    def _sal_tbl(rows):
        tots = {c: sum(r[c] for r in rows) for c in _SCOLS}
        tots["일별합계"] = sum(r["일별합계"] for r in rows)
        def _f(v): return f"{v:.2f}" if v else ""
        _hs = "".join(
            f'<th style="background:#4472C4;color:#fff;text-align:center;'
            f'padding:5px 3px;border:1px solid #2F5496;font-size:10px;white-space:nowrap;">{c}</th>'
            for c in _SCOLS)
        h = (f'<div style="overflow-x:auto;margin-top:8px;">'
             f'<table style="border-collapse:collapse;font-size:11px;width:100%;min-width:900px;">'
             f'<thead><tr>'
             f'<th rowspan="2" style="background:#4472C4;color:#fff;text-align:center;'
             f'padding:6px 4px;border:1px solid #2F5496;white-space:nowrap;">날짜</th>'
             f'<th colspan="{len(_SCOLS)}" style="background:#4472C4;color:#fff;'
             f'text-align:center;padding:6px 4px;border:1px solid #2F5496;">일일 급여시간</th>'
             f'<th rowspan="2" style="background:#2F5496;color:#fff;text-align:center;'
             f'padding:6px 4px;border:1px solid #2F5496;white-space:nowrap;">일별합계</th>'
             f'</tr><tr>{_hs}</tr></thead><tbody>')
        for i, r in enumerate(rows):
            bg = "#f0f4fb" if i % 2 == 0 else "#ffffff"
            cs = "".join(f'<td style="text-align:right;padding:3px 5px;border:1px solid #D0D7E4;background:{bg};">{_f(r[c])}</td>' for c in _SCOLS)
            h += (f'<tr><td style="text-align:center;padding:3px 5px;border:1px solid #D0D7E4;'
                  f'background:#EEF2FA;font-weight:600;">{r["날짜"]}</td>{cs}'
                  f'<td style="text-align:right;padding:3px 5px;border:1px solid #2F5496;'
                  f'background:#D9E1F2;font-weight:700;color:#1F3864;">{_f(r["일별합계"])}</td></tr>')
        tc = "".join(f'<td style="text-align:right;padding:4px 5px;border:1px solid #9DC3E6;background:#BDD7EE;font-weight:700;color:#1F3864;">{_f(tots[c])}</td>' for c in _SCOLS)
        h += (f'<tr><td style="text-align:center;padding:4px 5px;border:1px solid #9DC3E6;'
              f'background:#9DC3E6;font-weight:700;color:#1F3864;">근로별 월합계</td>{tc}'
              f'<td style="text-align:right;padding:4px 5px;border:1px solid #9DC3E6;'
              f'background:#9DC3E6;font-weight:700;color:#1F3864;">{_f(tots["일별합계"])}</td>'
              f'</tr></tbody></table></div>')
        return h

    if s["대근내역"]:
        with st.expander(f"대근 내역 ({s['대근횟수']}회 · 계 {_fhh(_tdh)}H)", expanded=False):
            for _dd in s["대근내역"]:
                st.write(f"{_dd['날짜']} | {_dd['구분']} | {_dd['휴가자']} {_dd['휴가구분']}으로 대근 | {_fhh(_dd['시간'])}H")
    else:
        st.caption("대근 없음")

    if s["휴가내역"]:
        with st.expander(f"{selected_month}월 휴가 내역 ({s['휴가일수']}일)", expanded=False):
            for _hh3 in s["휴가내역"]: st.write(_hh3)
    else:
        st.caption("이번달 휴가 없음")

    if yr_lv:
        with st.expander(f"{selected_year}년 전체 휴가 ({len(yr_lv)}일)", expanded=False):
            _tc2 = {}
            for _lv2 in yr_lv: _tc2[_lv2["구분"]] = _tc2.get(_lv2["구분"], 0) + 1
            st.info(" | ".join(f"{k}: {v}일" for k, v in _tc2.items()))
            for _lv2 in yr_lv: st.write(f"{_lv2['날짜']} — {_lv2['구분']}")

    with st.expander(f"{selected_month}월 급여시간표", expanded=False):
        _srows = []
        for _fd in range(1, days_in_month + 1):
            _fd_date = datetime.date(selected_year, selected_month, _fd)
            _fds = _fd_date.strftime("%Y-%m-%d")
            _fhol = _fd_date in _kr2
            _frow = {c: 0.0 for c in _SCOLS}; _frow["날짜"] = f"{selected_month:02d}/{_fd:02d}"
            if _fds in daily_details:
                _sh2 = daily_details[_fds].get("shift", {})
                if _sh2.get("is_2person"):
                    _lp=_sh2.get("leave_person","") or _sh2.get("3근_근무자","")
                    _dw3=_sh2.get("주간_근무자","") or _sh2.get("1근_근무자","")
                    _nw3=_sh2.get("야간_근무자","") or _sh2.get("2근_근무자","")
                    _lv3=_sh2.get("leave_type","") or _sh2.get("3근_비고","")
                    _nb=NIGHT_HOURS.get("야간",7.5)
                    _dot2=float(_sh2.get("1근_연장",4) or 4); _not2=float(_sh2.get("2근_연장",4) or 4)
                    if nm==_lp:
                        if _lv3=="공가": _frow["공가"]=8.0
                        elif _lv3!="공휴": _frow["휴가비근로"]=8.0
                    elif nm==_dw3:
                        if _fhol or _lv3=="공휴": _frow["유휴근로"]=8.0;_frow["휴일연장"]=_dot2;_frow["휴일비근로"]=8.0
                        else: _frow["정상근로"]=8.0;_frow["연장근로"]=_dot2
                    elif nm==_nw3:
                        if _fhol or _lv3=="공휴": _frow["유휴근로"]=8.0;_frow["야간근로"]=_nb;_frow["휴일연장"]=_not2;_frow["휴일비근로"]=8.0
                        else: _frow["정상근로"]=8.0;_frow["야간근로"]=_nb;_frow["연장근로"]=_not2
                else:
                    for _sk3,_근3 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        if _sh2.get(_sk3)==nm:
                            _nb3=NIGHT_HOURS.get(_근3,0);_ot3b=float(_sh2.get(f"{_근3}_연장",0) or 0)
                            _dy3=float(_sh2.get(f"{_근3}_주간연장") or _ot3b);_ny3=float(_sh2.get(f"{_근3}_야간연장") or 0)
                            if _fhol: _frow["유휴근로"]=8.0;_frow["휴일비근로"]=8.0;_frow["휴일연장"]=_dy3+_ny3;_frow["야간근로"]=_nb3 if _nb3>0 else 0
                            else: _frow["정상근로"]=8.0;_frow["연장근로"]=_dy3;_frow["야간근로"]=(_nb3+_ny3) if _nb3>0 else 0
                            break
            else:
                _fs=_shift_for_date(_fd_date,MEMBERS); _fs=_apply_leaves_stat(_fs,_fd_date,base_leaves)
                if _fs.get("is_2person"):
                    _nb4=NIGHT_HOURS.get("야간",7.5);_lp4=_fs.get("leave_person","");_dw4=_fs.get("주간_근무자","");_nw4=_fs.get("야간_근무자","");_lv4=_fs.get("leave_type","")
                    if nm==_lp4:
                        if _lv4=="공가": _frow["공가"]=8.0
                        elif _lv4!="공휴": _frow["휴가비근로"]=8.0
                    elif nm==_dw4:
                        if _fhol or _lv4=="공휴": _frow["유휴근로"]=8.0;_frow["휴일연장"]=4.0;_frow["휴일비근로"]=8.0
                        else: _frow["정상근로"]=8.0;_frow["연장근로"]=4.0
                    elif nm==_nw4:
                        if _fhol or _lv4=="공휴": _frow["유휴근로"]=8.0;_frow["야간근로"]=_nb4;_frow["휴일연장"]=4.0;_frow["휴일비근로"]=8.0
                        else: _frow["정상근로"]=8.0;_frow["야간근로"]=_nb4;_frow["연장근로"]=4.0
                else:
                    for _sk4,_근4 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        if _fs.get(_sk4)==nm:
                            _nb4b=NIGHT_HOURS.get(_근4,0)
                            if _fhol: _frow["유휴근로"]=8.0;_frow["휴일비근로"]=8.0;_frow["야간근로"]=_nb4b if _nb4b>0 else 0
                            else: _frow["정상근로"]=8.0;_frow["야간근로"]=_nb4b if _nb4b>0 else 0
                            break
            _ft=sum(_frow[c] for c in _SCOLS); _frow["일별합계"]=_ft
            if _ft>0: _srows.append(_frow)
        if _srows:
            st.markdown(_sal_tbl(_srows), unsafe_allow_html=True)
        else:
            st.info("저장된 근무 데이터가 없습니다.")

    with st.expander("교대주기별 연장 시간", expanded=False):
        st.caption("교대 주기(연속 근무 5일)별 연장 현황. 주기당 최대 12H — 초과 시 빨간색 경고.")
        if not _mblks:
            st.info("해당 월 교대 주기 없음")
        else:
            _OT=12; _n2=len(_mblks)
            _TH2="background:#374151;color:#D1D5DB;padding:5px 4px;border:1px solid #4B5563;text-align:center;font-size:12px;"
            _TDB="padding:6px 4px;border:1px solid #4B5563;text-align:center;font-size:13px;"
            _TDG=_TDB+"background:#1F2937;color:#9CA3AF;"
            _TDO=_TDB+"background:#1F2937;color:#E5E7EB;font-weight:600;"
            _TDR=_TDB+"background:#DC2626;color:#fff;font-weight:700;"
            _TDW=_TDB+"background:#78350F;color:#FDE68A;font-weight:600;"
            _dh="".join(f'<th colspan="3" style="{_TH2}">{b[0][0].strftime("%m/%d")}~{b[-1][0].strftime("%m/%d")}</th>' for b in _mblks)
            _cs2=('<th style="'+_TH2+'">연장(발생)</th><th style="'+_TH2+'">연장(잔여)</th><th style="'+_TH2+'">탄력근로 사용</th>')*_n2
            _dc=""; _aw=False
            for _b2 in _mblks:
                _bot=int(sum(_x[1] for _x in _b2)); _rem=max(0,_OT-_bot)
                _io=_bot>=_OT; _in2=not _io and _bot>=_OT-2; _aw=_aw or _io
                _cst=_TDR if _io else (_TDW if _in2 else _TDO)
                _dc+=f'<td style="{_cst}">{_bot}</td><td style="{_TDG}">{_rem}</td><td style="{_TDG}">0</td>'
            _hot=(f'<div style="overflow-x:auto;margin-top:8px;">'
                  f'<table style="border-collapse:collapse;font-size:12px;min-width:100%;">'
                  f'<thead><tr><th style="{_TH2}min-width:90px;">구분</th>{_dh}</tr>'
                  f'<tr><th style="{_TH2}"></th>{_cs2}</tr></thead>'
                  f'<tbody><tr><td style="{_TH2}text-align:left;white-space:nowrap;">연장/잔여(H)</td>{_dc}</tr>'
                  f'</tbody></table></div>')
            st.markdown(_hot, unsafe_allow_html=True)
            if _aw:
                _wl=[f'{b[0][0].strftime("%m/%d")}~{b[-1][0].strftime("%m/%d")} ({int(sum(_x[1] for _x in b))}H)' for b in _mblks if int(sum(_x[1] for _x in b))>=_OT]
                st.error("⚠️ 주 52시간 위배 주기: "+", ".join(_wl))


# 근태관리 (근무표 + 근무통계 통합)
# ══════════════════════════════════════
def page_attendance():
    import calendar as _cal

    MEMBERS = dict(st.secrets.get("members", {'A': '직원A', 'B': '직원B', 'C': '직원C', 'D': '직원D'}))
    ALL_MEMBERS = list(MEMBERS.values())
    today = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()

    # ── 전체화면 달력 모드 ──
    if st.session_state.get("att_fullscreen"):
        _bc, _ = st.columns([1, 4])
        with _bc:
            if st.button("← 돌아가기", key="att_back", use_container_width=True):
                st.session_state["att_fullscreen"] = False
                st.rerun()
        page_my_schedule()
        return

    # ── 연/월 session state 초기화 ──
    if "att_cal_year" not in st.session_state:
        st.session_state["att_cal_year"] = today.year
    if "att_cal_month" not in st.session_state:
        st.session_state["att_cal_month"] = today.month
    selected_year  = st.session_state["att_cal_year"]
    selected_month = st.session_state["att_cal_month"]

    # ── 헤더 행: 제목 | 휴가/연장신청서 | 근무형태 ──
    _h1, _h2, _h3, _h4 = st.columns([4, 1.5, 0.8, 0.4])
    with _h1:
        st.markdown("## 근태관리")
    with _h2:
        if st.button("휴가/연장 신청서", key="att_lv_open", use_container_width=True):
            st.session_state["_att_lv_dlg_data"] = {"ALL_MEMBERS": ALL_MEMBERS, "today": today}
            if "leave_list" not in st.session_state:
                try:
                    from utils.supabase_db import load_leaves as _ll2
                    st.session_state["leave_list"] = _ll2()
                except Exception:
                    st.session_state["leave_list"] = []
            _att_lv_dialog()
    with _h3:
        shift_type = st.selectbox(
            "근무형태",
            ["4조3교대", "3조3교대", "2조2교대", "4조2교대"],
            key="att_shift_type",
            help="다른근무형태 보기",
            label_visibility="collapsed",
        )
    with _h4:
        if st.button("🔄", key="att_refresh_btn", use_container_width=True, help="새로고침"):
            st.rerun()

    _SHIFT_TEAMS = {"4조3교대": ALL_MEMBERS, "3조3교대": ["A조","B조","C조"], "2조2교대": ["A조","B조"], "4조2교대": ["A조","B조","C조","D조"]}
    selected_name = _SHIFT_TEAMS[shift_type][0]

    # 휴가 목록
    leave_list = []
    if shift_type == "4조3교대":
        try:
            from utils.supabase_db import load_leaves as _ll
            leave_list = _ll()
        except Exception:
            pass

    days_in_month = _cal.monthrange(selected_year, selected_month)[1]

    # ── 달력 메모 로드 (연월 기준) ──
    import json as _json
    from utils.supabase_db import save_schedule_note as _cal_sn, load_schedule_note as _cal_ln
    _cal_memo_date = datetime.date(selected_year, selected_month, 1)
    _cal_memo_key = f"att_cal_memos_{selected_year}_{selected_month}"
    if _cal_memo_key not in st.session_state:
        try:
            _raw = _cal_ln("전체", _cal_memo_date)
            st.session_state[_cal_memo_key] = _json.loads(_raw) if _raw else []
        except Exception:
            st.session_state[_cal_memo_key] = []
    _cal_memos = st.session_state[_cal_memo_key]
    # 메모가 걸리는 날짜 집합 (calendar cell 하이라이트용) — 키: "YYYY-MM-DD" 문자열
    _memo_dates = {}
    for _m in _cal_memos:
        try:
            _ms = datetime.date.fromisoformat(str(_m.get("start", "")))
            _me = datetime.date.fromisoformat(str(_m.get("end", _m.get("start", ""))))
            _md = _ms
            while _md <= _me:
                if _md.year == selected_year and _md.month == selected_month:
                    _memo_dates.setdefault(_md.strftime("%Y-%m-%d"), []).append(_m.get("text", ""))
                _md += datetime.timedelta(days=1)
        except Exception:
            pass

    # --------- 통계 사전 계산 ---------
    NIGHT_HOURS = {"1근": 0, "2근": 0.5, "3근": 7.5, "주간": 0, "야간": 7.5}
    daily_details = {}
    base_leaves_s = []
    if shift_type == "4조3교대":
        try:
            from utils.supabase_db import load_daily_detail_month
            daily_details = load_daily_detail_month(selected_year, selected_month)
        except Exception:
            pass
        try:
            from utils.supabase_db import load_leaves as _ll_s
            base_leaves_s = _ll_s()
        except Exception:
            pass

    def _sff_s(val):
        try: return float(val)
        except Exception:
            import re as _re_s; nums = _re_s.findall(r"[\d.]+", str(val)); return float(nums[0]) if nums else 0

    stats = {nm: {"근무일수":0,"연장근로_대근":0,"연장근로_주간":0,"연장근로_야간":0,
                  "야간근로":0,"휴가일수":0,"대근횟수":0,"휴가내역":[],"대근내역":[]}
             for nm in ALL_MEMBERS}

    if shift_type == "4조3교대":
        for _day in range(1, days_in_month + 1):
            _date = datetime.date(selected_year, selected_month, _day)
            _dstr = _date.strftime("%Y-%m-%d")
            if _dstr in daily_details:
                _sh = daily_details[_dstr].get("shift", {})
                if _sh.get("is_2person"):
                    _dw = _sh.get("1근_근무자",""); _nw = _sh.get("2근_근무자","")
                    _lw = _sh.get("3근_근무자",""); _lt = _sh.get("3근_비고","")
                    if _dw in stats:
                        _dt = _sff_s(_sh.get("1근_연장", 4))
                        stats[_dw]["근무일수"]+=1; stats[_dw]["연장근로_대근"]+=4
                        stats[_dw]["연장근로_주간"]+=max(0,_sff_s(_sh.get("1근_주간연장",4))-4)
                        stats[_dw]["연장근로_야간"]+=_sff_s(_sh.get("1근_야간연장",0))
                        stats[_dw]["대근횟수"]+=1
                        stats[_dw]["대근내역"].append({"날짜":_dstr,"휴가자":_lw,"휴가구분":_lt,"시간":_dt,"구분":"주간"})
                    if _nw in stats:
                        _nt = _sff_s(_sh.get("2근_연장", 4))
                        stats[_nw]["근무일수"]+=1; stats[_nw]["연장근로_대근"]+=4
                        stats[_nw]["연장근로_주간"]+=_sff_s(_sh.get("2근_주간연장",0))
                        stats[_nw]["연장근로_야간"]+=max(0,_sff_s(_sh.get("2근_야간연장",4))-4)
                        stats[_nw]["야간근로"]+=8; stats[_nw]["대근횟수"]+=1
                        stats[_nw]["대근내역"].append({"날짜":_dstr,"휴가자":_lw,"휴가구분":_lt,"시간":_nt,"구분":"야간"})
                    if _lw in stats and _lt:
                        stats[_lw]["휴가일수"]+=1; stats[_lw]["휴가내역"].append(f"{_dstr}: {_lt}")
                else:
                    for _sk,_nk2 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        _wk=_sh.get(_sk,""); _ot=_sff_s(_sh.get(f"{_nk2}_연장",0))
                        if _wk in stats:
                            stats[_wk]["근무일수"]+=1
                            _ed=_sh.get(f"{_nk2}_주간연장"); _en=_sh.get(f"{_nk2}_야간연장")
                            if _ed is not None or _en is not None:
                                _dot,_not=_sff_s(_ed or 0),_sff_s(_en or 0)
                            elif _sh.get(f"{_nk2}_연장유형")=="야간연장":
                                _dot,_not=0,_ot
                            else:
                                _dot,_not=_ot,0
                            stats[_wk]["연장근로_주간"]+=_dot; stats[_wk]["연장근로_야간"]+=_not
                            stats[_wk]["야간근로"]+=NIGHT_HOURS.get(_nk2,0)+_not
                _ow=_sh.get("휴무_근무자",""); _ot2=_sh.get("휴무_구분","")
                if _ow in stats and _ot2 not in ("교대휴무","주휴휴무",""):
                    stats[_ow]["휴가일수"]+=1; stats[_ow]["휴가내역"].append(f"{_dstr}: {_ot2}")
            else:
                _ss=_shift_for_date(_date,MEMBERS)
                _ss=_apply_leaves_stat(_ss,_date,base_leaves_s)
                if _ss.get("is_2person"):
                    _dw2=_ss.get("주간_근무자",""); _nw2=_ss.get("야간_근무자","")
                    _lw2=_ss.get("leave_person",""); _lt2=_ss.get("leave_type","")
                    if _dw2 in stats:
                        stats[_dw2]["근무일수"]+=1; stats[_dw2]["연장근로_대근"]+=4
                        stats[_dw2]["대근횟수"]+=1
                        stats[_dw2]["대근내역"].append({"날짜":_dstr,"휴가자":_lw2,"휴가구분":_lt2,"시간":4,"구분":"주간"})
                    if _nw2 in stats:
                        stats[_nw2]["근무일수"]+=1; stats[_nw2]["연장근로_대근"]+=4
                        stats[_nw2]["야간근로"]+=8; stats[_nw2]["대근횟수"]+=1
                        stats[_nw2]["대근내역"].append({"날짜":_dstr,"휴가자":_lw2,"휴가구분":_lt2,"시간":4,"구분":"야간"})
                    if _lw2 in stats and _lt2:
                        stats[_lw2]["휴가일수"]+=1; stats[_lw2]["휴가내역"].append(f"{_dstr}: {_lt2} (예정)")
                else:
                    for _wk2,_sk2 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                        _w=_ss.get(_wk2,"")
                        if _w in stats:
                            stats[_w]["근무일수"]+=1; stats[_w]["야간근로"]+=_NIGHT_HOURS_BASE.get(_sk2,0)
                    _ow2=_ss.get("휴무_근무자",""); _ot3=_ss.get("휴무_구분","")
                    if _ow2 in stats and _ot3 not in ("교대휴무","주휴휴무",""):
                        stats[_ow2]["휴가일수"]+=1; stats[_ow2]["휴가내역"].append(f"{_dstr}: {_ot3} (예정)")

        year_leaves = {nm: [] for nm in ALL_MEMBERS}
        _today_kst = (datetime.datetime.utcnow() + datetime.timedelta(hours=9)).date()
        for _lv in base_leaves_s:
            try: _lvs=datetime.date.fromisoformat(_lv["start"]); _lve=datetime.date.fromisoformat(_lv["end"])
            except Exception: continue
            _cur2=_lvs
            while _cur2<=_lve and _cur2<=_today_kst:
                if _cur2.year==selected_year:
                    _ds2=_cur2.strftime("%Y-%m-%d")
                    if _lv["name"] in year_leaves:
                        year_leaves[_lv["name"]].append({"날짜":_ds2,"구분":_lv["type"]})
                _cur2+=datetime.timedelta(days=1)
    else:
        year_leaves = {nm: [] for nm in ALL_MEMBERS}

    # --------- 근무 통계 타이틀 + 이름 버튼 ---------
    st.markdown(f"### {selected_year}년 {selected_month}월 근무 통계")
    if shift_type != "4조3교대":
        st.info("상세 통계는 4조3교대 근무 형태에서만 지원됩니다.")
    else:
        _조_map = {v: k for k, v in MEMBERS.items()}
        _btn_cols = st.columns(len(ALL_MEMBERS))
        for _bi, _nm in enumerate(ALL_MEMBERS):
            with _btn_cols[_bi]:
                _조 = _조_map.get(_nm, "")
                if st.button(
                    f"{_nm}\n({_조}조)",
                    key=f"att_stat_{_nm}",
                    use_container_width=True,
                ):
                    _tk2 = next((k for k,v in MEMBERS.items() if v==_nm), None)
                    _mblks2 = []
                    _obd2 = {}
                    if _tk2:
                        _ms2 = datetime.date(selected_year, selected_month, 1)
                        _next_mo = selected_month % 12 + 1
                        _next_yr = selected_year + (1 if selected_month==12 else 0)
                        _me2 = datetime.date(_next_yr, _next_mo, 1) - datetime.timedelta(days=1)
                        _ss3 = _ms2 - datetime.timedelta(days=6)
                        _se3 = _me2 + datetime.timedelta(days=6)
                        for _dd2 in stats[_nm].get("대근내역", []):
                            _obd2[_dd2["날짜"]] = _obd2.get(_dd2["날짜"], 0) + _dd2["시간"]
                        for _ds5,_dd5 in daily_details.items():
                            if _ds5 in _obd2: continue
                            _sh5=_dd5.get("shift",{})
                            if not _sh5.get("is_2person"):
                                for _sk5,_ok5 in [("1근_근무자","1근"),("2근_근무자","2근"),("3근_근무자","3근")]:
                                    if _sh5.get(_sk5)==_nm:
                                        _ot5=_sff_s(_sh5.get(f"{_ok5}_연장",0))
                                        if _ot5>0: _obd2[_ds5]=_ot5
                                        break
                        _wseq2 = []
                        _dd6 = _ss3
                        while _dd6 <= _se3:
                            _idx3 = (_dd6-_BASE_DATE).days % 20
                            _c1b,_c2b,_c3b,_ = _CYCLE_20[_idx3]
                            if _tk2 in (_c1b,_c2b,_c3b):
                                _wseq2.append((_dd6, _obd2.get(_dd6.strftime("%Y-%m-%d"),0)))
                            _dd6 += datetime.timedelta(days=1)
                        _blks2 = []
                        if _wseq2:
                            _cb2=[_wseq2[0]]
                            for _ii3 in range(1,len(_wseq2)):
                                if (_wseq2[_ii3][0]-_wseq2[_ii3-1][0]).days==1: _cb2.append(_wseq2[_ii3])
                                else: _blks2.append(_cb2); _cb2=[_wseq2[_ii3]]
                            _blks2.append(_cb2)
                        _mblks2=[b for b in _blks2 if b[-1][0]>=_ms2 and b[0][0]<=_me2]

                    try:
                        import holidays as _hh2
                        _kr2=_hh2.SouthKorea(years=[selected_year,selected_year-1,selected_year+1])
                    except Exception:
                        _kr2=set()

                    _SCOLS=["정상근로","유휴근로","휴일근로","연장근로","휴일연장","야간근로","휴일비근로","휴가비근로","스틸아카데미","항군교육","사내교육(1)","사내교육(1.5)","사외교육(1)","사외교육(1.5)","공가"]

                    st.session_state["_att_dlg_data"] = {
                        "nm": _nm, "s": stats[_nm],
                        "yr_lv": year_leaves.get(_nm, []),
                        "year": selected_year, "month": selected_month,
                        "days_in_month": days_in_month,
                        "daily_details": daily_details,
                        "NIGHT_HOURS": NIGHT_HOURS,
                        "MEMBERS": MEMBERS, "ALL_MEMBERS": ALL_MEMBERS,
                        "_kr2": _kr2, "_SCOLS": _SCOLS,
                        "base_leaves": base_leaves_s,
                        "_obd": _obd2, "_mblks": _mblks2,
                    }
                    _att_stats_dialog()

    # --------- Galaxy 달력 ---------
    MCLS = {"1근": "#1565C0", "2근": "#2E7D32", "3근": "#C62828", "휴무": "#9E9E9E",
            "주간": "#E65100", "야간": "#37474F", "휴가": "#F57F17", "대근": "#6A1B9A"}

    def _mini_shift(d):
        if shift_type == "3조3교대":
            s = _shift_for_date_3s3(d, selected_name[0]); return s, MCLS.get(s, "#ccc")
        if shift_type == "2조2교대":
            s = _shift_for_date_2s2(d, selected_name[0]); return s, MCLS.get(s, "#ccc")
        if shift_type == "4조2교대":
            s = _shift_for_date_4s2(d, selected_name[0]); return s, MCLS.get(s, "#ccc")
        shift_d = _shift_for_date(d, MEMBERS)
        if shift_d["1근_근무자"] == selected_name: base = "1근"
        elif shift_d["2근_근무자"] == selected_name: base = "2근"
        elif shift_d["3근_근무자"] == selected_name: base = "3근"
        else: base = "휴무"
        for lv in leave_list:
            try:
                if lv["name"] == selected_name and datetime.date.fromisoformat(lv["start"]) <= d <= datetime.date.fromisoformat(lv["end"]):
                    return "휴가", MCLS["휴가"]
            except Exception: pass
        if base != "휴무":
            for lv in leave_list:
                try:
                    if lv["name"] != selected_name and datetime.date.fromisoformat(lv["start"]) <= d <= datetime.date.fromisoformat(lv["end"]):
                        return "대근", MCLS["대근"]
                except Exception: pass
        return base, MCLS.get(base, "#ccc")

    # 달력 헤더: < year month >  오늘  날짜선택
    st.markdown(
        "<style>"
        "[data-testid='stHorizontalBlock']:has([data-testid='att_cal_hdr']) button {"
        "  min-height:0!important; padding:4px 10px!important; font-size:13px!important;"
        "}"
        "</style>",
        unsafe_allow_html=True,
    )
    # 달력 헤더: < [년월 버튼] >  오늘
    _gcl_p, _gcl_title, _gcl_n, _gcl_sp, _gcl_today = st.columns([0.5, 4, 0.5, 0.5, 1])
    with _gcl_p:
        if st.button("❮", key="att_prev_mo", use_container_width=True):
            if selected_month == 1:
                st.session_state["att_cal_year"]  -= 1
                st.session_state["att_cal_month"]  = 12
            else:
                st.session_state["att_cal_month"] -= 1
            st.rerun()
    with _gcl_title:
        if st.button(
            f"{selected_year}년 {selected_month}월",
            key="att_ym_btn",
            use_container_width=True,
            help="클릭하면 날짜를 선택할 수 있습니다",
        ):
            st.session_state["_att_dlg_pick_init"] = datetime.date(selected_year, selected_month, 1)
            _att_date_picker_dialog()
    with _gcl_n:
        if st.button("❯", key="att_next_mo", use_container_width=True):
            if selected_month == 12:
                st.session_state["att_cal_year"]  += 1
                st.session_state["att_cal_month"]  = 1
            else:
                st.session_state["att_cal_month"] += 1
            st.rerun()
    with _gcl_sp:
        pass
    with _gcl_today:
        if st.button("오늘", key="att_today_btn", use_container_width=True):
            st.session_state["att_cal_year"]  = today.year
            st.session_state["att_cal_month"] = today.month
            st.rerun()

    # 이전달 마지막 날짜
    _prev_mo = selected_month - 1 if selected_month > 1 else 12
    _prev_yr = selected_year if selected_month > 1 else selected_year - 1
    _prev_days_in_mo = _cal.monthrange(_prev_yr, _prev_mo)[1]

    first_wd = (datetime.date(selected_year, selected_month, 1).weekday() + 1) % 7
    WD_LABELS = ["일", "월", "화", "수", "목", "금", "토"]
    WD_CLR    = ["#E53935", "#424242", "#424242", "#424242", "#424242", "#424242", "#1565C0"]

    # ── CSS: 달력 그리드 ──
    st.markdown("""<style>
    div[data-testid="stHorizontalBlock"]:has(div.cal-cell-curr),
    div[data-testid="stHorizontalBlock"]:has(div.cal-cell-dim) {
        gap: 0 !important; background: #fff;
    }
    div[data-testid="stColumn"]:has(div.cal-cell-curr) button {
        background: transparent !important;
        border: none !important;
        box-shadow: none !important;
        padding: 1px 4px 3px !important;
        min-height: 0 !important;
        height: auto !important;
        font-size: 12px !important;
        line-height: 1 !important;
        color: #CCCCCC !important;
        width: 100% !important;
        text-align: right !important;
    }
    div[data-testid="stColumn"]:has(div.cal-cell-curr) button:hover {
        color: #F57F17 !important;
    }
    </style>""", unsafe_allow_html=True)

    # ── 요일 헤더 ──
    st.markdown(
        '<div style="display:flex;background:#fff;border-radius:16px 16px 0 0;'
        'box-shadow:0 -2px 6px rgba(0,0,0,0.06);margin:6px 0 0;overflow:hidden;">'
        + "".join(
            f'<div style="flex:1;font-size:22px;font-weight:700;text-align:center;'
            f'padding:10px 2px 8px;border-bottom:2px solid #EEEEEE;color:{wc};">{wd}</div>'
            for wd, wc in zip(WD_LABELS, WD_CLR)
        )
        + "</div>",
        unsafe_allow_html=True,
    )

    # ── 셀 데이터 미리 계산 ──
    _all_cells = []
    for _pi in range(first_wd):
        _all_cells.append({"type": "dim", "day": _prev_days_in_mo - first_wd + 1 + _pi})

    for dn in range(1, days_in_month + 1):
        d      = datetime.date(selected_year, selected_month, dn)
        wi     = (d.weekday() + 1) % 7
        is_tod = (d == today)
        hol    = _get_holiday_name(d)
        _d_str = d.strftime("%Y-%m-%d")
        _, _   = _mini_shift(d), None  # mini_shift 호출 유지

        if is_tod:
            _dnh = (
                '<div style="text-align:center;margin-bottom:2px;">'
                '<span style="display:inline-flex;align-items:center;justify-content:center;'
                'background:#1A1A1A;color:#fff;border-radius:50%;'
                f'width:42px;height:42px;font-size:22px;font-weight:900;line-height:42px;">{dn}</span></div>'
            )
        else:
            _dc = "#E53935" if (wi == 0 or hol) else ("#1565C0" if wi == 6 else "#212121")
            _dnh = f'<div style="text-align:center;margin-bottom:2px;"><span style="font-size:28px;color:{_dc};font-weight:700;">{dn}</span></div>'

        # 배지
        _badge_html = ""
        if shift_type == "4조3교대":
            _base4 = _shift_for_date(d, MEMBERS)
            _lv4   = _apply_leaves_stat(_base4, d, leave_list)
            _4s_badges = []
            if _lv4.get("is_2person"):
                _slots4 = [
                    ("주간", _lv4.get("주간_근무자", ""), "#1565C0"),
                    ("야간", _lv4.get("야간_근무자", ""), "#C62828"),
                    ("휴가", _lv4.get("leave_person", ""), "#F57F17"),
                    ("휴무", _lv4.get("휴무_근무자", ""), "#9E9E9E"),
                ]
            elif _d_str in daily_details and daily_details[_d_str].get("shift", {}).get("1근_근무자"):
                _sh_d = daily_details[_d_str].get("shift", {})
                if _sh_d.get("is_2person"):
                    _slots4 = [
                        ("주간", _sh_d.get("주간_근무자") or _sh_d.get("1근_근무자", ""), "#1565C0"),
                        ("야간", _sh_d.get("야간_근무자") or _sh_d.get("2근_근무자", ""), "#C62828"),
                        ("휴무", _sh_d.get("휴무_근무자", ""), "#9E9E9E"),
                    ]
                else:
                    _slots4 = [
                        ("1근", _sh_d.get("1근_근무자", ""), "#1565C0"),
                        ("2근", _sh_d.get("2근_근무자", ""), "#2E7D32"),
                        ("3근", _sh_d.get("3근_근무자", ""), "#C62828"),
                        ("휴무", _sh_d.get("휴무_근무자", ""), "#9E9E9E"),
                    ]
            else:
                # shift 데이터 없으면 자동계산 사용
                _slots4 = [
                    ("1근", _base4.get("1근_근무자", ""), "#1565C0"),
                    ("2근", _base4.get("2근_근무자", ""), "#2E7D32"),
                    ("3근", _base4.get("3근_근무자", ""), "#C62828"),
                    ("휴무", _base4.get("휴무_근무자", ""), "#9E9E9E"),
                ]
            for _slbl, _snm, _sclr in _slots4:
                if _snm:
                    _short = _snm[:3]
                    _4s_badges.append(
                        f'<div style="color:{_sclr};border-left:3px solid {_sclr};'
                        f'background:transparent;font-size:11px;font-weight:600;padding:0 4px;margin-top:3px;'
                        f'line-height:1.5;display:block;max-width:100%;overflow:hidden;'
                        f'white-space:nowrap;text-overflow:ellipsis;">{_slbl}&thinsp;{_short}</div>'
                    )
            _badge_html = "".join(_4s_badges)
        else:
            _ALL_TEAM_LABELS = {"3조3교대":["A조","B조","C조"],"2조2교대":["A조","B조"],"4조2교대":["A조","B조","C조","D조"]}
            _TEAM_COLORS = {"A조":"#1565C0","B조":"#2E7D32","C조":"#C62828","D조":"#6A1B9A"}
            _SHIFT_FN = {"3조3교대":_shift_for_date_3s3,"2조2교대":_shift_for_date_2s2,"4조2교대":_shift_for_date_4s2}
            for _tm in _ALL_TEAM_LABELS.get(shift_type, []):
                _fn2 = _SHIFT_FN.get(shift_type)
                _ts2 = _fn2(d, _tm[0]) if _fn2 else ""
                if _ts2 and _ts2 != "휴무":
                    _tc2 = _TEAM_COLORS.get(_tm, "#888")
                    _badge_html += (
                        f'<div style="background:{_tc2};color:#fff;border-radius:4px;'
                        f'font-size:10px;font-weight:600;padding:1px 5px;margin-top:2px;'
                        f'display:inline-block;max-width:100%;overflow:hidden;white-space:nowrap;'
                        f'text-overflow:ellipsis;">{_tm}&nbsp;{_ts2}</div>'
                    )

        _memo_html = ""
        if _d_str in _memo_dates:
            for _mt in _memo_dates[_d_str][:1]:
                _txt = str(_mt)[:7]
                _memo_html += (
                    f'<div style="background:#F59E0B;color:#fff;border-radius:5px;'
                    f'font-size:10px;padding:2px 5px;margin-top:2px;'
                    f'display:inline-block;max-width:calc(100% - 2px);overflow:hidden;'
                    f'white-space:nowrap;text-overflow:ellipsis;">{_txt}</div>'
                )

        _hol_html = ""
        if hol:
            _hol_html = (
                f'<div style="font-size:12px;font-weight:700;color:#E53935;margin-top:2px;line-height:1.2;'
                f'overflow:hidden;white-space:nowrap;text-overflow:ellipsis;">{hol[:6]}</div>'
            )

        _note_txt = (daily_details.get(_d_str, {}).get("note") or "").strip()
        _note_prev_html = ""
        if _note_txt:
            # 줄바꿈 유지: 각 줄 10자 이하로 잘라 <br> 연결
            _np_lines = []
            for _nl in _note_txt.splitlines()[:4]:
                _nl = _nl.strip()
                if _nl:
                    _np_lines.append(_nl[:10] + ("…" if len(_nl) > 10 else ""))
            _np_html = "<br>".join(_np_lines)
            _note_prev_html = (
                f'<div style="font-size:11px;font-weight:700;color:#856404;background:#FFF9E6;'
                f'border-radius:3px;padding:1px 4px;margin-top:2px;overflow:hidden;">'
                f'{_np_html}</div>'
            )

        _all_cells.append({
            "type": "curr", "day": dn, "date": d, "date_str": _d_str,
            "dnh": _dnh, "badge": _badge_html, "hol": _hol_html,
            "memo": _memo_html, "note_prev": _note_prev_html,
            "note_txt": _note_txt,
            "bg": "#E8F0FE" if is_tod else "#fff",
        })

    _rem_cells = (7 - len(_all_cells) % 7) % 7
    for _ni in range(1, _rem_cells + 1):
        _all_cells.append({"type": "dim", "day": _ni})

    # ── 주별 렌더링 ──
    _trigger_note = None
    for _ws in range(0, len(_all_cells), 7):
        _week  = _all_cells[_ws:_ws + 7]
        _bdr   = "border-top:1px solid #F0F0F0;" if _ws > 0 else ""
        _wc2   = st.columns(7)
        for _ci, _cell in enumerate(_week):
            with _wc2[_ci]:
                if _cell["type"] == "dim":
                    st.markdown(
                        f'<div class="cal-cell-dim" style="{_bdr}min-height:155px;'
                        f'padding:8px 5px 4px;background:#FAFAFA;">'
                        f'<div style="text-align:center;font-size:28px;font-weight:700;color:#DCDCDC;">{_cell["day"]}</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    _c = _cell
                    st.markdown(
                        f'<div class="cal-cell-curr" style="{_bdr}min-height:130px;'
                        f'padding:8px 5px 0;background:{_c["bg"]};">'
                        f'{_c["dnh"]}'
                        f'<div style="display:inline-flex;gap:4px;align-items:flex-start;width:100%;flex-wrap:wrap;">'
                        f'<div style="flex:0 0 auto;min-width:0;overflow:hidden;">{_c["badge"]}{_c["memo"]}</div>'
                        f'<div style="flex:0 0 auto;min-width:0;overflow:hidden;">{_c["hol"]}{_c["note_prev"]}</div>'
                        f'</div>'
                        f'</div>',
                        unsafe_allow_html=True,
                    )
                    _btn_lbl = "📝" if _c["note_txt"] else "✎"
                    if st.button(
                        _btn_lbl,
                        key=f'cal_note_{_c["date_str"]}',
                        use_container_width=True,
                        help="특이사항 입력/보기",
                    ):
                        _trigger_note = _c

    if _trigger_note:
        _det = daily_details.get(_trigger_note["date_str"], {})
        st.session_state["_att_cal_note_info"] = {
            "date":     _trigger_note["date"],
            "date_str": _trigger_note["date_str"],
            "note":     _det.get("note", ""),
            "shift":    _det.get("shift", {}),
            "safety":   _det.get("safety", []),
        }
        _att_cal_note_dialog()



# ──────────────────────────────────────
# 재고 엑셀 생성 헬퍼
# ──────────────────────────────────────
def _make_inventory_excel(df, col_map, sheet_name="재고현황", notes_col=None):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    thin = Side(style="thin")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill(start_color="4B2D8E", end_color="4B2D8E", fill_type="solid")
    alt_fill = PatternFill(start_color="F3F0FF", end_color="F3F0FF", fill_type="solid")

    headers = [lbl for lbl, _ in col_map] + ["비고"]
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = Font(color="FFFFFF", bold=True)
        c.border = bd
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for ri, (_, row) in enumerate(df.iterrows(), 2):
        row_fill = alt_fill if ri % 2 == 0 else None
        for ci, (_, field) in enumerate(col_map, 1):
            val = ri - 1 if field == "#" else str(row.get(field, "") or "")
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = bd
            c.alignment = Alignment(vertical="center")
            if row_fill:
                c.fill = row_fill
        note_val = str(row.get(notes_col, "") or "") if notes_col else ""
        note = ws.cell(row=ri, column=len(col_map) + 1, value=note_val)
        note.border = bd
        if row_fill:
            note.fill = row_fill

    col_widths = {"#": 5, "product": 12, "lot": 18, "maker": 12,
                  "sector": 12, "registered": 16}
    for ci, (_, field) in enumerate(col_map, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(field, 12)
    ws.column_dimensions[get_column_letter(len(col_map) + 1)].width = 28  # 비고

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────
# 재고 관리 페이지
# ──────────────────────────────────────
def page_inventory():
    import requests as _req
    import pandas as _pd

    BACKEND = "https://kgcounter.up.railway.app"
    st.subheader("재고 현황")

    _half_hours = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]
    _tab_sector, _tab_history = st.tabs(["섹터별 현황", "날짜별 이력"])

    # ── 날짜별 이력 탭 ──────────────────────────────────────────────────────
    with _tab_history:
        _now_kst_h = datetime.datetime.utcnow() + datetime.timedelta(hours=9)
        _default_date = (_now_kst_h - datetime.timedelta(hours=6, minutes=30)).date()

        _hd_c1, _hd_c2 = st.columns(2)
        with _hd_c1:
            st.caption("시작")
            _hd_from_date = st.date_input("시작일", _default_date, key="inv_hd_from_date",
                                           min_value=datetime.date(2026, 1, 1),
                                           max_value=datetime.date(2100, 12, 31),
                                           label_visibility="collapsed")
            _hd_from_time = st.selectbox("시작시간", _half_hours, index=0, key="inv_hd_from_time",
                                          label_visibility="collapsed")
        with _hd_c2:
            st.caption("종료")
            _hd_to_date = st.date_input("종료일", _default_date, key="inv_hd_to_date",
                                         min_value=datetime.date(2026, 1, 1),
                                         max_value=datetime.date(2100, 12, 31),
                                         label_visibility="collapsed")
            _hd_to_time = st.selectbox("종료시간", _half_hours, index=len(_half_hours)-1, key="inv_hd_to_time",
                                        label_visibility="collapsed")

        _hist_from_dt = f"{_hd_from_date} {_hd_from_time}"
        _hist_to_dt   = f"{_hd_to_date} {_hd_to_time}"

        if st.button("조회", key="hist_fetch"):
            st.session_state["hist_data"] = None

        _hist_data = st.session_state.get("hist_data", None)
        _hist_loaded_key = st.session_state.get("hist_loaded_key", None)
        _hist_key = f"{_hist_from_dt}~{_hist_to_dt}"
        if _hist_data is None or _hist_loaded_key != _hist_key:
            try:
                _hr = _req.get(f"{BACKEND}/api/inventory/history",
                               params={"from_dt": _hist_from_dt, "to_dt": _hist_to_dt}, timeout=15)
                if _hr.ok:
                    _hist_data = _hr.json().get("history", [])
                    st.session_state["hist_data"] = _hist_data
                    st.session_state["hist_loaded_key"] = _hist_key
                else:
                    st.error(f"조회 실패: {_hr.status_code}")
                    _hist_data = []
            except Exception as _he:
                st.error(f"연결 오류: {_he}")
                _hist_data = []

        # 검색
        _hist_search = st.text_input("이력 검색", placeholder="LOT 또는 품명 입력...", key="hist_search", label_visibility="collapsed")

        def _hist_filter(items):
            if not _hist_search.strip():
                return items
            _s = _hist_search.strip().upper()
            return [h for h in items if _s in h.get("lot","").upper() or _s in h.get("product","").upper()]

        _act_new  = _hist_filter([h for h in _hist_data if h["action"] == "신규등록"])
        _act_line = _hist_filter([h for h in _hist_data if h["action"] == "라인입고"])
        _act_ret  = _hist_filter([h for h in _hist_data if h["action"] == "반품완료"])

        _hc1, _hc2, _hc3 = st.tabs([
            f"신규등록 ({len(_act_new)})",
            f"라인입고 ({len(_act_line)})",
            f"반품완료 ({len(_act_ret)})",
        ])

        def _render_hist_table(items, tab_key, show_from=False, show_to=False):
            if not items:
                st.info("해당 항목 없음")
                return []
            _hh0, _hh1, _hh2, _hh3, _hh4, _hh5 = st.columns([0.5, 2.5, 2, 1.5, 2, 1.5])
            _hh0.markdown("**✓**"); _hh1.markdown("**일시**"); _hh2.markdown("**LOT**")
            _hh3.markdown("**품명**"); _hh4.markdown("**제조사**")
            if show_from: _hh5.markdown("**이전섹터**")
            elif show_to: _hh5.markdown("**섹터**")
            selected = []
            for _idx, _hi in enumerate(items):
                _r0, _r1, _r2, _r3, _r4, _r5 = st.columns([0.5, 2.5, 2, 1.5, 2, 1.5])
                _chk = _r0.checkbox("선택", key=f"hchk_{tab_key}_{_idx}", label_visibility="collapsed")
                _r1.text(_hi["timestamp"][:16] if len(_hi["timestamp"]) >= 16 else _hi["timestamp"])
                _r2.text(_hi["lot"])
                _r3.text(_hi["product"])
                _r4.text(_hi["maker"])
                if show_from: _r5.text(_hi["from_sector"])
                elif show_to: _r5.text(_hi["to_sector"])
                if _chk:
                    selected.append(_hi)
            return selected

        def _hist_to_excel(items, sector_key):
            import io, openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "이력"
            headers = ["시각", "LOT번호", "품명", "제조사", "섹터", "비고"]
            purple = PatternFill("solid", fgColor="4B2D8E")
            thin = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = Font(bold=True, color="FFFFFF")
                c.fill = purple
                c.alignment = Alignment(horizontal="center")
                c.border = thin
            for ri, item in enumerate(items, 2):
                sector_val = item.get(sector_key, "")
                ts = item.get("timestamp", "")
                row_vals = [ts[:16] if len(ts) >= 16 else ts, item.get("lot",""), item.get("product",""), item.get("maker",""), sector_val, ""]
                fill = PatternFill("solid", fgColor="F5F0FF") if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                for ci, val in enumerate(row_vals, 1):
                    c = ws.cell(row=ri, column=ci, value=val)
                    c.border = thin
                    c.fill = fill
            ws.column_dimensions["F"].width = 20
            for col in ws.columns:
                if col[0].column_letter == "F":
                    continue
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)
            buf = io.BytesIO()
            wb.save(buf)
            return buf.getvalue()

        _fd = str(_hd_from_date)
        _td = str(_hd_to_date)
        with _hc1:
            _sel_new = _render_hist_table(_act_new, "new", show_to=True)
            if _act_new:
                _dl_new = _sel_new if _sel_new else _act_new
                _lbl_new = f"📥 선택 {len(_sel_new)}건 다운로드" if _sel_new else "📥 엑셀 다운로드 (전체)"
                st.download_button(_lbl_new, data=_hist_to_excel(_dl_new, "to_sector"),
                    file_name=f"신규등록_{_fd}_{_td}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with _hc2:
            _sel_line = _render_hist_table(_act_line, "line", show_from=True)
            if _act_line:
                _dl_line = _sel_line if _sel_line else _act_line
                _lbl_line = f"📥 선택 {len(_sel_line)}건 다운로드" if _sel_line else "📥 엑셀 다운로드 (전체)"
                st.download_button(_lbl_line, data=_hist_to_excel(_dl_line, "from_sector"),
                    file_name=f"라인입고_{_fd}_{_td}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)
        with _hc3:
            _sel_ret = _render_hist_table(_act_ret, "ret", show_from=True)
            if _act_ret:
                _dl_ret = _sel_ret if _sel_ret else _act_ret
                _lbl_ret = f"📥 선택 {len(_sel_ret)}건 다운로드" if _sel_ret else "📥 엑셀 다운로드 (전체)"
                st.download_button(_lbl_ret, data=_hist_to_excel(_dl_ret, "from_sector"),
                    file_name=f"반품완료_{_fd}_{_td}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # ── 섹터별 현황 탭 ──────────────────────────────────────────────────────
    with _tab_sector:
        _c_inv_sort, _c_inv_ref = st.columns([5, 0.8])
        with _c_inv_sort:
            sort_mode = st.radio("정렬", ["섹터별", "제조사별", "품목별", "LOT순", "등록시간순"], horizontal=True, key="inv_sort", label_visibility="collapsed")
        with _c_inv_ref:
            if st.button("🔄", key="inv_refresh", use_container_width=True, help="새로고침"):
                st.rerun()

    with _tab_sector:
            # 등록시간순 선택 시 기간 선택기 표시
        _inv_dt_from = None
        _inv_dt_to = None
        if sort_mode == "등록시간순":
            _tc1, _tc2 = st.columns(2)
            with _tc1:
                _inv_d_from = st.date_input("시작일", datetime.date.today() - datetime.timedelta(days=7), key="inv_d_from")
                _inv_h_from = st.selectbox("시작 시각", _half_hours, index=0, key="inv_h_from")
            with _tc2:
                _inv_d_to = st.date_input("종료일", datetime.date.today(), key="inv_d_to")
                _inv_h_to = st.selectbox("종료 시각", _half_hours, index=len(_half_hours)-1, key="inv_h_to")
            _inv_dt_from = datetime.datetime.combine(_inv_d_from, datetime.time(int(_inv_h_from[:2]), int(_inv_h_from[3:])))
            _inv_dt_to = datetime.datetime.combine(_inv_d_to, datetime.time(int(_inv_h_to[:2]), int(_inv_h_to[3:])))

        return_filter = ""

        try:
            _inv_r = _req.get(f"{BACKEND}/api/inventory/sectors", timeout=15)
            if not _inv_r.ok:
                st.error(f"조회 실패: {_inv_r.status_code}")
                return
            sectors_raw = _inv_r.json().get("sectors", {})
        except Exception as e:
            st.error(f"연결 오류: {e}")
            return

        # 전체 드럼 목록 (sector 컬럼 추가)
        all_drums = []
        for sector, drums in sectors_raw.items():
            for d in drums:
                all_drums.append({**d, "sector": sector})

        if not all_drums:
            st.info("보관 중인 드럼 없음")
            return

        df_all = _pd.DataFrame(all_drums)
        total = len(df_all)

        col_lbl, col_inp, col_cap = st.columns([0.6, 3, 1.5])
        col_lbl.markdown("**검색**")
        search = col_inp.text_input("검색", placeholder="품명 또는 LOT 일부 입력...", key="inv_search", label_visibility="collapsed")
        col_cap.caption(f"전체 **{total}드럼**")

        # 검색 필터
        if search.strip():
            s = search.strip().upper()
            mask = df_all["lot"].str.upper().str.contains(s, na=False) | df_all["product"].str.upper().str.contains(s, na=False)
            df_filtered = df_all[mask].copy()
        else:
            df_filtered = df_all.copy()

        df_filtered["_rsort"] = 0

        # 등록시간 파싱 및 기간 필터
        if sort_mode == "등록시간순":
            # tz·초 유무 혼합 형식 대응 (구 Google Sheets: 'YYYY-MM-DD HH:MM', Supabase: 'YYYY-MM-DD HH:MM:SS')
            _reg_str = df_filtered["registered"].astype(str).str[:19].str.replace("T", " ", regex=False)
            df_filtered["_reg_dt"] = _pd.to_datetime(_reg_str, errors="coerce", format="mixed")
            if _inv_dt_from and _inv_dt_to:
                _from_ts = _pd.Timestamp(_inv_dt_from)
                _to_ts = _pd.Timestamp(_inv_dt_to)
                _mask_dt = (df_filtered["_reg_dt"] >= _from_ts) & (df_filtered["_reg_dt"] <= _to_ts)
                df_filtered = df_filtered[_mask_dt]

        # 그룹 키 + 정렬
        if sort_mode == "LOT순":
            group_col = "_all"
            df_filtered["_all"] = "전체 (LOT순)"
            df_filtered = df_filtered.sort_values(["_rsort", "lot"])
        elif sort_mode == "등록시간순":
            group_col = "_all"
            df_filtered["_all"] = "전체 (등록시간순)"
            df_filtered = df_filtered.sort_values("_reg_dt", ascending=False, na_position="last")
        else:
            group_col = {"섹터별": "sector", "제조사별": "maker", "품목별": "product"}[sort_mode]
            df_filtered = df_filtered.sort_values(["_rsort", group_col, "lot"])


        _inv_col_map = [
            ("번호", "#"), ("품명", "product"), ("LOT번호", "lot"), ("제조사", "maker"),
            ("섹터", "sector"), ("등록시간", "registered"),
        ]
        import datetime as _dt_mod

        # 전체선택 / 선택 해제 버튼
        _btn_c1, _btn_c2 = st.columns([1, 1])
        if _btn_c1.button("전체선택", key="inv_selall", use_container_width=True):
            for _l in df_filtered["lot"].tolist():
                st.session_state[f"chk_{_l}"] = True
            st.rerun()
        if _btn_c2.button("선택 해제", key="inv_desel", use_container_width=True):
            for _k in list(st.session_state.keys()):
                if _k.startswith("chk_"):
                    st.session_state[_k] = False
            st.rerun()

        # 드럼별 체크박스 선택
        selected_lots = set()
        _seen_groups = []
        for group_key, group_df in df_filtered.groupby(group_col, sort=False):
            cnt = len(group_df)
            # 반품필터 활성 시 그룹 내 해당 반품 항목 목록
            grp_rf_lots = group_df[group_df["returnStatus"] == return_filter]["lot"].tolist() if return_filter else []
            grp_all_rf_selected = bool(grp_rf_lots) and all(st.session_state.get(f"chk_{_l}", False) for _l in grp_rf_lots)

            with st.expander(f"**{group_key}** — {cnt}드럼", expanded=False):
                # 그룹별 전체선택 버튼 (섹터별/제조사별/품목별)
                if sort_mode in ("섹터별", "제조사별", "품목별"):
                    _grp_lots = group_df["lot"].tolist()
                    _grp_all_sel = all(st.session_state.get(f"chk_{_l}", False) for _l in _grp_lots)
                    _grp_btn_label = f"선택해제 ({cnt})" if _grp_all_sel else f"전체선택 ({cnt})"
                    if st.button(_grp_btn_label, key=f"grpsel_{group_key}", type="secondary"):
                        for _l in _grp_lots:
                            st.session_state[f"chk_{_l}"] = not _grp_all_sel
                        st.rerun()

                # 헤더
                h1, h2, h3, h4, h5, h6 = st.columns([0.5, 1.5, 2, 1.5, 1.5, 1.8])
                h1.markdown("**선택**"); h2.markdown("**품명**"); h3.markdown("**LOT**")
                h4.markdown("**제조사**")
                if sort_mode not in ("섹터별",): h5.markdown("**섹터**")
                h6.markdown("**등록시간**")
                # 행
                for _, row in group_df.iterrows():
                    rs = row.get("returnStatus", "")
                    sd = row.get("scanDisabled", "")
                    return_emoji = "🔴" if rs == "불량" else "🟡" if rs == "기술" else "🔵" if rs == "무상" else ""
                    scan_badge = " `스캔불가`" if sd == "Y" else ""
                    c1, c2, c3, c4, c5, c6 = st.columns([0.5, 1.5, 2, 1.5, 1.5, 1.8])
                    checked = c1.checkbox("", key=f"chk_{row['lot']}", label_visibility="collapsed")
                    if checked:
                        selected_lots.add(row["lot"])
                    _pfx = f"{return_emoji} " if return_emoji else ""
                    c2.markdown(f"{_pfx}**{row.get('product','')}**{scan_badge}")
                    c3.text(row.get("lot", ""))
                    c4.text(row.get("maker", ""))
                    if sort_mode not in ("섹터별",):
                        c5.text(row.get("sector", ""))
                    c6.text(row.get("registered", ""))

        # 선택 항목 엑셀 다운로드
        if selected_lots:
            _sel_df = df_filtered[df_filtered["lot"].isin(selected_lots)].copy()
            _sel_df["_notes"] = _sel_df.apply(lambda r: ", ".join(filter(None, [
                r.get("returnStatus", "") or "",
                "스캔불가" if r.get("scanDisabled") == "Y" else "",
            ])), axis=1)
            _inv_excel_bytes = _make_inventory_excel(_sel_df, _inv_col_map, "재고현황", notes_col="_notes")
            _inv_fname = f"재고현황_{_dt_mod.date.today().strftime('%Y%m%d')}.xlsx"
            st.download_button("📥 선택 항목 엑셀 다운로드", data=_inv_excel_bytes, file_name=_inv_fname,
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="inv_excel_dl")

        # 액션 버튼바
        if selected_lots:
            st.markdown("---")
            selected_drums_list = [d for d in all_drums if d["lot"] in selected_lots]
            all_in_return = all(d.get("returnStatus") for d in selected_drums_list)
            st.warning(f"**{len(selected_lots)}드럼** 선택됨")

            # 편집 상태가 현재 선택과 다르면 자동 초기화
            if st.session_state.get("inv_confirm") == "edit" and st.session_state.get("inv_edit_lot") not in selected_lots:
                st.session_state.pop("inv_confirm", None)
                st.session_state.pop("inv_edit_lot", None)

            _inv_confirm = st.session_state.get("inv_confirm")

            # ── 재확인 화면 ──
            if _inv_confirm in ("checkout", "checkout_r", "return_done"):
                if _inv_confirm in ("checkout", "checkout_r"):
                    st.error(f"선택하신 **{len(selected_lots)}드럼** 제품이 라인입고 처리되어 목록에서 삭제됩니다. 계속하시겠습니까?")
                    _sector = "라인입고"
                    _success_msg = f"{len(selected_drums_list)}드럼 라인입고 완료!"
                else:
                    st.error(f"선택하신 반품 **{len(selected_lots)}드럼** 제품이 반품완료 처리되어 목록에서 삭제됩니다. 계속하시겠습니까?")
                    _sector = "반품완료"
                    _success_msg = f"{len(selected_drums_list)}드럼 반품완료 처리!"
                _yc, _nc = st.columns(2)
                if _yc.button("✅ 확인", type="primary", key="inv_confirm_yes"):
                    try:
                        res = _req.post(f"{BACKEND}/api/inventory/register",
                                        json={"drums": selected_drums_list, "sector": _sector}, timeout=15)
                        if res.ok:
                            st.success(_success_msg)
                            st.session_state.pop("inv_confirm", None)
                            st.rerun()
                        else:
                            st.error(f"실패: {res.text}")
                            st.session_state.pop("inv_confirm", None)
                    except Exception as e:
                        st.error(f"오류: {e}")
                        st.session_state.pop("inv_confirm", None)
                if _nc.button("❌ 취소", key="inv_confirm_no"):
                    st.session_state.pop("inv_confirm", None)
                    st.rerun()

            elif _inv_confirm == "edit":
                _edit_lot = st.session_state.get("inv_edit_lot")
                _edit_drum = next((d for d in all_drums if d["lot"] == _edit_lot), None)
                if _edit_drum is None:
                    st.error("수정할 드럼 정보를 찾을 수 없습니다.")
                    st.session_state.pop("inv_confirm", None)
                else:
                    st.info(f"✏️ **{_edit_drum['product']}** ({_edit_lot}) 정보 수정")
                    _ec1, _ec2 = st.columns(2)
                    _new_lot = _ec1.text_input("LOT번호", value=_edit_drum["lot"], key="edit_lot_inp")
                    _new_product = _ec2.text_input("품명", value=_edit_drum["product"], key="edit_prod_inp")
                    _ec3, _ec4 = st.columns(2)
                    _maker_list = ["고려(KCC)", "대한(노루)", "건설(제비)", "삼화", "애경", "동주(PPG)"]
                    _cur_mkr_idx = _maker_list.index(_edit_drum["maker"]) if _edit_drum["maker"] in _maker_list else 0
                    _new_maker = _ec3.selectbox("제조사", _maker_list, index=_cur_mkr_idx, key="edit_mkr_inp")
                    _sector_list = sorted(sectors_raw.keys())
                    _cur_sidx = _sector_list.index(_edit_drum["sector"]) if _edit_drum["sector"] in _sector_list else 0
                    _new_sector = _ec4.selectbox("섹터", _sector_list, index=_cur_sidx, key="edit_sec_inp")
                    _sy, _sn = st.columns(2)
                    if _sy.button("💾 저장", type="primary", key="inv_edit_save"):
                        try:
                            from utils.supabase_db import update_drum_fields as _udf
                            _udf(_edit_lot, _new_lot.strip(), _new_product.strip(), _new_maker.strip(), _new_sector)
                            st.success("수정 완료!")
                            st.session_state.pop("inv_confirm", None)
                            st.session_state.pop("inv_edit_lot", None)
                            st.rerun()
                        except Exception as _ee:
                            st.error(f"오류: {_ee}")
                            st.session_state.pop("inv_confirm", None)
                            st.session_state.pop("inv_edit_lot", None)
                    if _sn.button("❌ 취소", key="inv_edit_cancel"):
                        st.session_state.pop("inv_confirm", None)
                        st.session_state.pop("inv_edit_lot", None)
                        st.rerun()

            # ── 일반 버튼 ──
            elif all_in_return:
                # 반품 항목 선택: 반품완료 / 반품 해제 / 라인입고
                ba, bb, bc = st.columns(3)
                if ba.button(f"↩️ 반품완료 ({len(selected_lots)})", type="primary", key="btn_return_done"):
                    st.session_state["inv_confirm"] = "return_done"
                    st.rerun()
                if bb.button(f"🔓 반품 해제 ({len(selected_lots)})", key="btn_return_cancel"):
                    try:
                        from utils.supabase_db import set_return_status as _srs
                        _srs(selected_drums_list, "")
                        st.success(f"{len(selected_drums_list)}드럼 반품 해제!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류: {e}")
                if bc.button(f"라인입고 ({len(selected_lots)})", key="btn_checkout_r"):
                    st.session_state["inv_confirm"] = "checkout_r"
                    st.rerun()
                if len(selected_lots) == 1:
                    if st.button("✏️ 정보 수정", key="btn_edit_r"):
                        st.session_state["inv_confirm"] = "edit"
                        st.session_state["inv_edit_lot"] = next(iter(selected_lots))
                        st.rerun()
            else:
                # 입고존 드럼이 포함된 경우 스캔불가 토글 표시
                _sel_ingo = [d for d in selected_drums_list if d.get("sector") == "입고존"]
                _sel_ingo_dis = [d for d in _sel_ingo if d.get("scanDisabled") == "Y"]
                if _sel_ingo:
                    _sd_cols = st.columns([2, 2, 4])
                    if _sd_cols[0].button(f"스캔불가 설정 ({len(_sel_ingo)})", key="btn_sd_on"):
                        try:
                            res = _req.post(f"{BACKEND}/api/inventory/scan-disabled",
                                            json={"drums": _sel_ingo, "disabled": True}, timeout=15)
                            if res.ok:
                                st.success(f"{len(_sel_ingo)}드럼 스캔불가 설정!")
                                st.rerun()
                            else:
                                st.error(f"실패: {res.text}")
                        except Exception as e:
                            st.error(f"오류: {e}")
                    if _sel_ingo_dis and _sd_cols[1].button(f"스캔불가 해제 ({len(_sel_ingo_dis)})", key="btn_sd_off"):
                        try:
                            res = _req.post(f"{BACKEND}/api/inventory/scan-disabled",
                                            json={"drums": _sel_ingo_dis, "disabled": False}, timeout=15)
                            if res.ok:
                                st.success(f"{len(_sel_ingo_dis)}드럼 스캔불가 해제!")
                                st.rerun()
                            else:
                                st.error(f"실패: {res.text}")
                        except Exception as e:
                            st.error(f"오류: {e}")

                # 일반 항목 선택: 라인입고 + 반품 3종
                ca, cb, cc, cd = st.columns(4)
                if ca.button(f"라인입고 ({len(selected_lots)})", type="primary", key="btn_checkout"):
                    st.session_state["inv_confirm"] = "checkout"
                    st.rerun()
                if cb.button(f"🔴 불량반품 ({len(selected_lots)})", key="btn_return_bad"):
                    try:
                        from utils.supabase_db import set_return_status as _srs
                        _srs(selected_drums_list, "불량")
                        st.success(f"{len(selected_drums_list)}드럼 불량반품 등록!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류: {e}")
                if cc.button(f"🟡 기술반품 ({len(selected_lots)})", key="btn_return_tech"):
                    try:
                        from utils.supabase_db import set_return_status as _srs
                        _srs(selected_drums_list, "기술")
                        st.success(f"{len(selected_drums_list)}드럼 기술반품 등록!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류: {e}")
                if cd.button(f"🔵 무상반품 ({len(selected_lots)})", key="btn_return_free"):
                    try:
                        from utils.supabase_db import set_return_status as _srs
                        _srs(selected_drums_list, "무상")
                        st.success(f"{len(selected_drums_list)}드럼 무상반품 등록!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"오류: {e}")
                if len(selected_lots) == 1:
                    if st.button("✏️ 정보 수정", key="btn_edit"):
                        st.session_state["inv_confirm"] = "edit"
                        st.session_state["inv_edit_lot"] = next(iter(selected_lots))
                        st.rerun()


# ══════════════════════════════════════
# 반품 관리 페이지
# ══════════════════════════════════════
def page_inventory_return():
    import requests as _req
    import pandas as _pd

    BACKEND = "https://kgcounter.up.railway.app"
    _half_hours = [f"{h:02d}:{m:02d}" for h in range(24) for m in (0, 30)]
    st.subheader("반품 관리")
    st.caption("기술·불량·무상 반품 드럼 현황 및 처리")

    # 정렬 + 반품 필터 + 새로고침 한 줄
    if "ret_return_filter" not in st.session_state:
        st.session_state["ret_return_filter"] = ""
    _c_sort, _c_bad, _c_tech, _c_free, _c_ref = st.columns([4, 1, 1, 1, 0.8])
    with _c_sort:
        sort_mode = st.radio("정렬", ["섹터별", "제조사별", "품목별", "LOT순", "등록시간순"], horizontal=True, key="ret_sort", label_visibility="collapsed")
    for _rfk, _rfl, _col in [("불량", "🔴 불량", _c_bad), ("기술", "🟡 기술", _c_tech), ("무상", "🔵 무상", _c_free)]:
        _act = st.session_state["ret_return_filter"] == _rfk
        with _col:
            if st.button(_rfl, key=f"retbtn_{_rfk}", type="primary" if _act else "secondary", use_container_width=True):
                st.session_state["ret_return_filter"] = "" if _act else _rfk
                st.rerun()
    with _c_ref:
        if st.button("🔄", key="ret_refresh", use_container_width=True, help="새로고침"):
            st.rerun()
    return_filter = st.session_state["ret_return_filter"]

    # 등록시간순 선택 시 기간 선택기
    _ret_dt_from = None
    _ret_dt_to = None
    if sort_mode == "등록시간순":
        _rtc1, _rtc2 = st.columns(2)
        with _rtc1:
            _ret_d_from = st.date_input("시작일", datetime.date.today() - datetime.timedelta(days=30), key="ret_d_from")
            _ret_h_from = st.selectbox("시작 시각", _half_hours, index=0, key="ret_h_from")
        with _rtc2:
            _ret_d_to = st.date_input("종료일", datetime.date.today(), key="ret_d_to")
            _ret_h_to = st.selectbox("종료 시각", _half_hours, index=len(_half_hours)-1, key="ret_h_to")
        _ret_dt_from = datetime.datetime.combine(_ret_d_from, datetime.time(int(_ret_h_from[:2]), int(_ret_h_from[3:])))
        _ret_dt_to = datetime.datetime.combine(_ret_d_to, datetime.time(int(_ret_h_to[:2]), int(_ret_h_to[3:])))

    try:
        _inv_r2 = _req.get(f"{BACKEND}/api/inventory/sectors", timeout=15)
        if not _inv_r2.ok:
            st.error(f"조회 실패: {_inv_r2.status_code}")
            return
        sectors_raw = _inv_r2.json().get("sectors", {})
    except Exception as e:
        st.error(f"연결 오류: {e}")
        return

    all_drums = []
    for sector, drums in sectors_raw.items():
        for d in drums:
            all_drums.append({**d, "sector": sector})

    # ─── 반품 리스트 자동 매칭 ───────────────────────────────────
    with st.expander("📋 반품 리스트 자동 매칭 (이미지/엑셀 업로드)", expanded=False):
        st.caption("기술·무상·불량 반품 리스트를 업로드하면 일반 재고와 자동 매칭 후 반품대기 상태로 전환합니다.")
        rl_files = st.file_uploader(
            "반품 리스트 파일 선택 (여러 장 동시 업로드 가능)",
            type=["jpg", "jpeg", "png", "xlsx", "xls", "csv"],
            key="ret_rl_upload",
            accept_multiple_files=True,
        )
        if rl_files:
            st.caption(f"{len(rl_files)}개 파일 선택됨: " + ", ".join(f.name for f in rl_files))
            col_rl_btn, _ = st.columns([1, 3])
            if col_rl_btn.button("🔍 리스트 분석", key="ret_btn_rl_parse", use_container_width=True):
                import base64 as _b64
                all_items, parse_errors = [], []
                seen_lots = set()
                for _fi, rl_file in enumerate(rl_files):
                    with st.spinner(f"분석 중... ({_fi+1}/{len(rl_files)}) {rl_file.name}"):
                        file_b64 = _b64.b64encode(rl_file.read()).decode()
                        try:
                            res_rl = _req.post(
                                f"{BACKEND}/api/inventory/parse-return-list",
                                json={"file_data": file_b64, "filename": rl_file.name, "api_key": ""},
                                timeout=90,
                            )
                            if res_rl.ok:
                                for item in res_rl.json().get("items", []):
                                    if item["lot_no"] not in seen_lots:
                                        seen_lots.add(item["lot_no"])
                                        all_items.append(item)
                            else:
                                parse_errors.append(f"{rl_file.name}: {res_rl.text}")
                        except Exception as _e:
                            parse_errors.append(f"{rl_file.name}: {_e}")
                if parse_errors:
                    for _err in parse_errors:
                        st.error(f"분석 실패 — {_err}")
                if all_items:
                    st.session_state["ret_rl_parsed"] = all_items
                    st.session_state.pop("ret_rl_sel", None)
                    st.session_state.pop("ret_rl_confirm_pending", None)

        if st.session_state.get("ret_rl_parsed"):
            parsed = st.session_state["ret_rl_parsed"]
            normal_drums = [d for d in all_drums if not d.get("returnStatus")]
            lot_map = {d["lot"]: d for d in normal_drums}
            matched, unmatched = [], []
            for item in parsed:
                lot = item["lot_no"]
                if lot in lot_map:
                    matched.append({**lot_map[lot], "new_return_type": item["return_type"]})
                else:
                    unmatched.append(item)

            type_label = {"기술": "🟡 기술반품", "무상": "🔵 무상반품", "불량": "🔴 불량반품"}
            st.success(f"**추출 {len(parsed)}건** | 일반 재고 매칭 **{len(matched)}건** | 미매칭 {len(unmatched)}건")

            if matched:
                import pandas as _pdm
                df_m = _pdm.DataFrame(matched)

                ret_rl_sort = st.radio("정렬", ["제조사별", "품목별", "LOT순"], horizontal=True, key="ret_rl_sort")
                if ret_rl_sort == "제조사별":
                    df_m = df_m.sort_values(["maker", "lot"])
                elif ret_rl_sort == "품목별":
                    df_m = df_m.sort_values(["product", "lot"])
                else:
                    df_m = df_m.sort_values("lot")

                if "ret_rl_sel" not in st.session_state:
                    st.session_state["ret_rl_sel"] = {row["lot"]: True for _, row in df_m.iterrows()}

                c_toggle, c_cnt = st.columns([1.5, 3])
                all_checked = all(st.session_state["ret_rl_sel"].get(row["lot"], True) for _, row in df_m.iterrows())
                if c_toggle.button("전체 선택 해제" if all_checked else "전체 선택", key="ret_rl_all_toggle"):
                    st.session_state["ret_rl_sel"] = {row["lot"]: not all_checked for _, row in df_m.iterrows()}
                    st.rerun()

                h1, h2, h3, h4, h5 = st.columns([0.5, 1.5, 2, 1.5, 1.5])
                h1.markdown("**✓**"); h2.markdown("**품명**"); h3.markdown("**LOT**")
                h4.markdown("**제조사**"); h5.markdown("**반품유형**")

                sel_map = st.session_state.get("ret_rl_sel", {})
                for _, row in df_m.iterrows():
                    c1, c2, c3, c4, c5 = st.columns([0.5, 1.5, 2, 1.5, 1.5])
                    checked = c1.checkbox("", key=f"ret_rl_chk_{row['lot']}", value=sel_map.get(row["lot"], True), label_visibility="collapsed")
                    st.session_state["ret_rl_sel"][row["lot"]] = checked
                    c2.text(row.get("product", ""))
                    c3.text(row.get("lot", ""))
                    c4.text(row.get("maker", ""))
                    c5.markdown(type_label.get(row.get("new_return_type", ""), row.get("new_return_type", "")))

                selected_for_apply = [row for _, row in df_m.iterrows() if st.session_state.get("ret_rl_sel", {}).get(row["lot"], True)]
                c_cnt.caption(f"선택: {len(selected_for_apply)} / {len(matched)}건")

                if selected_for_apply:
                    if st.button(f"🔄 선택 {len(selected_for_apply)}드럼 → 반품대기 전환", type="primary", key="ret_btn_rl_confirm"):
                        st.session_state["ret_rl_confirm_pending"] = True
                        st.rerun()

                    if st.session_state.get("ret_rl_confirm_pending"):
                        st.warning(f"⚠️ 선택된 **{len(selected_for_apply)}드럼**을 반품대기 상태로 전환하시겠습니까?")
                        cy, cn = st.columns(2)
                        if cy.button("✅ 예, 전환합니다", type="primary", key="ret_rl_yes"):
                            groups = {}
                            for d in selected_for_apply:
                                rt = d.get("new_return_type", "무상")
                                groups.setdefault(rt, []).append(d)
                            ok_count, errs = 0, []
                            from utils.supabase_db import set_return_status as _srs
                            for status, grp in groups.items():
                                if not grp:
                                    continue
                                try:
                                    _srs([dict(d) for d in grp], status)
                                    ok_count += len(grp)
                                except Exception as _e2:
                                    errs.append(str(_e2))
                            if errs:
                                st.error(f"일부 오류: {errs}")
                            else:
                                st.success(f"{ok_count}드럼 반품대기 전환 완료!")
                                st.session_state.pop("ret_rl_parsed", None)
                                st.session_state.pop("ret_rl_sel", None)
                                st.session_state.pop("ret_rl_confirm_pending", None)
                                st.rerun()
                        if cn.button("❌ 취소", key="ret_rl_no"):
                            st.session_state.pop("ret_rl_confirm_pending", None)
                            st.rerun()

            if unmatched:
                import pandas as _pdum
                st.warning(f"⚠️ 아래 **{len(unmatched)}건**은 일반 재고에 없어 제외됩니다.")
                df_um = _pdum.DataFrame(unmatched)[["product", "lot_no", "return_type"]]
                df_um.columns = ["품명", "LOT-NO", "반품유형"]
                st.dataframe(df_um, use_container_width=True, hide_index=True)

    st.markdown("---")

    # 반품 항목만 필터
    return_drums = [d for d in all_drums if d.get("returnStatus")]
    if not return_drums:
        st.info("반품 드럼 없음")
        return

    df_all = _pd.DataFrame(return_drums)
    total = len(df_all)

    col_lbl, col_inp, col_cap = st.columns([0.6, 3, 1.5])
    col_lbl.markdown("**검색**")
    search = col_inp.text_input("검색", placeholder="품명 또는 LOT 일부...", key="ret_search", label_visibility="collapsed")
    col_cap.caption(f"반품 **{total}드럼**")

    if search.strip():
        s = search.strip().upper()
        mask = df_all["lot"].str.upper().str.contains(s, na=False) | df_all["product"].str.upper().str.contains(s, na=False)
        df_filtered = df_all[mask].copy()
    else:
        df_filtered = df_all.copy()

    # 반품 유형 필터
    if return_filter:
        df_filtered = df_filtered[df_filtered["returnStatus"] == return_filter].copy()

    # 등록시간 파싱 및 기간 필터
    if sort_mode == "등록시간순":
        _reg_str2 = df_filtered["registered"].astype(str).str[:19].str.replace("T", " ", regex=False)
        df_filtered["_reg_dt"] = _pd.to_datetime(_reg_str2, errors="coerce", format="mixed")
        if _ret_dt_from and _ret_dt_to:
            _from_ts = _pd.Timestamp(_ret_dt_from)
            _to_ts = _pd.Timestamp(_ret_dt_to)
            df_filtered = df_filtered[(df_filtered["_reg_dt"] >= _from_ts) & (df_filtered["_reg_dt"] <= _to_ts)]

    # 그룹 키 + 정렬
    if sort_mode == "LOT순":
        group_col = "_all"
        df_filtered["_all"] = "전체 (LOT순)"
        df_filtered = df_filtered.sort_values("lot")
    elif sort_mode == "등록시간순":
        group_col = "_all"
        df_filtered["_all"] = "전체 (등록시간순)"
        df_filtered = df_filtered.sort_values("_reg_dt", ascending=False, na_position="last")
    else:
        group_col = {"섹터별": "sector", "제조사별": "maker", "품목별": "product"}[sort_mode]
        df_filtered = df_filtered.sort_values([group_col, "lot"])

    if df_filtered.empty:
        st.info("해당 조건의 반품 드럼 없음")
        return

    st.markdown("---")

    _ret_col_map = [
        ("번호", "#"), ("품명", "product"), ("LOT번호", "lot"), ("제조사", "maker"),
        ("섹터", "sector"), ("등록시간", "registered"),
    ]
    import datetime as _dt_mod2

    # 전체선택 / 선택 해제
    _rb1, _rb2 = st.columns([1, 1])
    if _rb1.button("전체선택", key="ret_selall", use_container_width=True):
        for _l in df_filtered["lot"].tolist():
            st.session_state[f"ret_chk_{_l}"] = True
        st.rerun()
    if _rb2.button("선택 해제", key="ret_desel", use_container_width=True):
        for _l in df_filtered["lot"].tolist():
            st.session_state[f"ret_chk_{_l}"] = False
        st.rerun()

    # 드럼 목록
    selected_lots = set()
    for group_key, group_df in df_filtered.groupby(group_col, sort=False):
        cnt = len(group_df)
        with st.expander(f"**{group_key}** — {cnt}드럼", expanded=False):
            if sort_mode in ("섹터별", "제조사별", "품목별"):
                _rgrp_lots = group_df["lot"].tolist()
                _rgrp_all_sel = all(st.session_state.get(f"ret_chk_{_l}", False) for _l in _rgrp_lots)
                _rgrp_btn_label = f"선택해제 ({cnt})" if _rgrp_all_sel else f"전체선택 ({cnt})"
                if st.button(_rgrp_btn_label, key=f"ret_grpsel_{group_key}", type="secondary"):
                    for _l in _rgrp_lots:
                        st.session_state[f"ret_chk_{_l}"] = not _rgrp_all_sel
                    st.rerun()
            h1, h2, h3, h4, h5, h6 = st.columns([0.5, 1.5, 2, 1.5, 1.5, 1.8])
            h1.markdown("**선택**"); h2.markdown("**품명**"); h3.markdown("**LOT**")
            h4.markdown("**제조사**"); h5.markdown("**섹터**"); h6.markdown("**등록시간**")
            for _, row in group_df.iterrows():
                rs = row.get("returnStatus", "")
                emoji = "🔴" if rs == "불량" else "🟡" if rs == "기술" else "🔵" if rs == "무상" else ""
                c1, c2, c3, c4, c5, c6 = st.columns([0.5, 1.5, 2, 1.5, 1.5, 1.8])
                if c1.checkbox("", key=f"ret_chk_{row['lot']}", label_visibility="collapsed"):
                    selected_lots.add(row["lot"])
                c2.markdown(f"{emoji} **{row.get('product','')}**")
                c3.text(row.get("lot", ""))
                c4.text(row.get("maker", ""))
                c5.text(row.get("sector", ""))
                c6.text(row.get("registered", ""))

    # 선택 항목 엑셀 다운로드
    if selected_lots:
        _sel_ret_df = df_filtered[df_filtered["lot"].isin(selected_lots)].copy()
        _sel_ret_df["_notes"] = _sel_ret_df["returnStatus"].fillna("").astype(str)
        _ret_excel_bytes = _make_inventory_excel(_sel_ret_df, _ret_col_map, "반품관리", notes_col="_notes")
        _ret_fname = f"반품관리_{_dt_mod2.date.today().strftime('%Y%m%d')}.xlsx"
        st.download_button("📥 선택 항목 엑셀 다운로드", data=_ret_excel_bytes, file_name=_ret_fname,
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="ret_excel_dl")

    # 액션 버튼
    if selected_lots:
        st.markdown("---")
        selected_drums_list = [d for d in return_drums if d["lot"] in selected_lots]
        st.warning(f"**{len(selected_lots)}드럼** 선택됨")

        _ret_confirm = st.session_state.get("ret_confirm")

        # ── 재확인 화면 ──
        if _ret_confirm in ("ret_done", "ret_checkout"):
            if _ret_confirm == "ret_checkout":
                st.error(f"선택하신 **{len(selected_lots)}드럼** 제품이 라인입고 처리되어 목록에서 삭제됩니다. 계속하시겠습니까?")
                _ret_sector = "라인입고"
                _ret_ok_msg = f"{len(selected_drums_list)}드럼 라인입고 완료!"
            else:
                st.error(f"선택하신 반품 **{len(selected_lots)}드럼** 제품이 반품완료 처리되어 목록에서 삭제됩니다. 계속하시겠습니까?")
                _ret_sector = "반품완료"
                _ret_ok_msg = f"{len(selected_drums_list)}드럼 반품완료!"
            _ry, _rn = st.columns(2)
            if _ry.button("✅ 확인", type="primary", key="ret_confirm_yes"):
                try:
                    res = _req.post(f"{BACKEND}/api/inventory/register",
                                    json={"drums": selected_drums_list, "sector": _ret_sector}, timeout=15)
                    if res.ok:
                        st.success(_ret_ok_msg)
                        st.session_state.pop("ret_confirm", None)
                        st.rerun()
                    else:
                        st.error(f"실패: {res.text}")
                        st.session_state.pop("ret_confirm", None)
                except Exception as e:
                    st.error(f"오류: {e}")
                    st.session_state.pop("ret_confirm", None)
            if _rn.button("❌ 취소", key="ret_confirm_no"):
                st.session_state.pop("ret_confirm", None)
                st.rerun()

        # ── 일반 버튼 ──
        else:
            ba, bb, bc = st.columns(3)
            if ba.button(f"↩️ 반품완료 ({len(selected_lots)})", type="primary", key="ret_done"):
                st.session_state["ret_confirm"] = "ret_done"
                st.rerun()
            if bb.button(f"🔓 반품 해제 ({len(selected_lots)})", key="ret_cancel"):
                try:
                    from utils.supabase_db import set_return_status as _srs
                    _srs(selected_drums_list, "")
                    st.success(f"{len(selected_drums_list)}드럼 반품 해제!")
                    st.rerun()
                except Exception as e:
                    st.error(f"오류: {e}")
            if bc.button(f"라인입고 ({len(selected_lots)})", key="ret_checkout"):
                st.session_state["ret_confirm"] = "ret_checkout"
                st.rerun()


# ──────────────────────────────────────
# 메뉴 라우팅
# ──────────────────────────────────────
if page == "근태관리":
    page_attendance()
elif page == "일일 작업 일지":
    page_work_log()
elif page == "재고 현황":
    page_inventory()
elif page == "입고 관리":
    page_cross_check()
elif page == "반품 관리":
    page_inventory_return()
