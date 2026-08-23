# =====================================================================
# Project: paint-crosschecker
# Copyright (c) 2026 kmm851010-maker. All rights reserved.
# Unauthorized copying, modification, or distribution is strictly prohibited.
# =====================================================================
"""
서식 엑셀 리포트 자동 생성기
OpenPyXL을 활용하여 조건부 서식이 적용된 엑셀 보고서를 생성합니다.
"""

from io import BytesIO
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# 색상 정의
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FILL_MATCH = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_EXCESS = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
FILL_SHORT = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_MISSING = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

FONT_HEADER = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=11)
FONT_NORMAL = Font(name="맑은 고딕", size=10)
FONT_BOLD = Font(name="맑은 고딕", bold=True, size=10)
FONT_TITLE = Font(name="맑은 고딕", bold=True, size=14)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

STATUS_FILL_MAP = {
    "일치": FILL_MATCH,
    "초과": FILL_EXCESS,
    "부족": FILL_SHORT,
    "미입고": FILL_MISSING,
}

COLUMNS = [
    ("No.", 6),
    ("라인", 10),
    ("위치", 10),
    ("색상코드", 20),
    ("제조사", 15),
    ("계획수량", 12),
    ("입고수량", 12),
    ("차이", 10),
    ("상태", 12),
    ("총중량(kg)", 14),
]


def generate_report(result_df: pd.DataFrame) -> bytes:
    """교차검증 결과 DataFrame을 서식이 적용된 엑셀 파일로 생성합니다."""
    wb = Workbook()
    ws = wb.active
    ws.title = "KG스틸 페인트 입고 검증"

    # 타이틀
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "KG스틸 페인트 입고 검증 리포트"
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_CENTER

    # 날짜
    ws.merge_cells("A2:J2")
    date_cell = ws["A2"]
    date_cell.value = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    date_cell.font = FONT_NORMAL
    date_cell.alignment = Alignment(horizontal="right")

    # 헤더 (row 4)
    header_row = 4
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 데이터 rows
    if result_df.empty:
        ws.merge_cells(f"A{header_row + 1}:J{header_row + 1}")
        ws.cell(row=header_row + 1, column=1, value="검증 데이터가 없습니다.").font = FONT_NORMAL
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    for data_idx, (_, row) in enumerate(result_df.iterrows()):
        excel_row = header_row + 1 + data_idx
        status = str(row.get("상태", ""))
        row_fill = STATUS_FILL_MAP.get(status, None)

        values = [
            data_idx + 1,
            row.get("라인", ""),
            row.get("위치", ""),
            row.get("색상코드", ""),
            row.get("제조사", ""),
            row.get("계획수량", 0),
            row.get("입고수량", 0),
            row.get("차이", 0),
            status,
            row.get("총중량_kg", 0),
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val)
            cell.font = FONT_NORMAL
            cell.border = THIN_BORDER

            if col_idx in (1, 6, 7, 8, 10):
                cell.alignment = ALIGN_CENTER
            elif col_idx == 9:
                cell.alignment = ALIGN_CENTER
                cell.font = FONT_BOLD
            else:
                cell.alignment = Alignment(vertical="center")

            if row_fill:
                cell.fill = row_fill

    # 합계 행
    last_data_row = header_row + len(result_df)
    sum_row = last_data_row + 2

    ws.merge_cells(f"A{sum_row}:E{sum_row}")
    sum_label = ws.cell(row=sum_row, column=1, value="합계")
    sum_label.font = FONT_BOLD
    sum_label.alignment = ALIGN_CENTER
    sum_label.border = THIN_BORDER
    for c in range(2, 6):
        ws.cell(row=sum_row, column=c).border = THIN_BORDER

    # SUM 수식
    plan_start = f"F{header_row + 1}"
    plan_end = f"F{last_data_row}"
    actual_start = f"G{header_row + 1}"
    actual_end = f"G{last_data_row}"
    weight_start = f"J{header_row + 1}"
    weight_end = f"J{last_data_row}"

    sum_plan = ws.cell(row=sum_row, column=6, value=f"=SUM({plan_start}:{plan_end})")
    sum_plan.font = FONT_BOLD
    sum_plan.alignment = ALIGN_CENTER
    sum_plan.border = THIN_BORDER

    sum_actual = ws.cell(row=sum_row, column=7, value=f"=SUM({actual_start}:{actual_end})")
    sum_actual.font = FONT_BOLD
    sum_actual.alignment = ALIGN_CENTER
    sum_actual.border = THIN_BORDER

    diff_cell = ws.cell(row=sum_row, column=8, value=f"=G{sum_row}-F{sum_row}")
    diff_cell.font = FONT_BOLD
    diff_cell.alignment = ALIGN_CENTER
    diff_cell.border = THIN_BORDER

    ws.cell(row=sum_row, column=9).border = THIN_BORDER

    sum_weight = ws.cell(row=sum_row, column=10, value=f"=SUM({weight_start}:{weight_end})")
    sum_weight.font = FONT_BOLD
    sum_weight.alignment = ALIGN_CENTER
    sum_weight.border = THIN_BORDER

    # 범례
    legend_row = sum_row + 2
    ws.cell(row=legend_row, column=1, value="[범례]").font = FONT_BOLD
    legends = [
        ("일치", FILL_MATCH, "계획 수량 = 입고 수량"),
        ("초과", FILL_EXCESS, "입고 수량 > 계획 수량"),
        ("부족", FILL_SHORT, "입고 수량 < 계획 수량"),
        ("미입고", FILL_MISSING, "입고 내역 없음"),
    ]
    for i, (label, fill, desc) in enumerate(legends):
        r = legend_row + 1 + i
        cell_label = ws.cell(row=r, column=2, value=label)
        cell_label.fill = fill
        cell_label.font = FONT_BOLD
        cell_label.alignment = ALIGN_CENTER
        ws.cell(row=r, column=3, value=desc).font = FONT_NORMAL

    output = BytesIO()
    wb.save(output)
    return output.getvalue()
