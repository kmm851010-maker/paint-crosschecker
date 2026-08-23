# =====================================================================
# Project: paint-crosschecker
# Copyright (c) 2026 kmm851010-maker. All rights reserved.
# Unauthorized copying, modification, or distribution is strictly prohibited.
# =====================================================================
"""
서식 적용 엑셀 변환 모듈
추출된 테이블 데이터를 실무에 바로 사용 가능한
스타일이 적용된 엑셀(.xlsx) 파일로 변환합니다.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="2F3542", end_color="2F3542", fill_type="solid")
HEADER_FONT = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
DATA_FONT = Font(name="맑은 고딕", size=10)
SUM_FILL = PatternFill(start_color="DFE4EA", end_color="DFE4EA", fill_type="solid")
SUM_FONT = Font(name="맑은 고딕", size=10, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="A4B0BE"),
    right=Side(style="thin", color="A4B0BE"),
    top=Side(style="thin", color="A4B0BE"),
    bottom=Side(style="thin", color="A4B0BE"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")


def _is_numeric_column(rows: list, col_idx: int) -> bool:
    """해당 컬럼의 데이터가 대부분 숫자인지 판별합니다."""
    numeric_count = 0
    total = 0
    for row in rows:
        if col_idx < len(row):
            val = row[col_idx]
            if val is None or val == "":
                continue
            total += 1
            if isinstance(val, (int, float)):
                numeric_count += 1
    return total > 0 and numeric_count / total >= 0.5


def _is_sum_row(row: list) -> bool:
    """합계/소계 행인지 판별합니다."""
    for val in row:
        if isinstance(val, str) and any(k in val for k in ["합계", "소계", "TOTAL", "Total", "total", "SUM", "Sum"]):
            return True
    return False


def convert_to_excel(headers: list, rows: list) -> bytes:
    """
    헤더와 행 데이터를 서식 적용 엑셀 파일로 변환합니다.

    Args:
        headers: 컬럼명 리스트
        rows: 2D 데이터 리스트

    Returns:
        엑셀 파일 바이트
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "변환 결과"

    num_cols = len(headers)
    numeric_cols = set()
    for i in range(num_cols):
        if _is_numeric_column(rows, i):
            numeric_cols.add(i)

    # 헤더 행
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 데이터 행
    has_sum_row = False
    for row_idx, row_data in enumerate(rows, 2):
        is_sum = _is_sum_row(row_data)
        if is_sum:
            has_sum_row = True

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER

            if is_sum:
                cell.fill = SUM_FILL
                cell.font = SUM_FONT
            else:
                cell.font = DATA_FONT

            if col_idx - 1 in numeric_cols and isinstance(value, (int, float)):
                cell.alignment = RIGHT
                if isinstance(value, float) and value != int(value):
                    cell.number_format = "#,##0.0"
                else:
                    cell.number_format = "#,##0"
            else:
                cell.alignment = CENTER if col_idx - 1 in numeric_cols else LEFT

    # 합계 행이 없으면 자동 추가
    if not has_sum_row and numeric_cols and len(rows) > 0:
        sum_row = len(rows) + 2
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=sum_row, column=col_idx)
            cell.fill = SUM_FILL
            cell.font = SUM_FONT
            cell.border = THIN_BORDER

            if col_idx - 1 in numeric_cols:
                col_letter = get_column_letter(col_idx)
                cell.value = f"=SUM({col_letter}2:{col_letter}{sum_row - 1})"
                cell.number_format = "#,##0"
                cell.alignment = RIGHT
            elif col_idx == 1:
                cell.value = "합계"
                cell.alignment = CENTER

    # 열 너비 자동 최적화
    for col_idx in range(1, num_cols + 1):
        max_len = len(str(headers[col_idx - 1])) if col_idx - 1 < len(headers) else 5
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    val_len = len(str(cell.value))
                    if val_len > max_len:
                        max_len = val_len
        adjusted = min(max_len * 1.5 + 4, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(adjusted, 8)

    # 첫 행 고정
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_incoming_plan_excel(plan_df) -> bytes:
    """입고 예정 품목만 추려서 깔끔한 엑셀을 생성합니다."""
    import pandas as pd

    cols = ["색상코드", "제조사", "신규"]
    has_note = "비고" in plan_df.columns
    if has_note:
        cols.append("비고")
    incoming = plan_df[plan_df["신규"] > 0][cols].copy()
    col_names = ["품목코드", "제조사", "입고예정수량"]
    if has_note:
        col_names.append("비고")
    incoming.columns = col_names
    incoming = incoming.reset_index(drop=True)

    if incoming.empty:
        raise ValueError("입고 예정 품목이 없습니다.")

    wb = Workbook()
    ws = wb.active
    ws.title = "입고 예정"

    headers = ["No.", "품목코드", "제조사", "입고예정수량"]
    if has_note:
        headers.append("비고")
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for row_idx, (_, row) in enumerate(incoming.iterrows(), 2):
        row_data = [
            (row_idx - 1, CENTER),
            (row["품목코드"], CENTER),
            (row["제조사"], CENTER),
            (int(row["입고예정수량"]), CENTER),
        ]
        if has_note:
            row_data.append((str(row.get("비고", "")), CENTER))
        for col_idx, (val, align) in enumerate(row_data, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = DATA_FONT if col_idx < 4 else Font(name="맑은 고딕", size=10, bold=True)
            c.border = THIN_BORDER
            c.alignment = align
            if col_idx == 4:
                c.number_format = "#,##0"

    # 합계 행
    sum_row = len(incoming) + 2
    sum_fill = PatternFill(start_color="DFE4EA", end_color="DFE4EA", fill_type="solid")
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=sum_row, column=col)
        c.fill = sum_fill
        c.border = THIN_BORDER
        c.font = SUM_FONT
        c.alignment = CENTER
    ws.cell(row=sum_row, column=1, value="합계")
    col_letter = get_column_letter(4)
    ws.cell(row=sum_row, column=4, value=f"=SUM({col_letter}2:{col_letter}{sum_row - 1})")
    ws.cell(row=sum_row, column=4).number_format = "#,##0"

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    if has_note:
        ws.column_dimensions["E"].width = 12
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
